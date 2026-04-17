"""
clear_ticket_data.py
====================
One-time utility to wipe all ticket data rows from ticket_tracker_2601.xlsx
while leaving the header rows (1-8), human-filled columns (G, H, I, K), and
the TOC tab completely untouched.

Steps performed:
  1. Download  /Shared Documents/MJHughes OPEN JOBS/2601/ticket_tracker_2601.xlsx
               from SharePoint to a local temp file.
  2. For every non-TOC tab, clear values and formatting in rows 9+ for the
     AI-written columns only: A, B, C, D, E, F, J.
  3. Re-upload the cleaned workbook to the same SharePoint path, overwriting
     the existing file.

Azure App Registration permissions required (Application, not Delegated):
    - Files.ReadWrite.All
    - Sites.ReadWrite.All

Usage:
    python clear_ticket_data.py
"""

import os
import sys
import logging
import tempfile
from pathlib import Path

import msal
import openpyxl
import requests
from openpyxl.styles import PatternFill
from dotenv import load_dotenv

# ── Configuration ──────────────────────────────────────────────────────────────
_env_path = Path(__file__).resolve().parent / ".env"
load_dotenv(dotenv_path=_env_path)

AZURE_CLIENT_ID     = os.getenv("AZURE_CLIENT_ID")
AZURE_CLIENT_SECRET = os.getenv("AZURE_CLIENT_SECRET")
AZURE_TENANT_ID     = os.getenv("AZURE_TENANT_ID")

SHAREPOINT_HOST   = "vancouvermjhughes.sharepoint.com"
SHAREPOINT_FOLDER = "MJHughes OPEN JOBS"
JOB_NUMBER        = "2601"
EXCEL_FILENAME    = f"ticket_tracker_{JOB_NUMBER}.xlsx"
EXCEL_SP_PATH     = f"{SHAREPOINT_FOLDER}/{JOB_NUMBER}/{EXCEL_FILENAME}"

GRAPH_BASE   = "https://graph.microsoft.com/v1.0"
GRAPH_SCOPES = ["https://graph.microsoft.com/.default"]

# Columns written by the automation (1-based openpyxl indices)
_COL_TICKET_DATE = 1   # A
_COL_LOGGED_DATE = 2   # B
_COL_FACILITY    = 3   # C
_COL_MATERIAL    = 4   # D
_COL_TICKET_NUM  = 5   # E
_COL_QTY_TN      = 6   # F
_COL_NOTES       = 10  # J
_WRITTEN_COLS    = [
    _COL_TICKET_DATE, _COL_LOGGED_DATE, _COL_FACILITY,
    _COL_MATERIAL, _COL_TICKET_NUM, _COL_QTY_TN, _COL_NOTES,
]

_DATA_START_ROW = 9   # rows 1-8 are the pre-built header
_TOC_TAB_NAME   = "TOC"

_NO_FILL = PatternFill(fill_type=None)   # removes any background colour

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler()],
)


# ── Authentication ─────────────────────────────────────────────────────────────
def _get_token() -> str:
    """Acquire an app-only access token from Azure AD."""
    authority = f"https://login.microsoftonline.com/{AZURE_TENANT_ID}"
    app = msal.ConfidentialClientApplication(
        AZURE_CLIENT_ID,
        authority=authority,
        client_credential=AZURE_CLIENT_SECRET,
    )
    result = app.acquire_token_for_client(scopes=GRAPH_SCOPES)
    if "access_token" not in result:
        raise RuntimeError(
            "Authentication failed: "
            + result.get("error_description", result.get("error", "unknown"))
        )
    return result["access_token"]


def _auth_header(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


def _json_headers(token: str) -> dict:
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


# ── SharePoint helpers ─────────────────────────────────────────────────────────
def _get_site_id(token: str) -> str:
    resp = requests.get(
        f"{GRAPH_BASE}/sites/{SHAREPOINT_HOST}:/",
        headers=_auth_header(token),
    )
    resp.raise_for_status()
    return resp.json()["id"]


def _get_drive_id(token: str, site_id: str) -> str:
    """Return the 'Shared Documents' drive ID, falling back to the first drive."""
    resp = requests.get(
        f"{GRAPH_BASE}/sites/{site_id}/drives",
        headers=_auth_header(token),
    )
    resp.raise_for_status()
    drives = resp.json().get("value", [])
    for drive in drives:
        if drive.get("name", "").lower() in ("shared documents", "documents"):
            return drive["id"]
    if not drives:
        raise RuntimeError("No document libraries found on the SharePoint site.")
    logging.warning(
        f"'Shared Documents' drive not found; using '{drives[0]['name']}' as fallback."
    )
    return drives[0]["id"]


def _download_excel(token: str, site_id: str, drive_id: str, dest: Path) -> None:
    """Download EXCEL_SP_PATH from SharePoint and write it to *dest*."""
    url = (
        f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}"
        f"/root:/{EXCEL_SP_PATH}:/content"
    )
    resp = requests.get(url, headers=_auth_header(token))
    resp.raise_for_status()
    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_bytes(resp.content)
    logging.info(
        f"Downloaded {EXCEL_FILENAME} from SharePoint "
        f"({len(resp.content):,} bytes → {dest})"
    )


def _upload_excel(token: str, site_id: str, drive_id: str, src: Path) -> None:
    """Upload *src* to EXCEL_SP_PATH on SharePoint, overwriting the existing file."""
    url = (
        f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}"
        f"/root:/{EXCEL_SP_PATH}:/content"
    )
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    }
    data = src.read_bytes()
    resp = requests.put(url, headers=headers, data=data)
    resp.raise_for_status()
    logging.info(f"Uploaded clean file to SharePoint ({len(data):,} bytes)")


# ── Excel clearing ─────────────────────────────────────────────────────────────
def _clear_tab(sheet) -> None:
    """Clear values and fill colours from all data rows in the written columns."""
    max_row = sheet.max_row
    if max_row < _DATA_START_ROW:
        return  # nothing to clear

    for row_idx in range(_DATA_START_ROW, max_row + 1):
        for col in _WRITTEN_COLS:
            cell = sheet.cell(row=row_idx, column=col)
            cell.value = None
            cell.fill  = _NO_FILL

    logging.info(f"Cleared data from tab [{sheet.title}]")


# ── Entry point ────────────────────────────────────────────────────────────────
def main() -> None:
    # Validate credentials before making any network calls
    missing = [
        v for v in ("AZURE_CLIENT_ID", "AZURE_CLIENT_SECRET", "AZURE_TENANT_ID")
        if not os.getenv(v)
    ]
    if missing:
        logging.error(f"Missing required .env variables: {', '.join(missing)}")
        sys.exit(1)

    logging.info("Authenticating with Microsoft Graph API...")
    token    = _get_token()
    site_id  = _get_site_id(token)
    drive_id = _get_drive_id(token, site_id)
    logging.info("Authenticated.")

    # Download to a temp file so we never touch a local copy by accident
    tmp = Path(tempfile.gettempdir()) / EXCEL_FILENAME
    _download_excel(token, site_id, drive_id, tmp)

    # Clear each non-TOC tab
    workbook = openpyxl.load_workbook(tmp)
    tabs_cleared = 0
    for sheet in workbook.worksheets:
        if sheet.title == _TOC_TAB_NAME:
            logging.info(f"Skipping TOC tab [{sheet.title}]")
            continue
        _clear_tab(sheet)
        tabs_cleared += 1

    workbook.save(tmp)
    logging.info(f"Cleared {tabs_cleared} tab(s). Saving workbook...")

    # Upload cleaned file back to SharePoint
    _upload_excel(token, site_id, drive_id, tmp)

    # Clean up temp file
    tmp.unlink(missing_ok=True)
    logging.info("Done.")


if __name__ == "__main__":
    main()
