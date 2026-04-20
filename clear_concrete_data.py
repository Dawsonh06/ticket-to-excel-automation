"""
clear_concrete_data.py
======================
One-time utility to wipe all ticket data rows from every non-TOC tab in
ticket_tracker_2601_concrete.xlsx, then re-upload the cleaned file.

What it clears (rows 5 and below, columns A–G only):
    A  Ticket Date
    B  Logged Date
    C  Supplier
    D  Slump
    E  Qty Delivered
    F  Ticket #
    G  Helper Notes

What it does NOT touch:
    • Rows 1–4 (title / header rows)
    • Column H (Human Notes — human-filled, never cleared)
    • The TOC tab

SharePoint path:
    /Shared Documents/MJHughes OPEN JOBS/2601/ticket_tracker_2601_concrete.xlsx

Azure App Registration permissions required (Application, not Delegated):
    - Files.ReadWrite.All
    - Sites.ReadWrite.All

Usage:
    python clear_concrete_data.py
"""

import os
import sys
import logging
from pathlib import Path

import msal
import requests
import openpyxl
from openpyxl.styles import PatternFill
from dotenv import load_dotenv

# ── Configuration ──────────────────────────────────────────────────────────────
_env_path = Path(__file__).resolve().parent / ".env"
load_dotenv(dotenv_path=_env_path)

AZURE_CLIENT_ID     = os.getenv("AZURE_CLIENT_ID")
AZURE_CLIENT_SECRET = os.getenv("AZURE_CLIENT_SECRET")
AZURE_TENANT_ID     = os.getenv("AZURE_TENANT_ID")

SHAREPOINT_HOST = "vancouvermjhughes.sharepoint.com"
FILENAME        = "ticket_tracker_2601_concrete.xlsx"
SP_PATH         = f"MJHughes OPEN JOBS/2601/{FILENAME}"
LOCAL_TEMP      = Path(r"C:\Users\dawson.h\AppData\Local\Temp") / FILENAME

TOC_TAB         = "TOC"
DATA_START_ROW  = 5
CLEAR_COLS      = list(range(1, 8))   # columns A–G (1-indexed)
NO_FILL         = PatternFill(fill_type=None)

GRAPH_BASE   = "https://graph.microsoft.com/v1.0"
GRAPH_SCOPES = ["https://graph.microsoft.com/.default"]

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler()],
)


# ── Authentication ─────────────────────────────────────────────────────────────
def _get_token() -> str:
    authority = f"https://login.microsoftonline.com/{AZURE_TENANT_ID}"
    app = msal.ConfidentialClientApplication(
        AZURE_CLIENT_ID,
        authority=authority,
        client_credential=AZURE_CLIENT_SECRET,
    )
    result = app.acquire_token_for_client(scopes=GRAPH_SCOPES)
    if "access_token" not in result:
        raise RuntimeError(
            "Authentication failed: "
            + result.get("error_description", result.get("error", "unknown"))
        )
    return result["access_token"]


def _auth_header(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


# ── SharePoint helpers ─────────────────────────────────────────────────────────
def _get_site_id(token: str) -> str:
    resp = requests.get(
        f"{GRAPH_BASE}/sites/{SHAREPOINT_HOST}:/",
        headers=_auth_header(token),
    )
    resp.raise_for_status()
    return resp.json()["id"]


def _get_drive_id(token: str, site_id: str) -> str:
    resp = requests.get(
        f"{GRAPH_BASE}/sites/{site_id}/drives",
        headers=_auth_header(token),
    )
    resp.raise_for_status()
    drives = resp.json().get("value", [])
    for drive in drives:
        if drive.get("name", "").lower() in ("shared documents", "documents"):
            return drive["id"]
    if not drives:
        raise RuntimeError("No document libraries found on the SharePoint site.")
    logging.warning(
        f"'Shared Documents' drive not found; using '{drives[0]['name']}' as fallback."
    )
    return drives[0]["id"]


def _download(token: str, site_id: str, drive_id: str) -> None:
    url = (
        f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}"
        f"/root:/{SP_PATH}:/content"
    )
    resp = requests.get(url, headers=_auth_header(token))
    resp.raise_for_status()
    LOCAL_TEMP.parent.mkdir(parents=True, exist_ok=True)
    LOCAL_TEMP.write_bytes(resp.content)
    logging.info(
        f"Downloaded {FILENAME} from SharePoint "
        f"({len(resp.content):,} bytes → {LOCAL_TEMP})"
    )


def _upload(token: str, site_id: str, drive_id: str) -> None:
    data = LOCAL_TEMP.read_bytes()
    url = (
        f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}"
        f"/root:/{SP_PATH}:/content"
    )
    resp = requests.put(
        url,
        headers={
            **_auth_header(token),
            "Content-Type": (
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            ),
        },
        data=data,
    )
    resp.raise_for_status()
    logging.info(f"Uploaded clean file to SharePoint ({len(data):,} bytes)")


# ── Clear logic ────────────────────────────────────────────────────────────────
def clear_data() -> None:
    """Open the local workbook and wipe data rows from every non-TOC tab."""
    wb = openpyxl.load_workbook(LOCAL_TEMP)

    data_tabs = [name for name in wb.sheetnames if name != TOC_TAB]
    if not data_tabs:
        logging.warning("No data tabs found (only TOC or empty workbook). Nothing to clear.")
        wb.close()
        return

    for tab_name in data_tabs:
        ws = wb[tab_name]

        # Find the actual last row that has any content in A–G so we don't
        # iterate thousands of empty rows in a sparse sheet.
        last_row = DATA_START_ROW - 1
        for row in ws.iter_rows(min_row=DATA_START_ROW, min_col=1, max_col=7):
            if any(cell.value is not None for cell in row):
                last_row = row[0].row

        if last_row < DATA_START_ROW:
            logging.info(f"  Tab '{tab_name}': no data rows to clear — skipping.")
            continue

        for row_idx in range(DATA_START_ROW, last_row + 1):
            for col_idx in CLEAR_COLS:
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = None
                cell.fill  = NO_FILL

        logging.info(f"  Cleared data from tab '{tab_name}' (rows {DATA_START_ROW}–{last_row}, cols A–G)")

    wb.save(LOCAL_TEMP)
    logging.info(f"Workbook saved locally: {LOCAL_TEMP}")


# ── Entry point ────────────────────────────────────────────────────────────────
def main() -> None:
    missing = [
        v for v in ("AZURE_CLIENT_ID", "AZURE_CLIENT_SECRET", "AZURE_TENANT_ID")
        if not os.getenv(v)
    ]
    if missing:
        logging.error(f"Missing required .env variables: {', '.join(missing)}")
        sys.exit(1)

    logging.info("Authenticating with Microsoft Graph API...")
    token    = _get_token()
    site_id  = _get_site_id(token)
    drive_id = _get_drive_id(token, site_id)
    logging.info("Authenticated.")

    _download(token, site_id, drive_id)
    clear_data()
    _upload(token, site_id, drive_id)

    LOCAL_TEMP.unlink()
    logging.info(f"Deleted local temp copy: {LOCAL_TEMP}")


if __name__ == "__main__":
    main()
