"""
Ticket Processing Automation
==============================
Connects to the help@mjhughes.com Exchange mailbox via Microsoft Graph API,
processes PDF scanned tickets, extracts QR codes and OCR data, logs entries
to an Excel workbook organised by cost code, uploads PDFs to SharePoint,
and marks emails as read.

Azure App Registration permissions required (Application, not Delegated):
    - Mail.Read
    - Mail.ReadWrite
    - Mail.Send
    - Files.ReadWrite.All
    - Sites.ReadWrite.All

External dependencies (see requirements.txt):
    pip install -r requirements.txt

External tools required:
    - Tesseract OCR  → https://github.com/UB-Mannheim/tesseract/wiki
      Windows: install to C:\\Program Files\\Tesseract-OCR\\tesseract.exe
      and set TESSERACT_CMD in your .env if not on PATH.
    - opencv-python (pip install opencv-python; no external DLLs required)

.env file variables:
    AZURE_CLIENT_ID     = <your Azure app client ID>
    AZURE_CLIENT_SECRET = <your Azure app client secret>
    AZURE_TENANT_ID     = <your Azure tenant ID>
    TESSERACT_CMD       = (optional) full path to tesseract.exe on Windows
    OCR_DEBUG           = (optional) set to 1 to save ocr_debug.txt and ocr_debug.png
"""

# ============================================================
# IMPORTS
# ============================================================
import io
import os
import platform
import re
import json
import base64
import difflib
import logging
import statistics
import traceback
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

import anthropic
import msal
import requests
from dotenv import load_dotenv
import fitz                          # PyMuPDF – PDF → image conversion
from PIL import Image, ImageDraw, ImageEnhance
import pytesseract                   # OCR engine wrapper
import cv2                             # image utility operations
import numpy as np
try:
    from pyzbar.pyzbar import decode as pyzbar_decode
    from pyzbar.pyzbar import ZBarSymbol
    PYZBAR_AVAILABLE = True
    logging.info("pyzbar loaded successfully")
except Exception:
    PYZBAR_AVAILABLE = False
    logging.info("pyzbar not available - using AI vision for QR detection")

try:
    import zxingcpp
    ZXINGCPP_AVAILABLE = True
except Exception:
    ZXINGCPP_AVAILABLE = False
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURATION  (loaded from .env)
# ============================================================
# Always resolve .env relative to this script file, not the current working
# directory.  Without an explicit path, load_dotenv() only searches cwd, so
# running the script from any other directory silently skips the .env file.
_env_path = Path(__file__).resolve().parent / ".env"
_dotenv_loaded = load_dotenv(dotenv_path=_env_path)

# ── .env diagnostic (always prints so key-loading failures are visible) ──────
print(f"[.env] path      : {_env_path}")
print(f"[.env] file exists: {_env_path.exists()}")
print(f"[.env] load_dotenv returned: {_dotenv_loaded}")
_env_keys = [k for k in os.environ if not k.startswith("_")]
print(f"[.env] env keys loaded ({len(_env_keys)} total): {sorted(_env_keys)}")
print(f"[debug] ANTHROPIC_API_KEY raw value length: {len(os.getenv('ANTHROPIC_API_KEY', ''))}")
_api_key_raw = (os.environ.get("ANTHROPIC_API_KEY") or os.getenv("ANTHROPIC_API_KEY") or "").strip()
print(f"[.env] ANTHROPIC_API_KEY found: {bool(_api_key_raw)}")
if _api_key_raw:
    print(f"[.env] ANTHROPIC_API_KEY first 10 chars: {_api_key_raw[:10]}...")
# ─────────────────────────────────────────────────────────────────────────────

AZURE_CLIENT_ID     = os.getenv("AZURE_CLIENT_ID")
AZURE_CLIENT_SECRET = os.getenv("AZURE_CLIENT_SECRET")
AZURE_TENANT_ID     = os.getenv("AZURE_TENANT_ID")
ANTHROPIC_API_KEY   = (os.environ.get("ANTHROPIC_API_KEY") or os.getenv("ANTHROPIC_API_KEY") or "").strip()

# Allow overriding Tesseract path via .env.
# Default: Windows install path on Windows, plain 'tesseract' on Linux/macOS.
_tesseract_default = (
    r"C:\Program Files\Tesseract-OCR\tesseract.exe"
    if platform.system() == "Windows"
    else "tesseract"
)
_tesseract_cmd = os.getenv("TESSERACT_CMD", _tesseract_default)
pytesseract.pytesseract.tesseract_cmd = _tesseract_cmd

MAILBOX           = os.getenv("MAILBOX_ADDRESS",       "help@mjhughes.com")
SUMMARY_RECIPIENT = os.getenv("SUMMARY_EMAIL",         "dawson.h@mjhughes.com")
_JOB_NUMBER       = os.getenv("JOB_NUMBER",           "2601")          # matches QR codes
_SP_JOB_FOLDER    = os.getenv("SHAREPOINT_JOB_FOLDER", _JOB_NUMBER)    # actual SharePoint folder name

_EXCEL_FILENAME          = os.getenv("EXCEL_FILE_ROCK",     f"ticket_tracker_{_JOB_NUMBER}.xlsx")
_EXCEL_FILENAME_CONCRETE = os.getenv("EXCEL_FILE_CONCRETE", f"ticket_tracker_{_JOB_NUMBER}_concrete.xlsx")

# Local temp paths are derived from the OS temp directory and the (possibly
# overridden) filenames above.  No user-specific path is hardcoded.
_LOCAL_TEMP_DIR     = Path(os.environ.get("TEMP", r"C:\Windows\Temp"))
EXCEL_FILE          = str(_LOCAL_TEMP_DIR / _EXCEL_FILENAME)
EXCEL_FILE_CONCRETE = str(_LOCAL_TEMP_DIR / _EXCEL_FILENAME_CONCRETE)

ERROR_LOG              = "error_log.txt"
KNOWN_FACILITIES_FILE  = "known_facilities.txt"
TICKET_PROFILES_DIR    = "ticket_profiles"
UNKNOWN_SUPPLIERS_LOG  = "unknown_suppliers.txt"
OCR_DEBUG          = os.getenv("OCR_DEBUG", "").lower() in ("1", "true", "yes")
TICKET_TYPES_FILE  = "ticket_types.json"
# "all" or unset → no type filter; "rock" or "concrete" → only that type
_raw_ticket_type   = os.getenv("TICKET_TYPE", "all").lower().strip()
_TICKET_TYPE       = "" if _raw_ticket_type in ("all", "") else _raw_ticket_type
SHAREPOINT_HOST    = "vancouvermjhughes.sharepoint.com"
SHAREPOINT_FOLDER  = "MJHughes OPEN JOBS"   # Top-level folder inside Shared Documents
# Per-job folder layout:
#   /MJHughes OPEN JOBS/{job_number}/                        ← job root
#   /MJHughes OPEN JOBS/{job_number}/Ticket Scans/           ← PDFs go here
#   /MJHughes OPEN JOBS/{job_number}/ticket_tracker_NNN.xlsx ← Excel workbook

GRAPH_BASE         = "https://graph.microsoft.com/v1.0"
GRAPH_SCOPES       = ["https://graph.microsoft.com/.default"]

# OCR corrections are now defined per-supplier inside each ticket profile JSON file.
# See ticket_profiles/ directory.  The global dict below is kept for reference only
# and is no longer applied during processing.
# OCR_CORRECTIONS = {"378-0": "3/4-0", "3)4-0": "3/4-0", "3)4": "3/4", ...}


# ============================================================
# TICKET PROFILE SYSTEM
# ============================================================
@dataclass
class TicketProfile:
    """Supplier-specific ticket layout and extraction configuration.

    Loaded from a JSON file in the ticket_profiles/ directory.
    Each supplier gets one file; no code changes are needed to add a supplier.
    """
    supplier_name:                str
    detection_keywords:           list
    layout:                       dict
    labels:                       dict
    weight_table:                 dict
    ocr_corrections:              dict
    confidence_checks:            list
    source_file:                  str            = ""
    detection_context_keywords:   Optional[dict] = None
    is_fallback:                  bool           = False
    ticket_type:                  str            = ""
    detection_scan_lines:         int            = 3


_REQUIRED_PROFILE_KEYS: dict[str, type] = {
    "supplier_name":      str,
    "detection_keywords": list,
    "layout":             dict,
    "labels":             dict,
    "weight_table":       dict,
    "ocr_corrections":    dict,
    "confidence_checks":  list,
}


def validate_profile(data: dict, filepath: str) -> list[str]:
    """Return a list of validation error strings for a raw profile dict.

    Returns an empty list when the profile is valid.  Called during startup
    so invalid profiles are reported before any emails are processed.
    """
    errors: list[str] = []
    for key, expected_type in _REQUIRED_PROFILE_KEYS.items():
        if key not in data:
            errors.append(f"missing required key '{key}'")
        elif not isinstance(data[key], expected_type):
            errors.append(
                f"'{key}' must be {expected_type.__name__}, "
                f"got {type(data[key]).__name__}"
            )
    if isinstance(data.get("detection_keywords"), list) and not data["detection_keywords"]:
        if not data.get("is_fallback"):
            errors.append("'detection_keywords' must not be empty")
    if isinstance(data.get("confidence_checks"), list) and not data["confidence_checks"]:
        errors.append("'confidence_checks' must not be empty")
    return errors


def load_ticket_profiles() -> list[TicketProfile]:
    """Load and validate all *.json profile files from the ticket_profiles/ directory.

    Invalid profiles are logged to error_log.txt and skipped.
    Returns a list of valid TicketProfile objects sorted by filename.
    """
    profiles_dir = Path(TICKET_PROFILES_DIR)
    if not profiles_dir.is_dir():
        logging.warning(
            f"Ticket profiles directory not found: {profiles_dir.resolve()!s}. "
            "All tickets will be flagged for manual review (unknown supplier)."
        )
        return []

    profiles: list[TicketProfile] = []
    for json_path in sorted(profiles_dir.glob("*.json")):
        try:
            data = json.loads(json_path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError) as exc:
            msg = f"Profile load error — {json_path.name}: {exc}"
            logging.error(msg)
            log_error("profile_load", msg)
            continue

        errors = validate_profile(data, str(json_path))
        if errors:
            for err in errors:
                msg = f"Profile validation error — {json_path.name}: {err}"
                logging.error(msg)
                log_error("profile_validation", msg)
            continue

        profile = TicketProfile(
            supplier_name               = data["supplier_name"],
            detection_keywords          = data["detection_keywords"],
            layout                      = data.get("layout", {}),
            labels                      = data.get("labels", {}),
            weight_table                = data.get("weight_table", {}),
            ocr_corrections             = data.get("ocr_corrections", {}),
            confidence_checks           = data["confidence_checks"],
            source_file                 = str(json_path),
            detection_context_keywords  = data.get("detection_context_keywords"),
            is_fallback                 = bool(data.get("is_fallback", False)),
            ticket_type                 = data.get("ticket_type", ""),
            detection_scan_lines        = int(data.get("detection_scan_lines", 3)),
        )
        profiles.append(profile)
        logging.info(f"Loaded ticket profile: {profile.supplier_name!r} ({json_path.name})")

    if not profiles:
        logging.warning(
            f"No valid profiles loaded from {profiles_dir.resolve()!s}. "
            "All tickets will be flagged (unknown supplier)."
        )
    return profiles


def load_ticket_types() -> dict:
    """Load ticket type definitions from ticket_types.json.

    Returns a dict mapping type name → config dict with keys:
        keywords, target_folder, excel_file.
    Returns an empty dict if the file is missing or invalid.
    """
    path = Path(TICKET_TYPES_FILE)
    if not path.exists():
        logging.warning(
            f"Ticket types file not found: {path.resolve()!s}. "
            "Ticket type detection will be skipped."
        )
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError) as exc:
        logging.error(f"Failed to load ticket types: {exc}")
        return {}
    if not isinstance(data, dict):
        logging.error(f"ticket_types.json must be a JSON object, got {type(data).__name__}")
        return {}
    logging.info(f"Loaded {len(data)} ticket type(s): {list(data.keys())}")
    return data


def detect_ticket_type(ocr_text: str, ticket_types: dict) -> "tuple[str, int]":
    """Score OCR text against each ticket type's keyword list.

    Each keyword that appears (case-insensitive substring match) in the text
    adds one point to that type's score.

    Returns:
        (type_name, score) for the highest-scoring type, or ("unknown", 0)
        when no keywords match or multiple types tie for the top score.
    """
    if not ticket_types:
        return ("unknown", 0)

    text_lower = ocr_text.lower()
    scores: dict[str, int] = {}
    for type_name, config in ticket_types.items():
        keywords = config.get("keywords", [])
        scores[type_name] = sum(1 for kw in keywords if kw.lower() in text_lower)

    max_score = max(scores.values(), default=0)
    if max_score == 0:
        return ("unknown", 0)

    winners = [t for t, s in scores.items() if s == max_score]
    if len(winners) != 1:
        return ("unknown", 0)  # tie

    return (winners[0], max_score)


def _keyword_matches(keyword: str, text: str, cutoff: float = 0.80) -> bool:
    """Return True if keyword appears in text exactly or via fuzzy n-gram match.

    Exact substring check is tried first (handles clean OCR output).
    Falls back to a sliding n-gram fuzzy comparison at `cutoff` SequenceMatcher
    ratio to tolerate single-character OCR noise (e.g., "Teevin" vs "Teevim").
    """
    kw_lower   = keyword.lower().strip()
    text_lower = text.lower()

    if kw_lower in text_lower:
        return True

    kw_words   = kw_lower.split()
    text_words = text_lower.split()
    n = len(kw_words)

    for i in range(max(len(text_words) - n + 1, 1)):
        chunk = " ".join(text_words[i : i + n])
        if difflib.SequenceMatcher(None, kw_lower, chunk).ratio() >= cutoff:
            return True

    return False


def _identify_profile_with_ai(
    image: "Image.Image",
    candidates: list[TicketProfile],
) -> "Optional[TicketProfile]":
    """Ask Claude AI to identify which supplier profile matches the ticket image.

    Called when OCR keyword detection fails.  Sends the image to Claude with a
    list of candidate supplier names and asks it to pick one.  Returns the
    matching TicketProfile or None.
    """
    if not ANTHROPIC_API_KEY:
        return None

    names_list = "\n".join(f"- {p.supplier_name}" for p in candidates)
    prompt = (
        "Look at this delivery ticket image. Which supplier issued this ticket? "
        "The possible suppliers are:\n"
        f"{names_list}\n\n"
        "Reply with ONLY the exact supplier name from the list above. "
        "If you cannot identify the supplier from the list, reply with 'unknown'."
    )
    try:
        import base64, io as _io
        _client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        buf = _io.BytesIO()
        image.save(buf, format="PNG")
        b64 = base64.standard_b64encode(buf.getvalue()).decode()
        response = _client.messages.create(
            model=_AI_MODEL,
            max_tokens=64,
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {"type": "base64", "media_type": "image/png", "data": b64},
                    },
                    {"type": "text", "text": prompt},
                ],
            }],
        )
        identified = response.content[0].text.strip()
        logging.info(f"  AI profile detection: {identified!r}")
        if identified.lower() == "unknown":
            return None
        for p in candidates:
            if p.supplier_name.lower() == identified.lower():
                return p
        for p in candidates:
            if identified.lower() in p.supplier_name.lower():
                return p
        logging.warning(f"  AI profile detection: no profile matched name {identified!r}")
    except Exception as exc:
        logging.warning(f"  AI profile detection failed: {exc}")
    return None


def detect_profile(
    full_text: str, profiles: list[TicketProfile],
    image: "Optional[Image.Image]" = None,
) -> Optional[TicketProfile]:
    """Identify which supplier profile matches the OCR text.

    Primary check: scans the first three non-empty lines (where the company
    name appears) and tests each profile's detection_keywords.

    Context-dependent check: if a profile defines detection_context_keywords,
    secondary keywords (e.g. "CUSTOMER COPY") are also tested against the
    header — but only accepted when a required context string (e.g.
    "KNIFE RIVER") is also found within the first N lines of the text.
    This prevents false-positive matches on headers that share common words
    with other ticket formats.

    Returns the first matching profile, or None if no match is found.
    """
    all_lines: list[str] = [l.strip() for l in full_text.splitlines() if l.strip()]
    default_header_text = " ".join(all_lines[:3])

    fallback: Optional[TicketProfile] = None

    for profile in profiles:
        # Fallback profiles match only when no specific profile does — skip
        # them in the main loop and remember the first one found.
        if profile.is_fallback:
            if fallback is None:
                fallback = profile
            continue

        # ── Primary keywords ───────────────────────────────────────────────────
        # Use detection_scan_lines from the profile (default 3) so suppliers
        # whose company name appears lower on the ticket can still be detected.
        scan_n = profile.detection_scan_lines
        header_text = (
            " ".join(all_lines[:scan_n]) if scan_n != 3 else default_header_text
        )
        for keyword in profile.detection_keywords:
            if _keyword_matches(keyword, header_text):
                return profile

        # ── Context-dependent secondary keywords ───────────────────────────────
        ctx = profile.detection_context_keywords
        if ctx:
            ctx_keywords  = ctx.get("keywords", [])
            required_ctx  = ctx.get("required_context", [])
            ctx_lines_n   = ctx.get("context_lines", 10)
            wide_text     = " ".join(all_lines[:ctx_lines_n])

            for keyword in ctx_keywords:
                if _keyword_matches(keyword, header_text):
                    if any(_keyword_matches(req, wide_text) for req in required_ctx):
                        logging.info(
                            f"  Profile {profile.supplier_name!r} matched via "
                            f"context keyword {keyword!r} (required context "
                            f"confirmed in first {ctx_lines_n} lines)"
                        )
                        return profile

    # No specific profile matched via keywords — try AI vision before falling back.
    if image is not None and fallback is not None:
        specific = [p for p in profiles if not p.is_fallback]
        if specific:
            ai_match = _identify_profile_with_ai(image, specific)
            if ai_match:
                logging.info(
                    f"  Profile detected via AI vision: {ai_match.supplier_name!r}"
                )
                return ai_match

    return fallback


def _log_unknown_supplier(company_name: str) -> None:
    """Append an unknown supplier entry to unknown_suppliers.txt."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with open(UNKNOWN_SUPPLIERS_LOG, "a", encoding="utf-8") as fh:
        fh.write(f"[{timestamp}] {company_name}\n")
    logging.info(
        f"Unknown supplier logged to {UNKNOWN_SUPPLIERS_LOG}: {company_name!r}"
    )

# QR code pattern:  XXXX  -  YYYY  -  CCCCCC  -  CC
#   XXXX   = job number (4 alphanumeric chars)
#   YYYY   = location   (4 alphanumeric chars)
#   CCCCCC = cost code part 1 (6 alphanumeric chars)
#   CC     = cost code part 2 (2 alphanumeric chars)
QR_PATTERN = re.compile(
    r'\b([A-Z0-9]{4})-([A-Z0-9]{4})-([A-Z0-9]{6})-([A-Z0-9]{2})\b',
    re.IGNORECASE,
)

# ── Excel column mapping for ticket_tracker_2601.xlsx ───────────────────────
# Columns A-F and J are written by this script; G, H, I, K are human-filled.
_COL_TICKET_DATE = 1   # A — date from the ticket
_COL_LOGGED_DATE = 2   # B — date this script ran
_COL_FACILITY    = 3   # C — quarry / source facility
_COL_MATERIAL    = 4   # D — material / product description
_COL_TICKET_NUM  = 5   # E — 5-digit ticket number
_COL_QTY_TN      = 6   # F — net quantity in tons
_COL_NOTES       = 10  # J — helper notes (flags, outlier warnings, etc.)
_DATA_START_ROW  = 9   # rows 1-8 are the pre-built header; data starts here
_WRITTEN_COLS    = [_COL_TICKET_DATE, _COL_LOGGED_DATE, _COL_FACILITY,
                    _COL_MATERIAL, _COL_TICKET_NUM, _COL_QTY_TN, _COL_NOTES]

# ── Concrete Excel column mapping (ticket_tracker_2601_concrete.xlsx) ────────
# Data starts at row 5; columns A–G written; H is left blank for human notes.
_CONCRETE_COL_TICKET_DATE = 1   # A — ticket date
_CONCRETE_COL_LOGGED_DATE = 2   # B — script run date
_CONCRETE_COL_SUPPLIER    = 3   # C — supplier name (e.g. "CalPortland")
_CONCRETE_COL_SLUMP       = 4   # D — slump on arrival
_CONCRETE_COL_QTY_CY      = 5   # E — qty delivered (CY)
_CONCRETE_COL_TICKET_NUM  = 6   # F — 7-digit ticket number
_CONCRETE_COL_NOTES       = 7   # G — helper notes (material description)
_CONCRETE_WRITTEN_COLS    = [1, 2, 3, 4, 5, 6, 7]
_CONCRETE_DATA_START_ROW  = 5

_TOC_TAB_NAME = "TOC"   # never read or written to; skipped in all tab scans


# ============================================================
# LOGGING SETUP
# ============================================================
def configure_logging() -> None:
    """Configure logging to both the console and a rolling log file."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[
            logging.StreamHandler(),
            logging.FileHandler("processor.log", encoding="utf-8"),
        ],
    )


def log_error(subject: str, error_detail: str) -> None:
    """
    Append a structured error entry to error_log.txt.
    Called when any processing step fails; the email is intentionally
    left unread so the next run will retry it.
    """
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    separator = "-" * 60
    entry = f"[{timestamp}]\nSubject : {subject}\nError   : {error_detail}\n{separator}\n"

    with open(ERROR_LOG, "a", encoding="utf-8") as fh:
        fh.write(entry)

    logging.error(f"Error logged → '{subject}': {error_detail}")


# ============================================================
# MICROSOFT GRAPH API CLIENT
# ============================================================
class GraphClient:
    """
    Thin wrapper around the Microsoft Graph REST API.

    Uses MSAL's ConfidentialClientApplication with the client-credentials
    flow (app-only auth), which is appropriate for background services that
    operate without a signed-in user.
    """

    def __init__(self) -> None:
        self.access_token: str = ""
        self._authenticate()

    def _authenticate(self) -> None:
        """Acquire an access token from Azure AD using client credentials."""
        authority = f"https://login.microsoftonline.com/{AZURE_TENANT_ID}"

        app = msal.ConfidentialClientApplication(
            AZURE_CLIENT_ID,
            authority=authority,
            client_credential=AZURE_CLIENT_SECRET,
        )

        result = app.acquire_token_for_client(scopes=GRAPH_SCOPES)

        if "access_token" not in result:
            raise RuntimeError(
                "Graph API authentication failed: "
                + result.get("error_description", result.get("error", "unknown error"))
            )

        self.access_token = result["access_token"]
        logging.info("Authenticated with Microsoft Graph API.")

    # ----------------------------------------------------------
    # Helper: standard JSON request headers
    # ----------------------------------------------------------
    def _json_headers(self) -> dict:
        return {
            "Authorization":   f"Bearer {self.access_token}",
            "Content-Type":    "application/json",
            "ConsistencyLevel": "eventual",   # required by some SharePoint endpoints
        }

    # ----------------------------------------------------------
    # Generic HTTP verbs
    # ----------------------------------------------------------
    def get(self, url: str, **kwargs) -> requests.Response:
        resp = requests.get(url, headers=self._json_headers(), **kwargs)
        resp.raise_for_status()
        return resp

    def post(self, url: str, **kwargs) -> requests.Response:
        resp = requests.post(url, headers=self._json_headers(), **kwargs)
        resp.raise_for_status()
        return resp

    def patch(self, url: str, **kwargs) -> requests.Response:
        resp = requests.patch(url, headers=self._json_headers(), **kwargs)
        resp.raise_for_status()
        return resp

    def put_binary(self, url: str, data: bytes, content_type: str) -> requests.Response:
        """PUT raw bytes (used for SharePoint file upload)."""
        headers = {
            "Authorization": f"Bearer {self.access_token}",
            "Content-Type": content_type,
        }
        resp = requests.put(url, headers=headers, data=data)
        resp.raise_for_status()
        return resp


# ============================================================
# EMAIL OPERATIONS
# ============================================================
def get_unread_emails_with_pdf(client: GraphClient) -> list:
    """
    Retrieve all unread messages from the mailbox that have at least one
    PDF attachment.  Handles OData pagination automatically.

    Graph API filter: isRead eq false AND hasAttachments eq true
    We then inspect each message's attachment list to confirm PDF presence.
    """
    url = (
        f"{GRAPH_BASE}/users/{MAILBOX}/mailFolders/inbox/messages"
        "?$filter=isRead eq false and hasAttachments eq true"
        "&$select=id,subject,from,receivedDateTime,hasAttachments"
        "&$top=50"
    )

    emails_with_pdfs = []

    # Walk through pages until no @odata.nextLink is returned
    while url:
        response = client.get(url)
        data = response.json()

        for message in data.get("value", []):
            pdf_attachments = _get_pdf_attachments(client, message["id"])
            if pdf_attachments:
                message["pdf_attachments"] = pdf_attachments
                emails_with_pdfs.append(message)

        url = data.get("@odata.nextLink")  # None when last page reached

    logging.info(f"Found {len(emails_with_pdfs)} unread email(s) with PDF attachments.")
    return emails_with_pdfs


def _get_pdf_attachments(client: GraphClient, message_id: str) -> list:
    """
    Fetch attachment metadata for a message and return only PDF items.
    Metadata only (no content bytes) to keep this call lightweight.
    """
    url = (
        f"{GRAPH_BASE}/users/{MAILBOX}/messages/{message_id}/attachments"
        "?$select=id,name,contentType,size"
    )
    attachments = client.get(url).json().get("value", [])

    return [
        att for att in attachments
        if att.get("contentType", "").lower() in ("application/pdf", "application/octet-stream")
        or att.get("name", "").lower().endswith(".pdf")
    ]


def get_attachment_content(client: GraphClient, message_id: str, attachment_id: str) -> bytes:
    """
    Download the binary content of a single attachment.
    Graph API returns it base64-encoded inside a JSON envelope.
    """
    url = (
        f"{GRAPH_BASE}/users/{MAILBOX}/messages/{message_id}"
        f"/attachments/{attachment_id}"
    )
    data = client.get(url).json()

    # Decode from base64 string to raw bytes
    return base64.b64decode(data["contentBytes"])


def mark_email_as_read(client: GraphClient, message_id: str) -> None:
    """
    Mark an email as read.
    Only called after ALL processing steps succeed so that failed
    emails remain unread and are retried on the next run.
    """
    url = f"{GRAPH_BASE}/users/{MAILBOX}/messages/{message_id}"
    client.patch(url, json={"isRead": True})
    logging.info(f"Marked email {message_id} as read.")


def mark_email_as_unread(client: GraphClient, message_id: str) -> None:
    """
    Mark an email as unread.
    Called after moving a low-confidence ticket to the REVIEW REQUIRED folder
    so the message stands out as requiring attention.
    """
    url = f"{GRAPH_BASE}/users/{MAILBOX}/messages/{message_id}"
    client.patch(url, json={"isRead": False})
    logging.info(f"Marked email {message_id} as unread (flagged for review).")


def rename_email_subject(
    client: GraphClient, message_id: str, new_subject: str, current_subject: str = ""
) -> None:
    """Update the subject line of an email via a Graph API PATCH request.

    Must be called using the current message ID for the email's location.
    After a move operation Exchange assigns a new ID, so always rename
    before moving to avoid 404 errors.

    If current_subject is provided and new_subject would produce a duplicate
    prefix (e.g. "REVIEW REQUIRED - REVIEW REQUIRED - …") the rename is
    skipped and a warning is logged instead.
    """
    # Guard against double-prefixing when the email was already renamed on a
    # previous run (e.g. the move succeeded but mark-as-unread failed and the
    # email was picked up again).
    check = current_subject or new_subject
    for prefix in ("REVIEW REQUIRED", "DUPLICATE"):
        if check.startswith(prefix) and new_subject.startswith(prefix):
            logging.warning(
                f"  Skipping rename — subject already starts with {prefix!r}: "
                f"{current_subject!r}"
            )
            return

    url = f"{GRAPH_BASE}/users/{MAILBOX}/messages/{message_id}"
    client.patch(url, json={"subject": new_subject})
    logging.info(f"Renamed email {message_id} subject to: {new_subject!r}")


def flag_email_category(client: GraphClient, message_id: str) -> None:
    """Add the 'REVIEW REQUIRED' category to an email in Exchange.

    The category will appear in Outlook in the colour configured for that
    category name in Exchange.  If the category has not been set up it will
    show as the default (no colour) but the label is still searchable.
    """
    url = f"{GRAPH_BASE}/users/{MAILBOX}/messages/{message_id}"
    client.patch(url, json={"categories": ["REVIEW REQUIRED"]})
    logging.info(f"Flagged email {message_id} with category 'REVIEW REQUIRED'.")


def _processed_folder_name(ticket_type: str) -> str:
    """Return the processed-mail folder name for *ticket_type*.

    Examples:
        "rock"     → "Rock Tickets - Processed"
        "concrete" → "Concrete Tickets - Processed"
        ""         → "Processed"  (fallback)
    """
    t = (ticket_type or "").lower().strip()
    return f"{t.title()} Tickets - Processed" if t else "Processed"


def _review_folder_name(ticket_type: str) -> str:
    """Return the review-required folder name for *ticket_type*.

    Examples:
        "rock"     → "Rock Tickets - Review Required"
        "concrete" → "Concrete Tickets - Review Required"
        "unknown"  → "Unknown Tickets - Review Required"
        ""         → "Unknown Tickets - Review Required"
    """
    t = (ticket_type or "unknown").lower().strip()
    return f"{t.title()} Tickets - Review Required"


def get_or_create_review_folder(client: GraphClient, ticket_type: str = "") -> str:
    """Return the Exchange folder ID for the review folder for *ticket_type*.

    Folder name is determined by :func:`_review_folder_name`.
    Creates the folder if it does not already exist.

    Returns:
        The Graph API folder ID string.
    """
    folder_name = _review_folder_name(ticket_type)
    list_url = (
        f"{GRAPH_BASE}/users/{MAILBOX}/mailFolders"
        f"?$filter=displayName eq '{folder_name}'&$select=id,displayName"
    )
    data    = client.get(list_url).json()
    folders = data.get("value", [])

    if folders:
        folder_id = folders[0]["id"]
        logging.debug(f"Found existing mail folder '{folder_name}' (id={folder_id}).")
        return folder_id

    create_url = f"{GRAPH_BASE}/users/{MAILBOX}/mailFolders"
    created    = client.post(create_url, json={"displayName": folder_name}).json()
    folder_id  = created["id"]
    logging.info(f"Created mail folder '{folder_name}' (id={folder_id}).")
    return folder_id


def move_email_to_review_folder(
    client: GraphClient, message_id: str, ticket_type: str = ""
) -> str:
    """Move an email to the type-specific review folder, creating it if needed.

    Returns:
        The new message ID assigned by Exchange after the move.
    """
    folder_name = _review_folder_name(ticket_type)
    folder_id   = get_or_create_review_folder(client, ticket_type)
    url         = f"{GRAPH_BASE}/users/{MAILBOX}/messages/{message_id}/move"
    moved       = client.post(url, json={"destinationId": folder_id}).json()
    new_id      = moved["id"]
    logging.info(f"Moved email {message_id} to '{folder_name}' folder (new id={new_id}).")
    return new_id


def get_or_create_processed_folder(client: GraphClient, ticket_type: str = "") -> str:
    """Return the Exchange folder ID for the processed folder for *ticket_type*.

    Folder name is determined by :func:`_processed_folder_name`.
    Creates the folder if it does not already exist.
    """
    folder_name = _processed_folder_name(ticket_type)
    list_url = (
        f"{GRAPH_BASE}/users/{MAILBOX}/mailFolders"
        f"?$filter=displayName eq '{folder_name}'&$select=id,displayName"
    )
    data    = client.get(list_url).json()
    folders = data.get("value", [])

    if folders:
        folder_id = folders[0]["id"]
        logging.debug(f"Found existing mail folder '{folder_name}' (id={folder_id}).")
        return folder_id

    create_url = f"{GRAPH_BASE}/users/{MAILBOX}/mailFolders"
    created    = client.post(create_url, json={"displayName": folder_name}).json()
    folder_id  = created["id"]
    logging.info(f"Created mail folder '{folder_name}' (id={folder_id}).")
    return folder_id


def move_email_to_processed_folder(
    client: GraphClient, message_id: str, subject: str, ticket_type: str = ""
) -> None:
    """Move a fully-processed email to the type-specific processed folder.

    Creates the folder if it does not already exist.
    """
    folder_name = _processed_folder_name(ticket_type)
    folder_id   = get_or_create_processed_folder(client, ticket_type)
    url         = f"{GRAPH_BASE}/users/{MAILBOX}/messages/{message_id}/move"
    client.post(url, json={"destinationId": folder_id})
    logging.info(f"Archived: {subject!r} → '{folder_name}'")


def get_or_create_named_folder(client: GraphClient, folder_name: str) -> str:
    """Return the Exchange folder ID for *folder_name*, creating it if needed.

    Generic helper used for ticket-type sort folders (e.g. 'Rock Tickets',
    'Concrete Tickets').  Follows the same pattern as the dedicated review and
    processed-folder helpers.
    """
    list_url = (
        f"{GRAPH_BASE}/users/{MAILBOX}/mailFolders"
        f"?$filter=displayName eq '{folder_name}'&$select=id,displayName"
    )
    data    = client.get(list_url).json()
    folders = data.get("value", [])

    if folders:
        return folders[0]["id"]

    create_url = f"{GRAPH_BASE}/users/{MAILBOX}/mailFolders"
    created    = client.post(create_url, json={"displayName": folder_name}).json()
    folder_id  = created["id"]
    logging.info(f"Created mail folder '{folder_name}' (id={folder_id}).")
    return folder_id


def sort_email_to_type_folder(
    client: GraphClient,
    message_id: str,
    subject: str,
    folder_name: str,
) -> str:
    """Move an email to a ticket-type sort folder and mark it as read.

    Creates the folder if it does not already exist.

    Returns:
        The new message ID assigned by Exchange after the move.
    """
    folder_id = get_or_create_named_folder(client, folder_name)
    url       = f"{GRAPH_BASE}/users/{MAILBOX}/messages/{message_id}/move"
    moved     = client.post(url, json={"destinationId": folder_id}).json()
    new_id    = moved["id"]
    mark_email_as_read(client, new_id)
    logging.info(f"Sorted email to: {folder_name!r}")
    return new_id


def send_run_summary_email(
    client: GraphClient,
    processed_count: int,
    review_count: int,
    duplicate_count: int,
    error_count: int,
    flagged_items: "list[tuple[str, list[str]]]",
    error_items: "list[dict]",
) -> None:
    """Send a plain-text run summary email to SUMMARY_RECIPIENT.

    Sent from SUMMARY_RECIPIENT's own mailbox so it does not depend on
    help@mjhughes.com having Mail.Send permission.  Any failure is logged
    and silently swallowed — a summary email problem must never fail the run.
    """
    now      = datetime.now()
    now_str  = now.strftime("%Y-%m-%d %H:%M:%S")
    subj_ts  = now.strftime("%Y-%m-%d %H:%M")
    subject  = f"Ticket Processor Run Summary - {subj_ts}"
    sp_path  = (
        f"/Shared Documents/{SHAREPOINT_FOLDER}"
        f"/{_SP_JOB_FOLDER}/{_EXCEL_FILENAME}"
    )

    lines: list[str] = [
        f"Run completed at {now_str}",
        "",
        "Results:",
        f"- Tickets processed successfully: {processed_count}",
        f"- Tickets flagged for review: {review_count}",
        f"- Duplicate tickets detected: {duplicate_count}",
        f"- Errors: {error_count}",
        "",
        "Flagged for review:",
    ]
    if flagged_items:
        for subj, reasons in flagged_items:
            reason_str = "; ".join(reasons) if reasons else "flagged for review"
            lines.append(f"- {subj}: {reason_str}")
    else:
        lines.append("(none)")

    lines += ["", "Errors:"]
    if error_items:
        for i, item in enumerate(error_items, start=1):
            lines.append(f"  Error {i}:")
            lines.append(f"    Email: {item.get('subject', '(unknown)')}")
            if item.get("att_name"):
                lines.append(f"    File:  {item['att_name']}")
            lines.append(f"    Step:  {item.get('error_step', 'unknown')}")
            lines.append(f"    Error: {item.get('error_detail', '')}")
            tb = (item.get("traceback") or "").strip()
            if tb:
                lines.append("    Traceback:")
                for tb_line in tb.splitlines():
                    lines.append(f"      {tb_line}")
            lines.append("")
    else:
        lines.append("(none)")

    lines += [
        "",
        f"Excel file updated: {_EXCEL_FILENAME}",
        f"SharePoint location: {sp_path}",
    ]

    body    = "\r\n".join(lines)
    url     = f"{GRAPH_BASE}/users/{SUMMARY_RECIPIENT}/sendMail"
    payload = {
        "message": {
            "subject": subject,
            "body": {"contentType": "Text", "content": body},
            "toRecipients": [
                {"emailAddress": {"address": SUMMARY_RECIPIENT}}
            ],
        },
        "saveToSentItems": False,
    }

    try:
        client.post(url, json=payload)
        logging.info(f"Run summary email sent to {SUMMARY_RECIPIENT}.")
    except Exception as exc:
        logging.warning(f"Could not send run summary email: {exc}")


# ============================================================
# PDF → IMAGE CONVERSION
# ============================================================
def pdf_to_images(pdf_bytes: bytes, dpi: int = 200) -> list:
    """
    Render every page of a PDF as a PIL Image at the requested DPI.

    200 DPI is a good balance between QR readability and OCR accuracy
    without creating very large images.  Increase to 300 for poor scans.

    Args:
        pdf_bytes : Raw PDF file content.
        dpi       : Render resolution (72 DPI is PyMuPDF's base unit).

    Returns:
        List of PIL Image objects, one per PDF page.
    """
    images = []
    zoom = dpi / 72                        # PyMuPDF zoom factor
    matrix = fitz.Matrix(zoom, zoom)

    doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    for page in doc:
        pixmap = page.get_pixmap(matrix=matrix)
        img = Image.open(io.BytesIO(pixmap.tobytes("png")))
        images.append(img)
    doc.close()

    return images


# ============================================================
# IMAGE PREPROCESSING
# ============================================================
def preprocess_for_ocr(image: Image.Image) -> Image.Image:
    """
    Apply a minimal preprocessing pipeline to a PIL Image before passing it
    to Tesseract.  Steps are applied in this order:

      1. Upscale  — 2× enlargement (LANCZOS) gives Tesseract more pixel
                    detail to work with on low-resolution scans.
      2. Grayscale — single-channel image required for contrast enhancement.
      3. Contrast  — PIL ImageEnhance.Contrast ×1.2, a subtle boost only.
                     Keeps character detail intact; Tesseract handles the rest.

    Sharpening, denoising, and adaptive thresholding are intentionally omitted:
    on these scanned tickets they destroy fine character detail and cause table
    borders to bleed into adjacent text, which worsens Tesseract accuracy.
    """
    # 1. Upscale 2×
    w, h = image.size
    image = image.resize((w * 2, h * 2), Image.LANCZOS)

    # 2. Grayscale
    image = image.convert("L")

    # 3. Subtle contrast boost — preserves character detail
    image = ImageEnhance.Contrast(image).enhance(1.2)

    return image


# ============================================================
# QR CODE EXTRACTION
# ============================================================
def _zxing_scan(image: "Image.Image", label: str, page_num: int) -> Optional[dict]:
    """
    Try QR detection on *image* using pyzbar with two variants:
      1. Original PIL image (colour)
      2. Grayscale

    Returns a parsed QR dict on the first match, or None.
    When OCR_DEBUG=1 every attempt is logged.
    """
    if not PYZBAR_AVAILABLE:
        return None

    variants = [
        ("original",   image),
        ("grayscale",  image.convert("L")),
    ]

    for var_name, img in variants:
        results = pyzbar_decode(img, symbols=[ZBarSymbol.QRCODE])
        for r in results:
            text = r.data.decode("utf-8").strip()
            if OCR_DEBUG:
                logging.info(
                    f"  QR p{page_num} | {label} [{var_name}]: found {text!r}"
                )
            if QR_PATTERN.search(text):
                if OCR_DEBUG:
                    logging.info(
                        f"  QR p{page_num} | {label} [{var_name}]: MATCH {text!r}"
                    )
                return _parse_qr_string(text)
            if OCR_DEBUG:
                logging.info(
                    f"  QR p{page_num} | {label} [{var_name}]: "
                    f"barcode found but pattern not matched"
                )
        if OCR_DEBUG and not results:
            logging.info(
                f"  QR p{page_num} | {label} [{var_name}]: no barcode detected"
            )
    return None


def extract_qr_code(images: list) -> Optional[dict]:
    """
    Scan each page image for a QR code matching the ticket pattern.

    Detection strategy per page (stops at first success):
      1. Full page at original resolution
      2. Full page at 2× resolution
      3. Right 25% strip (full height)
      4. Upper-right 25%×25% crop
      5. OCR fallback — pytesseract --psm 6 regex search on full page

    Each image region is tried with both colour and grayscale variants via
    pyzbar.  When OCR_DEBUG=1 every attempt and its outcome is logged, and
    the first page is saved to qr_debug.png for inspection.

    Returns:
        Parsed QR data dict, or None if no valid code is found.
    """
    if not PYZBAR_AVAILABLE:
        logging.warning(
            "  QR crop scan | Step 3: pyzbar not available — "
            "crop-based QR detection skipped, falling back to OCR only"
        )

    if OCR_DEBUG and images:
        images[0].save("qr_debug.png")
        logging.info(
            f"QR debug: first page saved to qr_debug.png "
            f"(size {images[0].size[0]}×{images[0].size[1]})"
        )

    for page_num, image in enumerate(images, start=1):
        w, h = image.size

        logging.info(f"  QR p{page_num} | Step 3: Attempting pyzbar crop scans...")
        # ── Step 1: full page, original resolution ────────────────────────
        if OCR_DEBUG:
            logging.info(f"  QR p{page_num} | step 1: full page ({w}×{h})")
        result = _zxing_scan(image, "full-page", page_num)
        if result:
            return result

        # ── Step 2: full page, 2× upscale ────────────────────────────────
        img_2x = image.resize((w * 2, h * 2), Image.LANCZOS)
        if OCR_DEBUG:
            logging.info(f"  QR p{page_num} | step 2: full page 2× ({w*2}×{h*2})")
        result = _zxing_scan(img_2x, "full-page-2x", page_num)
        if result:
            return result

        # ── Step 3: right 25% strip (full height) ────────────────────────
        rs_box = (int(w * 0.75), 0, w, h)
        rs_crop = image.crop(rs_box)
        if OCR_DEBUG:
            logging.info(
                f"  QR p{page_num} | step 3: right-strip "
                f"({rs_crop.size[0]}×{rs_crop.size[1]}, box={rs_box})"
            )
        result = _zxing_scan(rs_crop, "right-strip", page_num)
        if result:
            return result

        # ── Step 4: upper-right 25%×25% crop ─────────────────────────────
        ur_box = (int(w * 0.75), 0, w, int(h * 0.25))
        ur_crop = image.crop(ur_box)
        if OCR_DEBUG:
            logging.info(
                f"  QR p{page_num} | step 4: upper-right "
                f"({ur_crop.size[0]}×{ur_crop.size[1]}, box={ur_box})"
            )
        result = _zxing_scan(ur_crop, "upper-right", page_num)
        if result:
            return result

        # ── Step 5: OCR fallback ──────────────────────────────────────────
        logging.info(f"  QR p{page_num} | step 5: OCR text fallback on crop image")
        ocr_text = pytesseract.image_to_string(image, config="--psm 6 --oem 3")
        match = QR_PATTERN.search(ocr_text)
        if match:
            logging.info(
                f"  QR p{page_num} | QR found via OCR text: {match.group(0)!r}"
            )
            return _parse_qr_string(match.group(0))
        else:
            logging.info(
                f"  QR p{page_num} | OCR fallback: pattern not found in crop OCR text"
            )

    return None   # Pattern not found on any page


def _parse_qr_string(raw: str) -> dict:
    """
    Parse a raw QR string into component fields.

    Input format : XXXX-YYYY-CCCCCC-CC
    Returns a dict with keys: job_number, location, cost_code_1,
    cost_code_2, cost_code, raw.
    """
    match = QR_PATTERN.search(raw)
    if not match:
        raise ValueError(f"String does not match QR pattern: {raw!r}")

    job_number   = match.group(1).upper()
    location     = match.group(2).upper()
    cost_code_1  = match.group(3).upper()
    cost_code_2  = match.group(4).upper()

    return {
        "job_number":  job_number,
        "location":    location,
        "cost_code_1": cost_code_1,
        "cost_code_2": cost_code_2,
        "cost_code":   f"{cost_code_1}-{cost_code_2}",
        "raw":         raw,
    }


def _scan_full_page_for_qr(
    page_image: "Image.Image",
    page_num: int,
) -> "tuple[Optional[dict], Optional[int]]":
    """Scan the full page image for a QR code without any cropping.

    Tries pyzbar on the original colour image then grayscale.  Stops at
    the first result that matches ``QR_PATTERN``.

    Returns:
        ``(qr_data, y_center_px)`` when a matching code is found, where
        ``y_center_px`` is the average Y coordinate of the QR code's bounding
        polygon in the full-page coordinate space.
        ``(None, None)`` when no matching code is found.

    This is the primary QR scan for multi-ticket pages.  Cropped regions
    are used only for OCR/AI extraction — QR detection stays on the full
    page so that codes near a crop boundary are not lost.
    """
    if not PYZBAR_AVAILABLE:
        logging.warning(
            f"  QR p{page_num} | Step 2: pyzbar not available — "
            "install libzbar0 and the pyzbar package"
        )
        return None, None

    logging.info(f"  QR p{page_num} | Step 2: Attempting pyzbar full-page scan...")
    variants = [
        ("original",  page_image),
        ("grayscale", page_image.convert("L")),
    ]
    for _var_name, img in variants:
        results = pyzbar_decode(img, symbols=[ZBarSymbol.QRCODE])
        for r in results:
            text = r.data.decode("utf-8").strip()
            if not QR_PATTERN.search(text):
                logging.info(
                    f"  QR p{page_num} | pyzbar [{_var_name}]: found {text!r} "
                    "but pattern not matched"
                )
                continue
            polygon  = r.polygon
            y_center = (
                sum(p.y for p in polygon) // len(polygon)
                if polygon
                else r.rect.top + r.rect.height // 2
            )
            return _parse_qr_string(text), y_center
    logging.info(f"  QR p{page_num} | pyzbar full-page: no QR found")
    return None, None


def _ai_scan_qr_from_image(
    page_image: "Image.Image",
    page_num: int,
) -> "tuple[Optional[dict], Optional[int]]":
    """Try to read the QR code sticker from *page_image* using Claude vision.

    Sends the full page image to Claude with a structured prompt asking it to
    locate and decode any QR code sticker matching the job-location-costcode
    pattern.  This is the primary QR detection method; pyzbar is the fallback.

    Returns:
        ``(qr_data, y_center_px)`` when Claude finds a valid matching QR code,
        where ``y_center_px`` is estimated from the reported sticker location.
        ``(None, None)`` when Claude reports no QR code, the returned text does
        not match the expected pattern, or on any API/parsing error.
    """
    if not ANTHROPIC_API_KEY:
        logging.warning(
            f"  QR p{page_num} | AI vision skipped — ANTHROPIC_API_KEY not set"
        )
        return None, None

    logging.info(f"  QR p{page_num} | Step 1: Attempting AI vision scan...")

    # Downscale the page so the payload stays manageable.
    img_copy = page_image.copy()
    max_side = 1500
    if max(img_copy.size) > max_side:
        ratio    = max_side / max(img_copy.size)
        new_size = (int(img_copy.size[0] * ratio), int(img_copy.size[1] * ratio))
        img_copy = img_copy.resize(new_size, Image.LANCZOS)
    buf = io.BytesIO()
    img_copy.convert("RGB").save(buf, format="JPEG", quality=85)
    b64_data = base64.b64encode(buf.getvalue()).decode("ascii")

    prompt = (
        "Look at this ticket image carefully.\n\n"
        "Find any small label sticker on this ticket. "
        "The sticker may be on the left side, right side, top, bottom, or any corner. "
        "It may be rotated vertically or at an angle.\n\n"
        "The sticker contains a QR code image AND printed text. "
        "IMPORTANT: Even if you cannot decode the QR code image itself, "
        "read the printed text on the sticker label directly. "
        "The text is printed in the format:\n"
        "XXXX-XXXX-XXXXXX-XX\n"
        "Example: 2601-0180-313713-99\n\n"
        "Look carefully at all edges and corners of the ticket for this sticker. "
        "The text may be printed vertically (rotated 90 degrees) alongside the QR code.\n\n"
        'Return ONLY a JSON object:\n'
        '{\n'
        '  "qr_found": true or false,\n'
        '  "qr_text": "the XXXX-XXXX-XXXXXX-XX text from the sticker, or null",\n'
        '  "qr_location": "top-right/top-left/bottom-right/bottom-left/left-side/right-side"\n'
        '}\n'
        "No other text."
    )

    try:
        ai_client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        response  = ai_client.messages.create(
            model=_AI_MODEL,
            max_tokens=128,
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type":       "base64",
                            "media_type": "image/jpeg",
                            "data":       b64_data,
                        },
                    },
                    {"type": "text", "text": prompt},
                ],
            }],
        )
        raw = response.content[0].text.strip()
        if raw.startswith("```"):
            raw = re.sub(r"^```[a-z]*\n?", "", raw)
            raw = re.sub(r"\n?```$",        "", raw)
        logging.info(f"  QR p{page_num} | AI QR detection raw response: {raw!r}")
        ai_result = json.loads(raw)
    except Exception as exc:
        logging.warning(f"  QR p{page_num} | AI vision error: {exc}")
        return None, None

    qr_found = ai_result.get("qr_found")
    qr_text  = (ai_result.get("qr_text") or "").strip()
    location = (ai_result.get("qr_location") or "").lower()
    logging.info(
        f"  QR p{page_num} | AI result: qr_found={qr_found}  "
        f"qr_text={qr_text!r}  location={location!r}"
    )

    if not qr_found:
        logging.info(
            f"  QR p{page_num} | AI could not find QR code — trying pyzbar fallback"
        )
        return None, None

    if not QR_PATTERN.search(qr_text):
        logging.warning(
            f"  QR p{page_num} | AI vision: qr_found=true but "
            f"{qr_text!r} does not match expected pattern XXXX-XXXX-XXXXXX-XX "
            "— treating as not found"
        )
        return None, None

    # Estimate a Y coordinate from the reported sticker location so the
    # boundary-assignment logic can match the QR to the correct ticket crop.
    ph = page_image.size[1]
    if "top" in location:
        y_est = int(ph * 0.15)
    elif "bottom" in location:
        y_est = int(ph * 0.85)
    else:
        y_est = int(ph * 0.50)

    logging.info(
        f"  [p{page_num}] QR found via AI vision: "
        f"'{qr_text}' at {location or 'unknown location'}"
    )
    return _parse_qr_string(qr_text), y_est


# ============================================================
# OCR CONFIDENCE CHECK
# ============================================================
def _run_confidence_check(
    check_name: str, full_text: str, profile: TicketProfile,
    images: list = None,
) -> Optional[str]:
    """Run one named confidence check from the profile's confidence_checks list.

    Returns None when the check passes, or a human-readable failure description
    when it fails.  Unknown check names are logged and treated as passing so
    adding new names to a profile never silently breaks old installations.
    """
    if check_name == "valid_date_on_line_2":
        date_str, _ = _parse_ticket_header(full_text)
        return None if date_str else "date not found on header line"

    if check_name == "five_digit_ticket_on_line_2":
        # Primary: look for the number on the header line of the full-page text.
        _, ticket = _parse_ticket_header(full_text)
        if ticket:
            return None
        # Fallback: if the profile uses region OCR (bold corner number), try
        # cropping that region directly — many tickets place the number there
        # and it never appears in the full-page text stream.
        region_config = profile.layout.get("ticket_number_extraction", {})
        if region_config.get("method") == "region_ocr" and images:
            region_ticket = _ocr_ticket_number_from_region(images[0], region_config)
            if region_ticket:
                return None   # found via region OCR — check passes
        return "5-digit ticket number not found on header line or upper-right region"

    if check_name == "product_label_found":
        labels = profile.labels.get("material", ["Product"])
        if isinstance(labels, str):
            labels = [labels]
        found = any(
            re.search(re.escape(lbl.rstrip(":. ")), full_text, re.IGNORECASE)
            for lbl in labels
        )
        if found:
            return None
        # Print the full OCR text so we can see what Tesseract actually read
        logging.warning("  Full OCR text (product label not found):")
        for line in full_text.splitlines():
            logging.warning(f"    | {line}")
        return f"material/product label not found (expected one of: {labels})"

    if check_name == "net_row_found":
        net_label = profile.labels.get("net_row", "Net")
        # Primary: case-insensitive word search for the configured label
        # (handles NET, net, Net).
        if re.search(rf'\b{re.escape(net_label)}\b', full_text, re.IGNORECASE):
            return None
        # Primary extension: common single-character OCR misreads of the
        # first letter ("Met", "Wet" for "Net").  Only applied when the label
        # is a short word so the pattern stays tight.
        if len(net_label) <= 5:
            rest = re.escape(net_label[1:])
            misread_matches = re.findall(rf'\b[A-Za-z]{rest}\b', full_text, re.IGNORECASE)
            # Accept if a same-length word ending in the right suffix exists
            # that is NOT a longer word (avoids "Netric", "Netting", etc.).
            if any(len(m) == len(net_label) for m in misread_matches):
                return None
        # Fallback 1: classic weight-table row — 4+-digit integer (pounds)
        # followed by two decimals (tons, metric).
        if re.search(r'^\s*\d{4,}\s+\d+\.\d+\s+\d+\.\d+', full_text, re.MULTILINE):
            return None
        # Fallback 2: two adjacent decimal numbers separated by ≤5 arbitrary
        # characters — catches garbled pounds values and spurious degree
        # symbols from low-quality scans (e.g. "13.18° 11.96°").
        if re.search(r'\d+\.\d+.{0,5}\d+\.\d+', full_text):
            return None
        return f"'{net_label}' row not found and no weight-table row pattern detected"

    if check_name == "facility_label_found":
        labels = profile.labels.get("facility", ["Location"])
        if isinstance(labels, str):
            labels = [labels]
        found = any(
            re.search(re.escape(lbl.rstrip(":. ")), full_text, re.IGNORECASE)
            for lbl in labels
        )
        if not found:
            return f"facility label not found (expected one of: {labels})"
        facility_val = _ocr_facility(full_text, profile)
        if facility_val and re.search(r'[&$@#%]', facility_val):
            return f"facility name contains garble characters: {facility_val!r}"
        return None

    if check_name == "valid_date_found":
        # Accept any common date format anywhere in the text.
        date_patterns = [
            r'\b\d{1,2}/\d{1,2}/\d{2,4}\b',   # MM/DD/YYYY or M/D/YY
            r'\b\d{1,2}-\d{1,2}-\d{2,4}\b',    # MM-DD-YYYY
            r'\b\d{4}-\d{2}-\d{2}\b',           # YYYY-MM-DD (ISO)
        ]
        if any(re.search(p, full_text) for p in date_patterns):
            return None
        return "no date found in ticket text"

    if check_name == "ticket_number_found":
        # Look for a 5-8 digit number in the first 10 non-empty lines where
        # the ticket number is typically printed.
        all_lines  = [l.strip() for l in full_text.splitlines() if l.strip()]
        header     = "\n".join(all_lines[:10])
        if re.search(r'\b[1-9]\d{4,7}\b', header):
            return None
        return "no 5-8 digit ticket number found near top of ticket"

    if check_name == "six_digit_ticket_number_found":
        # Specific check for suppliers with 6-digit ticket numbers (e.g. Troutdale
        # Sand & Gravel).  Searches the first 15 non-empty lines so that tickets
        # where the number appears slightly lower on the page still pass.
        all_lines = [l.strip() for l in full_text.splitlines() if l.strip()]
        header    = "\n".join(all_lines[:15])
        if re.search(r'\b[1-9]\d{5}\b', header):
            return None
        return "no 6-digit ticket number found near top of ticket"

    if check_name == "supplier_name_in_text":
        # Verify that at least one of the profile's detection_keywords appears
        # somewhere in the full OCR text (case-insensitive).  Useful for suppliers
        # whose name is readable in the body of the ticket even when the header is
        # garbled by box formatting.
        full_upper = full_text.upper()
        for keyword in profile.detection_keywords:
            if keyword.upper() in full_upper:
                return None
        return (
            f"supplier name not found in ticket text "
            f"(expected one of: {profile.detection_keywords})"
        )

    logging.warning(
        f"Unknown confidence check {check_name!r} in profile "
        f"{profile.supplier_name!r} — skipping"
    )
    return None


def _check_ocr_confidence(
    full_text: str, profile: TicketProfile, images: list = None,
) -> tuple[bool, list[str]]:
    """Run all confidence checks defined in the profile.

    Returns:
        (all_passed, failed_descriptions) — all_passed is True only when
        the failed list is empty.
    """
    failed: list[str] = []
    for check_name in profile.confidence_checks:
        result = _run_confidence_check(check_name, full_text, profile, images=images)
        if result is not None:
            failed.append(result)
    return len(failed) == 0, failed


# ============================================================
# AUTO-PROFILE GENERATION
# ============================================================
def _generate_auto_profile(
    company_name: str,
    full_text: str,
    image: "Optional[Image.Image]",
    ai_fields: dict,
) -> "Optional[TicketProfile]":
    """Generate a ticket profile JSON for an unknown supplier via Claude API.

    Called when the generic fallback profile is used and AI extraction succeeds.
    Saves the generated profile JSON to ticket_profiles/ and returns a loaded
    TicketProfile so it can be used immediately for remaining tickets in this run.

    Returns None on any error (Claude API failure, invalid JSON, validation failure).
    """
    if not ANTHROPIC_API_KEY:
        logging.warning("  Auto-profile generation skipped — ANTHROPIC_API_KEY not set.")
        return None

    # Guess ticket type from the fields AI extracted via the generic profile.
    if ai_fields.get("qty_delivered") or ai_fields.get("slump"):
        guessed_type = "concrete"
    elif ai_fields.get("net_tons"):
        guessed_type = "rock"
    else:
        guessed_type = ""

    # Load a reference profile to show Claude the expected structure.
    _ref_path = Path(TICKET_PROFILES_DIR) / "teevin_bros.json"
    example_json = _ref_path.read_text(encoding="utf-8") if _ref_path.exists() else "{}"

    prompt = f"""\
You have successfully extracted data from a delivery ticket from the supplier "{company_name}".

Now generate a JSON profile for this supplier so future tickets can be processed automatically
without relying on the generic fallback.

The profile MUST be a valid JSON object containing ALL of these required keys:

  supplier_name        (string)  — full company name as it appears on the ticket
  detection_keywords   (array)   — words/phrases to scan in OCR text to identify this supplier
                                   Include at least 2-3 variants (e.g. abbreviations, all-caps)
  layout               (object)  — must include a "ticket_number_extraction" sub-object with:
                                     method: "region_ocr"
                                     region: {{ x_start_pct, y_start_pct, x_end_pct, y_end_pct }}
                                       (use 0.50–1.00 x, 0.00–0.25 y for the top-right area)
                                     tesseract_config: "--oem 3 --psm 6"
                                     ticket_number_length: (integer, typically 5-8)
  labels               (object)  — maps field names to arrays of label text variants found on
                                   the ticket (e.g. "customer": ["Customer:", "SOLD TO:"])
  weight_table         (object)  — {{ "columns": [...], "tons_column_index": N }} for rock/aggregate
                                   tickets; use {{}} for concrete tickets
  ocr_corrections      (object)  — common OCR misreads as "wrong": "correct" pairs; use {{}} if none
  confidence_checks    (array)   — always include "supplier_name_in_text"

Optional but recommended:
  ticket_type          (string)  — "rock", "concrete", or omit if unknown
  detection_scan_lines (integer) — lines to scan for keywords (default 3; use more if name is deep)
  notes                (object)  — human-readable extraction hints per field

Reference profile (teevin_bros.json) — use this as a structural template:
{example_json}

CONTEXT:
  Supplier name detected: {company_name}
  Guessed ticket type:    {guessed_type or "unknown"}

Fields extracted by the generic AI profile:
{json.dumps(ai_fields, indent=2)}

Return ONLY the JSON object — no markdown fences, no explanation.

OCR TEXT (first 3000 chars):
{full_text[:3000]}"""

    ai_client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    # Include the ticket image so Claude can see the layout directly.
    if image is not None:
        img_copy = image.copy()
        max_side = 1500
        if max(img_copy.size) > max_side:
            ratio    = max_side / max(img_copy.size)
            new_size = (int(img_copy.size[0] * ratio), int(img_copy.size[1] * ratio))
            img_copy = img_copy.resize(new_size, Image.LANCZOS)
        buf = io.BytesIO()
        img_copy.convert("RGB").save(buf, format="JPEG", quality=85)
        b64_data = base64.b64encode(buf.getvalue()).decode("ascii")
        content = [
            {
                "type": "image",
                "source": {
                    "type":       "base64",
                    "media_type": "image/jpeg",
                    "data":       b64_data,
                },
            },
            {"type": "text", "text": prompt},
        ]
    else:
        content = prompt

    try:
        response = ai_client.messages.create(
            model=_AI_MODEL,
            max_tokens=2048,
            messages=[{"role": "user", "content": content}],
        )
        raw_json = response.content[0].text.strip()

        # Strip accidental markdown fences.
        if raw_json.startswith("```"):
            raw_json = re.sub(r"^```[a-z]*\n?", "", raw_json)
            raw_json = re.sub(r"\n?```$",        "", raw_json)

        profile_data = json.loads(raw_json)
        logging.info(f"  Auto-profile AI response parsed OK for {company_name!r}")
    except Exception as exc:
        logging.warning(f"  Auto-profile generation failed for {company_name!r}: {exc}")
        return None

    # Validate the generated profile against the required schema.
    errors = validate_profile(profile_data, "<auto-generated>")
    if errors:
        logging.warning(
            f"  Auto-generated profile for {company_name!r} has validation errors: "
            + "; ".join(errors)
        )
        return None

    # Derive a safe filename: lowercase, special chars → underscores.
    sanitized = re.sub(r"[^a-z0-9]+", "_", company_name.lower()).strip("_")
    filename   = f"{sanitized}.json"
    profile_path = Path(TICKET_PROFILES_DIR) / filename

    try:
        profile_path.write_text(json.dumps(profile_data, indent=2), encoding="utf-8")
        logging.info(f"  Auto-profile saved: {profile_path}")
    except OSError as exc:
        logging.warning(f"  Could not save auto-profile to {profile_path}: {exc}")
        return None

    # Append a note to the unknown_suppliers log.
    try:
        with open(UNKNOWN_SUPPLIERS_LOG, "a", encoding="utf-8") as f:
            f.write(
                f"AUTO-PROFILE CREATED: {company_name} → {filename}"
                f"  Review this profile and verify all fields before relying on it.\n"
            )
    except OSError:
        pass

    # Build and return a TicketProfile for immediate in-memory use.
    new_profile = TicketProfile(
        supplier_name               = profile_data["supplier_name"],
        detection_keywords          = profile_data["detection_keywords"],
        layout                      = profile_data.get("layout", {}),
        labels                      = profile_data.get("labels", {}),
        weight_table                = profile_data.get("weight_table", {}),
        ocr_corrections             = profile_data.get("ocr_corrections", {}),
        confidence_checks           = profile_data["confidence_checks"],
        source_file                 = str(profile_path),
        is_fallback                 = False,
        ticket_type                 = str(profile_data.get("ticket_type", guessed_type)),
        detection_scan_lines        = int(profile_data.get("detection_scan_lines", 3)),
    )
    return new_profile


# ============================================================
# OCR TEXT EXTRACTION
# ============================================================
def extract_ticket_data_ocr(images: list, profiles: list[TicketProfile], subject: str = "") -> dict:
    """Run OCR, detect supplier profile, apply corrections, check confidence, extract fields.

    Returns a dict.  Special keys (all start with "_"):
        _ocr_confidence_passed : bool — False when extraction should not be trusted.
        _no_profile_match      : bool — True when no supplier profile was identified.
        _company_name          : str  — first OCR line when no profile matched.
        _failed_checks         : list — check failure descriptions on confidence failure.
    """
    # 1. Run OCR on all pages (no corrections yet — need raw text for profile detection).
    raw_text = "\n".join(
        pytesseract.image_to_string(preprocess_for_ocr(img), config="--oem 3 --psm 6")
        for img in images
    )

    # 2. Detect supplier profile from raw OCR text (AI vision fallback if needed).
    profile = detect_profile(raw_text, profiles, image=images[0] if images else None)
    if not profile:
        # No profiles loaded at all (not even a fallback) — flag for review.
        company_name = next(
            (line.strip() for line in raw_text.splitlines() if line.strip()), "unknown"
        )
        logging.warning(
            f"  No matching profile found. First OCR line: {company_name!r}"
        )
        _log_unknown_supplier(company_name)

        try:
            Path("ocr_debug.txt").write_text(raw_text, encoding="utf-8")
            if images:
                preprocess_for_ocr(images[0]).save("ocr_debug.png")
            logging.info(f"OCR debug saved for unknown supplier: {company_name!r}")
        except Exception as _dbg_exc:
            logging.warning(f"  Could not save OCR debug files: {_dbg_exc}")

        return {
            "_ocr_confidence_passed": False,
            "_no_profile_match": True,
            "_company_name": company_name,
        }

    if profile.is_fallback:
        # A specific supplier profile did not match — using the generic fallback.
        # Log the supplier name so a dedicated profile can be created later.
        company_name = next(
            (line.strip() for line in raw_text.splitlines() if line.strip()), "unknown"
        )
        logging.info(
            f"  No specific profile matched — using generic profile for: {company_name!r}"
        )
        _log_unknown_supplier(company_name)

        # Save OCR debug so we can inspect what Tesseract read and build a
        # dedicated profile for this supplier.
        try:
            Path("ocr_debug.txt").write_text(raw_text, encoding="utf-8")
            if images:
                preprocess_for_ocr(images[0]).save("ocr_debug.png")
            logging.info(f"OCR debug saved for unknown supplier: {company_name!r}")
        except Exception as _dbg_exc:
            logging.warning(f"  Could not save OCR debug files: {_dbg_exc}")
    else:
        logging.info(f"  Detected profile: {profile.supplier_name!r}")

    # 3. Apply profile-specific OCR corrections.
    full_text = _apply_ocr_corrections(raw_text, profile)
    logging.debug(f"OCR full text (after corrections):\n{full_text}")

    if OCR_DEBUG:
        Path("ocr_debug.txt").write_text(full_text, encoding="utf-8")
        logging.info("OCR debug: corrected text saved to ocr_debug.txt")
        if images:
            debug_img = preprocess_for_ocr(images[0])

            # If the profile uses region OCR for ticket number, overlay a red
            # rectangle on the debug image showing the exact crop area, and save
            # the preprocessed crop separately so the region can be verified.
            region_config = profile.layout.get("ticket_number_extraction", {})
            if region_config.get("method") == "region_ocr":
                orig_w, orig_h = images[0].size
                rgn   = region_config.get("region", {})
                x1    = int(orig_w * rgn.get("x_start_pct", 0.70))
                y1    = int(orig_h * rgn.get("y_start_pct", 0.00))
                x2    = int(orig_w * rgn.get("x_end_pct",   1.00))
                y2    = int(orig_h * rgn.get("y_end_pct",   0.15))

                # Save the preprocessed crop (what Tesseract actually sees)
                crop_debug = preprocess_for_ocr(images[0].crop((x1, y1, x2, y2)))
                crop_debug.save("ocr_debug_topright.png")
                logging.info(
                    "OCR debug: ticket-number crop saved to ocr_debug_topright.png "
                    f"(original coords x={x1}-{x2}, y={y1}-{y2})"
                )

                # Draw a red rectangle on the full-page debug image.
                # The preprocessed image is 2× the original, so scale coords.
                scale = 2
                draw  = ImageDraw.Draw(debug_img)
                draw.rectangle(
                    [x1 * scale, y1 * scale, x2 * scale - 1, y2 * scale - 1],
                    outline="red",
                    width=4,
                )

            debug_img.save("ocr_debug.png")
            logging.info("OCR debug: preprocessed first page saved to ocr_debug.png")

    # 4. Run profile confidence checks.
    confidence_passed, failed_checks = _check_ocr_confidence(full_text, profile, images=images)
    if not confidence_passed:
        for check in failed_checks:
            logging.warning(f"  OCR confidence FAILED: {check}")
        logging.warning(
            f"OCR confidence check failed ({len(failed_checks)} check(s) failed). "
            "Ticket will be flagged for manual review."
        )
        return {"_ocr_confidence_passed": False, "_failed_checks": failed_checks}

    # 5. Extract all fields — try Claude AI first, fall back to regex.
    ticket_data: dict = {"_ocr_confidence_passed": True}
    ai_used = False
    _label = subject or "(unknown email)"

    if ANTHROPIC_API_KEY:
        logging.info(f"  Using AI extraction for: {_label}")
        try:
            # Build a supplier-specific ticket number hint for the AI prompt.
            # If the profile provides an explicit 'ticket_number_ai_hint' use it
            # verbatim (useful when OCR is unreliable for that supplier).
            # Otherwise auto-generate a hint whenever the ticket number length
            # differs from the default 5 digits.
            tn_config  = profile.layout.get("ticket_number_extraction", {})
            tn_length  = tn_config.get("ticket_number_length", 5)
            tn_label   = tn_config.get("ticket_number_label", "")
            custom_hint = tn_config.get("ticket_number_ai_hint", "")
            if custom_hint:
                _tn_hint = custom_hint
            elif tn_length != 5:
                _tn_hint = (
                    f"For {profile.supplier_name}, the ticket number is {tn_length} digits"
                    + (f", found after the '{tn_label}' label" if tn_label else "")
                    + ". The OCR may split it across two consecutive lines — concatenate"
                    " the parts in order to form the full number."
                )
            else:
                _tn_hint = ""
            _generic_note = (
                "This is a material delivery ticket from an unknown supplier. "
                "Extract the following fields using common ticket conventions. "
                "The ticket number is typically a bold 5-8 digit number near the top right. "
                "The facility is where the material came from, not the customer receiving it. "
                "The net tons is in the Tons column of the Net row in the weight table."
            ) if profile.is_fallback else ""
            ai_fields = extract_fields_with_ai(
                full_text, profile.supplier_name,
                image=images[0] if images else None,
                ticket_number_hint=_tn_hint,
                generic_note=_generic_note,
                ticket_type=profile.ticket_type,
            )
            ticket_data.update(ai_fields)
            ai_used = True

            # Date: log which method produced a value; fall back to header-line
            # OCR when AI returns null.  Keeping "" (not REVIEW_REQUIRED) on
            # total failure lets the yellow-highlight logic flag the cell.
            _ai_date = ticket_data.get("date", "")
            if _ai_date:
                logging.info(f"  Date from AI vision: {_ai_date}")
            else:
                logging.info("  Date from AI null — trying OCR fallback")
                _ocr_date_fallback, _ = _parse_ticket_header(full_text)
                if _ocr_date_fallback:
                    ticket_data["date"] = _ocr_date_fallback
                    logging.info(f"  Date from OCR fallback: {_ocr_date_fallback}")
                else:
                    logging.info(
                        "  Date not found by either method — flagging cell yellow"
                    )
                    # ticket_data["date"] stays "" → yellow highlight in Excel

            # Auto-profile generation: when the generic fallback matched this
            # ticket and AI extraction succeeded, ask Claude to generate a
            # dedicated profile JSON for the supplier so future tickets from the
            # same company are processed with proper field mappings.
            if profile.is_fallback:
                _auto_company = next(
                    (l.strip() for l in raw_text.splitlines() if l.strip()),
                    "unknown supplier",
                )
                logging.info(
                    f"  Generic profile used — attempting auto-profile generation "
                    f"for: {_auto_company!r}"
                )
                _new_profile = _generate_auto_profile(
                    _auto_company,
                    full_text,
                    images[0] if images else None,
                    ai_fields,
                )
                if _new_profile:
                    profiles.append(_new_profile)
                    logging.info(
                        f"  Auto-profile added for immediate use: "
                        f"{_new_profile.supplier_name!r}"
                    )
                    # Tag the ticket so the reviewer knows the profile was auto-generated.
                    _caution = "Auto-generated profile used - verify all fields"
                    if ticket_data.get("helper_notes"):
                        ticket_data["helper_notes"] = (
                            _caution + "; " + ticket_data["helper_notes"]
                        )
                    elif ticket_data.get("material"):
                        ticket_data["material"] = (
                            _caution + "; " + ticket_data["material"]
                        )
                    else:
                        # No existing notes field — set whichever key is relevant.
                        if _new_profile.ticket_type == "concrete":
                            ticket_data["helper_notes"] = _caution
                        else:
                            ticket_data["material"] = _caution

        except Exception as ai_exc:
            logging.warning(f"  AI extraction failed, falling back to regex: {ai_exc}")
    else:
        logging.warning("  ANTHROPIC_API_KEY not set — using regex extraction.")

    if not ai_used:
        logging.info(f"  Using regex extraction for: {_label}")
        logging.warning(
            "  AI unavailable — ticket number will be empty (OCR is not used "
            "for ticket numbers). Ticket will be flagged for review."
        )
        if profile.ticket_type == "concrete":
            ticket_data.update({
                "date":          _ocr_date(full_text),
                "ticket_number": "",   # ticket number requires AI vision
                "supplier":      profile.supplier_name,
                "slump":         "",
                "qty_delivered": "",
                "helper_notes":  _ocr_material(full_text, profile),
            })
        else:
            ticket_data.update({
                "date":          _ocr_date(full_text),
                "facility":      _ocr_facility(full_text, profile),
                "customer":      _ocr_customer(full_text, profile),
                "material":      _ocr_material(full_text, profile),
                "ticket_number": "",   # ticket number requires AI vision
                "net_tons":      _ocr_net_tons(full_text, profile),
            })

    # 6. Generated ID: for profiles with no printed ticket number, build a unique
    # identifier from date + net_tons per the profile's generated_id config and
    # inject it as the ticket_number so duplicate detection and Excel entry work
    # the same as for any other profile.
    gen_id_cfg = profile.layout.get("generated_id")
    if gen_id_cfg and not ticket_data.get("ticket_number"):
        _date_raw = ticket_data.get("date", "")
        _tons_raw = ticket_data.get("net_tons", "")
        try:
            _date_seg = datetime.strptime(_date_raw, "%m/%d/%Y").strftime("%m%d%Y")
        except (ValueError, TypeError):
            _date_seg = re.sub(r"[^0-9]", "", str(_date_raw))
        _tons_seg    = str(_tons_raw) if _tons_raw else ""
        _prefix      = gen_id_cfg.get("format", "GENERATED-{MMDDYYYY}-{net_tons}").split("{")[0]
        generated_id = f"{_prefix}{_date_seg}-{_tons_seg}"
        ticket_data["ticket_number"] = generated_id
        logging.info(f"  Generated ticket ID (no printed number): {generated_id!r}")
        _gen_note = "No ticket number - ID generated from date and tonnage"
        if ticket_data.get("helper_notes"):
            ticket_data["helper_notes"] = _gen_note + " | " + ticket_data["helper_notes"]
        else:
            ticket_data["helper_notes"] = _gen_note

    # 7. Post-extraction check: warn on any required field that is still empty.
    if profile.ticket_type == "concrete":
        required_fields = ["date", "ticket_number", "qty_delivered"]
    else:
        required_fields = ["date", "facility", "customer", "material", "ticket_number", "net_tons"]
    for field_name in required_fields:
        if not ticket_data.get(field_name):
            src = "AI" if ai_used else "regex"
            logging.warning(f"  {src} extraction: could not extract '{field_name}'.")

    return ticket_data


REVIEW_REQUIRED = "REVIEW REQUIRED"


class AmbiguousTabError(Exception):
    """Raised when multiple Excel tabs match a QR code and AI cannot choose one."""


# ============================================================
# AI FIELD EXTRACTION
# ============================================================
_AI_MODEL  = "claude-sonnet-4-6"
_AI_PROMPT = """\
You are given both the original ticket image and raw OCR text extracted from it.
Use both to extract fields — the image is the authoritative source for values
that OCR may have missed or misread.

{ticket_number_preamble}\
OCR text may contain character-level errors from scan quality \
(e.g. "Bradiey" = "Bradley", "3)4-0" = "3/4-0").

Extract exactly these seven fields:

  ticket_number    — {ticket_number_field_hint}
  date             — CRITICAL FIELD — delivery date, normalised to MM/DD/YYYY.
                     Find this by looking DIRECTLY at the image; do not rely
                     on OCR text alone.  The date may be:
                     - Near the top of the ticket, next to the ticket number
                     - Partially covered by a QR sticker
                     - On the same line as the ticket number (e.g. Teevin Bros)
                     - After a DATE: or Date/Time: label (Knife River, others)
                     - In any format: M/D/YYYY, MM/DD/YYYY, MM/DD/YY, M/D/YY,
                       or written out (e.g. "Dec 8 2024")
                     Look carefully at the entire top quarter of the image.
                     Return null ONLY if you genuinely cannot find any date.
  facility         — quarry / pit / source location printed after the
                     "Location:" label.  NEVER a customer or contractor name.
                     Must not contain "MJ Hughes" or "Hughes Construction".
  customer         — company receiving the material, printed after "Customer:"
  material         — description of what was delivered.
                     The material/product field may be labeled as any of these
                     on the ticket: Product, Product., Material, Material:,
                     Item, Description, Mat, Prod. They all refer to the same
                     thing — extract the text after whichever label appears.
                     Include sizing notation (e.g. "3/4-0\"").
                     Exclude leading numeric product codes and quantity values.
  net_tons         — decimal number from the Tons column on the Net weight row.
                     Return the number only, no units.
  qr_sticker_text  — the short text label printed on the QR code sticker itself.
                     It appears rotated 90 degrees on the right side of the
                     ticket beside or below the QR code (e.g. "Class 2 RipRap",
                     "Aggregate Base", "3/4-0 Crushed").  Look at the image
                     directly for this text.  Return null if not visible.

Rules:
  • Correct obvious OCR errors based on context.
  • If a field genuinely cannot be determined, return null for that key.
  • Return ONLY a JSON object — no markdown fences, no explanation.

Supplier: {supplier_name}
{ticket_number_supplier_note}{generic_context}
OCR TEXT:
{ocr_text}"""

_AI_PROMPT_CONCRETE = """\
You are given both the original ticket image and raw OCR text extracted from it.
Use both to extract fields — the image is the authoritative source for values
that OCR may have missed or misread.

{ticket_number_preamble}\
OCR text may contain character-level errors from scan quality.

Extract exactly these seven fields:

  ticket_number  — {ticket_number_field_hint}
  date           — CRITICAL FIELD — delivery date, normalised to MM/DD/YYYY.
                   Find this by looking DIRECTLY at the image; do not rely
                   on OCR text alone.  The date may be:
                   - Near the top of the ticket, next to the ticket number
                   - Partially covered by a QR sticker
                   - After a DATE: or Date/Time: label
                   - In any format: M/D/YYYY, MM/DD/YYYY, MM/DD/YY, M/D/YY,
                     or written out (e.g. "Dec 8 2024")
                   Look carefully at the entire top quarter of the image.
                   Strip any time portion and normalise to MM/DD/YYYY.
                   Return null ONLY if you genuinely cannot find any date.
  supplier       — the concrete supplier company name printed at the top
                   of the ticket (e.g. "CalPortland").
  slump          — numeric value after the label "Slump:" in the delivery
                   data row (the row that also contains "Scheduled Arrival:"
                   and "Load#:"). Example: "Scheduled Arrival:  Slump: 4.00
                   Load#:" → 4.00. The "Slump on arrival" field in the upper
                   left section is a blank hand-fill line — ignore it entirely.
                   Return null if the slump value cannot be found.
  qty_delivered  — numeric value after "Qty This Load:" label.
                   This is the cubic yards delivered on this load.
                   Return the number only, no units.
  helper_notes   — product description from the "Product:" line.
                   Strip the leading product code and dash separator
                   (e.g. "0446FS - 4000 PSI WITH 25% SL" → "4000 PSI WITH 25% SL").
                   Return just the description text, no leading code.
  qr_sticker_text — the short text label printed on the QR code sticker itself.
                   Return null if not visible.

Rules:
  • Correct obvious OCR errors based on context.
  • If a field genuinely cannot be determined, return null for that key.
  • Return ONLY a JSON object — no markdown fences, no explanation.

Supplier: {supplier_name}
{ticket_number_supplier_note}{generic_context}
OCR TEXT:
{ocr_text}"""


def extract_fields_with_ai(
    ocr_text: str,
    supplier_name: str,
    image: "Optional[Image.Image]" = None,
    ticket_number_hint: str = "",
    generic_note: str = "",
    ticket_type: str = "",
) -> dict:
    """Send OCR text — and optionally the ticket image — to Claude for field extraction.

    When *image* is provided it is resized to max 1500 px on the longest side,
    base64-encoded as JPEG, and sent as a vision content block alongside the
    OCR text.  This lets Claude read fields (such as the bold upper-right ticket
    number) that OCR may have missed entirely.

    *ticket_number_hint* is an optional supplier-specific note injected into the
    prompt when the ticket number format differs from the default (e.g. 8-digit
    Knife River numbers that OCR splits across two lines).

    *generic_note* is injected when the generic fallback profile is used, giving
    Claude additional context about how to interpret an unknown ticket layout.

    *ticket_type* switches between rock/generic and concrete field sets.

    Rock/generic returns: ticket_number, date, facility, customer, material,
        net_tons, qr_sticker_text
    Concrete returns: ticket_number, date, supplier, slump, qty_delivered,
        helper_notes, qr_sticker_text

    Raises on API error or invalid JSON — caller falls back to regex extraction.
    """
    ai_client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    # Global ticket-number vision instruction — applied to every extraction.
    # OCR is unreliable for ticket numbers (thin fonts, digit-letter confusion);
    # AI reads the image directly and is the only source of truth for this field.
    _TN_VISION = (
        "IMPORTANT — Ticket Number: always read the ticket number directly from "
        "the image. It is a bold prominent number near the top of the ticket, "
        "typically labeled 'Ticket No.', 'Ticket #', 'TICKET NO.', or similar. "
        "Do not rely on OCR text for this field — OCR often misreads digits as "
        "letters (e.g. 0\u2192O, 6\u2192G, 9\u2192Q, 5\u2192S or I). "
        "Return only the numeric digits."
    )

    # Build the supplier-specific ticket number fragments for the prompt template.
    if ticket_number_hint:
        ticket_number_preamble   = _TN_VISION + "\n\n" + ticket_number_hint + "\n\n"
        ticket_number_field_hint = (
            "ticket number — see the supplier-specific note above for the exact"
            " format.  Read directly from the image; do not use OCR text."
        )
        ticket_number_supplier_note = (
            f"\nSupplier-specific note: {ticket_number_hint}"
        )
    else:
        ticket_number_preamble   = _TN_VISION + "\n\n"
        ticket_number_field_hint = (
            "bold prominent number near the top of the ticket — read from the"
            " image directly, not from OCR text."
        )
        ticket_number_supplier_note = ""

    generic_context = f"\n{generic_note}" if generic_note else ""

    prompt_template = _AI_PROMPT_CONCRETE if ticket_type == "concrete" else _AI_PROMPT
    prompt_text = prompt_template.format(
        supplier_name=supplier_name,
        ocr_text=ocr_text,
        ticket_number_preamble=ticket_number_preamble,
        ticket_number_field_hint=ticket_number_field_hint,
        ticket_number_supplier_note=ticket_number_supplier_note,
        generic_context=generic_context,
    )

    if image is not None:
        # Downscale to keep the payload manageable (~300 KB at q=85 for a typical ticket)
        img_copy = image.copy()
        max_side = 1500
        if max(img_copy.size) > max_side:
            ratio    = max_side / max(img_copy.size)
            new_size = (int(img_copy.size[0] * ratio), int(img_copy.size[1] * ratio))
            img_copy = img_copy.resize(new_size, Image.LANCZOS)
        buf = io.BytesIO()
        img_copy.convert("RGB").save(buf, format="JPEG", quality=85)
        b64_data = base64.b64encode(buf.getvalue()).decode("ascii")
        content = [
            {
                "type": "image",
                "source": {
                    "type":       "base64",
                    "media_type": "image/jpeg",
                    "data":       b64_data,
                },
            },
            {
                "type": "text",
                "text": prompt_text,
            },
        ]
        logging.info("  AI extraction: vision mode (image + OCR text)")
    else:
        content = prompt_text
        logging.info("  AI extraction: text-only mode (no image supplied)")

    response = ai_client.messages.create(
        model=_AI_MODEL,
        max_tokens=512,
        messages=[{"role": "user", "content": content}],
    )
    raw_json = response.content[0].text.strip()

    # Strip accidental markdown fences if the model adds them despite instructions
    if raw_json.startswith("```"):
        raw_json = re.sub(r"^```[a-z]*\n?", "", raw_json)
        raw_json = re.sub(r"\n?```$", "", raw_json)

    logging.info(f"  AI extraction result: {raw_json}")

    data = json.loads(raw_json)

    # Normalise: convert null → "" and coerce all values to str.
    # Log any field that was null before normalisation so we have a record.
    if ticket_type == "concrete":
        required_keys = {
            "ticket_number", "date", "supplier", "slump",
            "qty_delivered", "helper_notes", "qr_sticker_text",
        }
    else:
        required_keys = {
            "ticket_number", "date", "facility", "customer",
            "material", "net_tons", "qr_sticker_text",
        }
    for key in required_keys:
        val = data.get(key)
        if val is None:
            logging.warning(f"  AI returned null for field: {key}")
        data[key] = str(val).strip() if val is not None else ""

    return data


def _verify_material_match(anticipated: str, actual: str) -> bool:
    """Ask Claude API whether anticipated and actual materials refer to the same thing.

    Sends a short YES/NO prompt.  Returns True when the answer starts with YES.
    Raises on API error — callers catch and skip verification rather than flagging.
    """
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    prompt = (
        f"The anticipated material description is: {anticipated}\n"
        f"The material extracted from this ticket is: {actual}\n"
        "Do these refer to the same material? Consider common "
        "abbreviations, alternate names, and partial matches.\n"
        "Examples of matches: 'rip rap class 2' matches '808 "
        "Bradley Rock Rip Rap', 'boulders' matches 'Fish Boulder', "
        "'topsoil' matches 'Topsoil'\n"
        "Reply with only YES or NO."
    )
    response = client.messages.create(
        model=_AI_MODEL,
        max_tokens=10,
        messages=[{"role": "user", "content": prompt}],
    )
    answer = response.content[0].text.strip().upper()
    logging.info(
        f"  Material verification: anticipated={anticipated!r} "
        f"actual={actual!r} → {answer}"
    )
    return answer.startswith("YES")


# Matches a complete date with two separators: M/D/YYYY, MM/DD/YY, M-D-YYYY, etc.
# Intentionally uses only / and - so time strings like "10:18:01AM" are never matched.
_HEADER_DATE_RE   = re.compile(r'\b(\d{1,2}[/\-]\d{1,2}[/\-]\d{2,4})\b')
# Matches the first standalone 5-digit ticket number (10000-99999).
_HEADER_TICKET_RE = re.compile(r'\b([1-9]\d{4})\b')
# Lines containing weight-table keywords are skipped when scanning for the header.
_HEADER_SKIP_RE   = re.compile(
    r'\b(?:gross|tare|net|lbs?|pounds?|tons?|metric)\b', re.IGNORECASE
)


def _parse_ticket_header(text: str) -> tuple[str, str]:
    """Locate the header line and return (date_str, ticket_number).

    The Teevin Bros ticket layout is fixed:
      Line 1 — company name  (no date pattern → naturally skipped)
      Line 2 — date and ticket number, e.g. "8/6/2025 16800"
      Line 3 — time, e.g. "10:18:01AM"  (colon-separated → never matched as date)

    Strategy: scan non-empty lines from the top, skip weight-table rows, and
    return the first line that contains a complete date (two separators).  The
    ticket number is the first 5-digit number on that same line.

    Returns:
        (date_str, ticket_number) — either value is "" when not found on the
        header line.  Both being "" means the header is unreadable.
    """
    for line in text.splitlines():
        stripped = line.strip()
        if not stripped or _HEADER_SKIP_RE.search(stripped):
            continue
        date_m = _HEADER_DATE_RE.search(stripped)
        if not date_m:
            continue
        # Found the header line — extract ticket number from the same line.
        ticket_m = _HEADER_TICKET_RE.search(stripped)
        return date_m.group(1), (ticket_m.group(1) if ticket_m else "")

    return "", ""


def _ocr_date(text: str) -> str:
    """Return the date extracted from the ticket header line, or REVIEW_REQUIRED."""
    date_str, _ = _parse_ticket_header(text)
    return date_str if date_str else REVIEW_REQUIRED


def _extract_ticket_by_label(
    text: str, label: str, ticket_len: int
) -> str:
    """Search for *label* in OCR text and extract a *ticket_len*-digit ticket
    number from the same or following lines, with split-line reassembly.

    Handles three layouts:
      1. Label and number on the same line:  "Ticket No.: 20321064"
      2. Number split across two lines:      "Ticket No.: 20321" / "064"
      3. Label alone, number on next lines:  "Ticket No.:" / "20321" / "064"

    Reassembly rules (same as those in the region OCR fallback):
      - First part  : 4–6 digits
      - Second part : 2–4 digits
      - Combined    : exactly *ticket_len* digits, no leading zero

    Returns the ticket number string, or "" if not found.
    """
    lines      = text.splitlines()
    label_norm = label.rstrip(":. ").lower()

    for i, line in enumerate(lines):
        if label_norm not in line.lower():
            continue

        # ── Collect digit fragments starting from this line ────────────────────
        # Fragment 0: digits on the same line, after the label
        after = re.split(re.escape(label), line, maxsplit=1, flags=re.IGNORECASE)
        inline_digits = re.findall(r'\d+', after[1]) if len(after) > 1 else []

        # Fragments 1, 2: digits from the next two lines
        next_line_digits  = re.findall(r'\d+', lines[i + 1]) if i + 1 < len(lines) else []
        next2_line_digits = re.findall(r'\d+', lines[i + 2]) if i + 2 < len(lines) else []

        # ── Case 1: complete number on the same line ───────────────────────────
        for num in inline_digits:
            if len(num) == ticket_len and num[0] != '0':
                return num

        # ── Case 2 / 3: try reassembly ────────────────────────────────────────
        # Determine first and second fragments
        if inline_digits:
            first  = inline_digits[-1]               # last number on the label line
            nexts  = next_line_digits
        elif next_line_digits:
            # Complete number may be entirely on the next line
            for num in next_line_digits:
                if len(num) == ticket_len and num[0] != '0':
                    return num
            first  = next_line_digits[0]
            nexts  = next2_line_digits
        else:
            continue

        for second in (nexts[:1] if nexts else []):
            combined = first + second
            if (
                len(combined) == ticket_len
                and combined[0] != '0'
                and 4 <= len(first)  <= 6
                and 2 <= len(second) <= 4
            ):
                logging.info(
                    f"  Ticket number reassembled from label {label!r}: "
                    f"{first!r} + {second!r} → {combined!r}"
                )
                return combined

    return ""


def _ocr_ticket_number_from_region(
    image: Image.Image, region_config: dict
) -> str:
    """Crop the image to the configured region and extract the ticket number.

    The crop uses percentage coordinates from the profile.  Full-text OCR
    (--psm 6, no character whitelist) is run on the preprocessed crop.  The
    first *standalone* 5-digit number is returned — "standalone" meaning it
    is not surrounded by other digits or letters (i.e. not embedded in a
    longer number or word).  Any match whose position in the crop falls in
    the rightmost 15 % of the *full* page is skipped to avoid confusing a
    QR-code digit sequence with the ticket number.

    When OCR_DEBUG=1:
      - Logs the full page dimensions and crop box.
      - Logs the raw OCR text from the crop.
      - Saves the preprocessed crop to ocr_debug_topright.png and the raw
        OCR text to ocr_debug_topright.txt.

    Returns the first qualifying number of the configured length, or "" if none found.
    """
    # Build a regex for the profile-specified ticket number length (default 5).
    ticket_len = region_config.get("ticket_number_length", 5)
    _TICKET_RE = re.compile(
        rf'(?<![a-zA-Z\d])([1-9]\d{{{ticket_len - 1}}})(?![a-zA-Z\d])'
    )

    w, h = image.size

    if OCR_DEBUG:
        logging.info(f"  Page dimensions: {w}×{h} px")

    region = region_config.get("region", {})
    x1 = int(w * region.get("x_start_pct", 0.53))
    y1 = int(h * region.get("y_start_pct", 0.00))
    x2 = int(w * region.get("x_end_pct",   1.00))
    y2 = int(h * region.get("y_end_pct",   0.24))

    crop   = image.crop((x1, y1, x2, y2))
    prep   = preprocess_for_ocr(crop)
    config = region_config.get("tesseract_config", "--oem 3 --psm 6")
    raw    = pytesseract.image_to_string(prep, config=config).strip()

    if OCR_DEBUG:
        logging.info(
            f"  Region crop ({x1},{y1})–({x2},{y2})  raw OCR: {raw!r}"
        )
        try:
            prep.save("ocr_debug_topright.png")
            Path("ocr_debug_topright.txt").write_text(raw, encoding="utf-8")
            logging.info("  Saved ocr_debug_topright.png and ocr_debug_topright.txt")
        except Exception as dbg_exc:
            logging.warning(f"  Could not save region debug files: {dbg_exc}")
    else:
        logging.debug(f"  Region OCR raw: {raw!r}")

    # QR code sits in the rightmost 15 % of the full page.  The crop starts
    # at x1 pixels from the left edge of the full page, so the QR exclusion
    # zone begins at (0.85 * w - x1) pixels from the left edge of the crop.
    qr_zone_start_in_crop = max(0, int(w * 0.85) - x1)

    # Use pytesseract data to get per-word bounding boxes when available;
    # fall back to a plain regex scan if not.
    try:
        import pytesseract as _tess
        data = _tess.image_to_data(
            prep, config=config, output_type=_tess.Output.DICT
        )
        for i, word in enumerate(data["text"]):
            m = _TICKET_RE.fullmatch(word.strip()) if word.strip() else None
            if not m:
                continue
            word_x = data["left"][i]
            if word_x >= qr_zone_start_in_crop:
                if OCR_DEBUG:
                    logging.info(
                        f"  Skipping {word!r} at crop-x={word_x} "
                        f"(QR zone starts at {qr_zone_start_in_crop})"
                    )
                continue
            if OCR_DEBUG:
                logging.info(f"  Ticket number from region (bbox): {word.strip()!r}")
            return word.strip()
    except Exception:
        pass  # fall through to regex scan

    # Regex fallback — no positional info available
    for m in _TICKET_RE.finditer(raw):
        if OCR_DEBUG:
            logging.info(f"  Ticket number from region (regex): {m.group(1)!r}")
        return m.group(1)

    # Split-line reassembly fallback.  When ticket_len > 5 (e.g. Knife River
    # 8-digit tickets), OCR often breaks the number across two consecutive lines.
    # Look for two adjacent numeric fragments that concatenate to ticket_len digits.
    if ticket_len > 5:
        lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
        for i in range(len(lines) - 1):
            first_nums  = re.findall(r'\d+', lines[i])
            second_nums = re.findall(r'\d+', lines[i + 1])
            if not first_nums or not second_nums:
                continue
            first_part  = first_nums[-1]    # rightmost number on the line
            second_part = second_nums[0]    # leftmost number on the next line
            combined    = first_part + second_part
            if (
                len(combined) == ticket_len
                and combined[0] != '0'
                and 4 <= len(first_part)  <= 6
                and 2 <= len(second_part) <= 4
            ):
                logging.info(
                    f"  Ticket number reassembled from split lines "
                    f"(region): {first_part!r} + {second_part!r} → {combined!r}"
                )
                return combined

    return ""


def _ocr_ticket_number(
    text: str, images: list, profile: TicketProfile
) -> str:
    """Extract the ticket number using up to three strategies in priority order.

    1. Label-based search: if the profile defines ``ticket_number_label``
       (e.g. "Ticket No.:"), scan the full OCR text for that label and extract
       the following number, with split-line reassembly for multi-line reads.
    2. Region OCR: if ``method == "region_ocr"``, crop the configured corner
       region and run Tesseract, using the profile-specified ``ticket_number_length``
       (default 5).  Includes split-line reassembly for lengths > 5.
    3. Header-line fallback: parse the second non-empty line of the full OCR text
       for a 5-digit number (Teevin Bros style).
    """
    region_config = profile.layout.get("ticket_number_extraction", {})
    ticket_len    = region_config.get("ticket_number_length", 5)
    ticket_label  = region_config.get("ticket_number_label", "")

    # ── method: none — profile has no printed ticket number ──────────────────
    if region_config.get("method") == "none":
        return ""

    # ── Strategy 1: label-based search ──────────────────────────────────────
    if ticket_label:
        result = _extract_ticket_by_label(text, ticket_label, ticket_len)
        if result:
            logging.info(
                f"  Ticket number from label {ticket_label!r}: {result!r}"
            )
            return result

    # ── Strategy 2: region OCR ───────────────────────────────────────────────
    if region_config.get("method") == "region_ocr" and images:
        result = _ocr_ticket_number_from_region(images[0], region_config)
        if result:
            logging.info(f"  Ticket number from region OCR: {result!r}")
            return result
        logging.debug(
            f"  Region OCR returned no {ticket_len}-digit number "
            "— falling back to header line."
        )

    # ── Strategy 3: header-line fallback ─────────────────────────────────────
    _, ticket_number = _parse_ticket_header(text)
    return ticket_number


def _load_known_facilities() -> list:
    """Return the list of confirmed facility names from known_facilities.txt."""
    path = Path(KNOWN_FACILITIES_FILE)
    if not path.exists():
        return []
    return [l.strip() for l in path.read_text(encoding="utf-8").splitlines() if l.strip()]


def _save_known_facility(name: str) -> None:
    """Append a new facility name to known_facilities.txt if not already present."""
    if name in _load_known_facilities():
        return
    with open(KNOWN_FACILITIES_FILE, "a", encoding="utf-8") as fh:
        fh.write(name + "\n")
    logging.info(f"New facility saved to {KNOWN_FACILITIES_FILE}: {name!r}")


def _strip_facility_noise(candidate: str) -> str:
    """Filter OCR garbage tokens from a facility name candidate.

    Each token is tested against a set of noise rules.  Tokens that pass all
    rules are kept; tokens that fail any rule are dropped.  The leading numeric
    site code (e.g. "800") is always preserved.

    A token is kept when ALL of the following are true:
      1. It contains at least one alphabetic character  (drops "|", "?", "=")
      2. It has no characters other than letters, digits, hyphens, and
         apostrophes  (drops tokens with |, ?, @, #, etc.)
      3. It is not a single character  (drops stray "A", "I", etc.)
      4. It is not all-uppercase AND shorter than 4 letters
         (drops "OR", "Oo", "AR" but keeps "ROCK", "SAND", "MINE")
      5. It does not contain an uppercase letter after its first character
         when the token is lower-cased start  (drops "eRe", "oRe", "bRo")

    The leading numeric token (digits only, e.g. "800") is always kept
    regardless of the above rules.
    """
    tokens = candidate.split()
    clean: list[str] = []
    for i, tok in enumerate(tokens):
        # Always keep a leading numeric site code
        if i == 0 and tok.isdigit():
            clean.append(tok)
            continue
        # Rule 1: must contain at least one letter
        if not any(c.isalpha() for c in tok):
            continue
        # Rule 2: only letters, digits, hyphens, apostrophes allowed
        if re.search(r"[^a-zA-Z0-9\-']", tok):
            continue
        # Rule 3: must be longer than 1 character
        if len(tok) <= 1:
            continue
        # Rule 4: all-caps tokens must be 4+ letters  (drops "OR", "AR")
        if tok.isupper() and len(tok) < 4:
            continue
        # Rule 5: any 2-character token with an uppercase letter is noise
        #         (drops "Ar", "My", "Oo", "Or" — too short to be a real word
        #          in a facility name context)
        if len(tok) == 2 and any(c.isupper() for c in tok):
            continue
        # Rule 6: lowercase-start token with mid-uppercase (e.g. "eRe")
        if tok[0].islower() and any(c.isupper() for c in tok[1:]):
            continue
        clean.append(tok)
    return " ".join(clean)


def _fuzzy_correct_facility(raw: str) -> str:
    """Compare a raw facility name against known_facilities.txt and correct OCR errors.

    Strategy:
    - Split both the extracted value and each known facility into a leading
      numeric prefix ("800") and a text portion ("Bradley Quarry").
    - Fuzzy-match the text portions only so that different site numbers don't
      cause a false positive between two otherwise identical quarry names.
    - If the best text match scores >= 80%, return the numeric prefix from the
      extracted value combined with the corrected text from the known list.
    - If no match scores >= 80%, treat it as a new facility, save it, and
      return the raw value unchanged.
    """
    known = _load_known_facilities()

    def _split(s: str):
        """Split "800 Bradley Quarry" → ("800", "Bradley Quarry")."""
        m = re.match(r'^(\d[\d\s]*?)\s+([A-Za-z].*)$', s.strip())
        return (m.group(1).strip(), m.group(2).strip()) if m else ("", s.strip())

    raw_prefix, raw_text = _split(raw)

    # Build a map of text-portion → full known name
    known_text_map: dict[str, str] = {}
    for name in known:
        _, ktext = _split(name)
        known_text_map[ktext] = name

    matches = difflib.get_close_matches(
        raw_text, known_text_map.keys(), n=1, cutoff=0.80
    )

    if matches:
        best_text = matches[0]
        _, corrected_text = _split(known_text_map[best_text])
        corrected = f"{raw_prefix} {corrected_text}".strip() if raw_prefix else corrected_text
        if corrected != raw:
            logging.info(f"Facility fuzzy-corrected: {raw!r} → {corrected!r}")
        return corrected

    # New facility — strip trailing OCR bleed before persisting
    cleaned = _strip_facility_noise(raw)
    if cleaned != raw:
        logging.info(f"Facility trailing noise stripped: {raw!r} → {cleaned!r}")
    _save_known_facility(cleaned)
    return cleaned


def _ocr_facility(text: str, profile: TicketProfile) -> str:
    """Extract the source facility (quarry, mine, pit, plant) from OCR text.

    Primary label candidates come from profile.labels["facility"].
    Generic keyword fallback is retained for robustness.
    """
    company_name = re.compile(r'mj\s*hughes|hughes\s*construction', re.IGNORECASE)

    _STOP_RE = re.compile(
        r'\b(?:pays|customer|order|p\.o|product|carrier|vehicle|'
        r'weighmaster|gross|tare|net)\b',
        re.IGNORECASE,
    )

    def _clean(raw: str) -> str:
        cut = _STOP_RE.search(raw)
        return raw[:cut.start()].strip() if cut else raw.strip()

    # Build primary label patterns from the profile
    facility_labels = profile.labels.get("facility", ["Location"])
    if isinstance(facility_labels, str):
        facility_labels = [facility_labels]

    label_patterns = [
        rf'{re.escape(lbl.rstrip(":. "))}[.:\s]+([^\n]{{3,80}})'
        for lbl in facility_labels
    ]
    # Generic fallback labels (not supplier-specific)
    label_patterns.append(
        r'(?:shipped\s*from|from|source|origin|facility|pit|quarry|mine|'
        r'plant|producer|supplier|sold\s*by|vendor)[.:\s]+([^\n]{3,80})'
    )

    for pattern in label_patterns:
        for m in re.finditer(pattern, text, re.IGNORECASE):
            candidate = _clean(m.group(1))
            if not candidate or company_name.search(candidate):
                continue
            if len(candidate) >= 3:
                return _fuzzy_correct_facility(_apply_ocr_corrections(candidate, profile))

    # Fallback: line containing a facility keyword
    facility_keywords = re.compile(
        r'\b(?:quarry|mine|pit|plant|aggregate|gravel|sand|rock|stone)\b',
        re.IGNORECASE,
    )
    for line in text.splitlines():
        line = line.strip()
        if not line or company_name.search(line):
            continue
        if facility_keywords.search(line):
            return _fuzzy_correct_facility(_apply_ocr_corrections(_clean(line), profile))

    return ""


def _apply_ocr_corrections(value: str, profile: TicketProfile) -> str:
    """Apply OCR misread corrections defined in the supplier profile."""
    for wrong, right in profile.ocr_corrections.items():
        value = value.replace(wrong, right)
    return value


def _ocr_material(text: str, profile: TicketProfile) -> str:
    """Extract the material / product description from OCR text.

    Primary label candidates come from profile.labels["material"].
    Generic fallback labels are appended after profile labels.
    """
    material_labels = profile.labels.get("material", ["Product"])
    if isinstance(material_labels, str):
        material_labels = [material_labels]

    patterns = [
        rf'{re.escape(lbl.rstrip(":. "))}[:\s]+([^\n]{{3,80}})'
        for lbl in material_labels
    ]
    patterns += [
        r'(?:material(?:\s+type)?|description|item|commodity|aggregate)[:\s]+([^\n]{3,80})',
        r'(?:type|grade)[:\s]+([^\n]{3,80})',
    ]

    def _clean_material(raw: str) -> str:
        raw = _apply_ocr_corrections(raw.strip(), profile)
        # Strip a leading product code prefix in two forms:
        #   "—-801 3/4-0"" — special chars before the code (original pattern)
        #   "801 3/4-0""   — plain code with no leading special chars,
        #                    but only when the code is followed by a non-letter
        #                    so "808 Bradley Rock Rip Rap" is left untouched.
        raw = re.sub(r'^[-—–=!\s]*\d{2,4}\s+(?=[^a-zA-Z])', '', raw)
        # Truncate at the first digit that follows a non-digit word character
        cut = re.search(r'(?<=\D)\s+\d', raw)
        if cut:
            raw = raw[:cut.start()]
        # Strip trailing noise symbols and unit words.
        # NOTE: double-quote (") is intentionally excluded from the character
        # class — it is part of aggregate sizing notation (e.g. '3/4-0"'
        # meaning 0-inch minus) and must be preserved.
        raw = re.sub(
            r"[\s=*/@'!|]+(?:tons?|lbs?|pounds?|kg)?\s*$",
            '',
            raw,
            flags=re.IGNORECASE,
        ).strip()
        return raw

    for pattern in patterns:
        m = re.search(pattern, text, re.IGNORECASE)
        if m:
            result = _clean_material(m.group(1))
            if result:
                return result

    # Last-resort fallback: look for aggregate-style sizing patterns (e.g. "3/4-0"")
    # in the bottom third of the OCR text where the product line usually appears.
    lines = text.splitlines()
    bottom_lines = lines[len(lines) * 2 // 3:]
    for line in bottom_lines:
        m = re.search(r'(\d+/\d+[-\d"]*(?:\s+\w+)*)', line)
        if m:
            candidate = _clean_material(m.group(1))
            if candidate:
                logging.debug(f"  Material from bottom-third fallback: {candidate!r}")
                return candidate

    return ""


def _ocr_net_tons(text: str, profile: TicketProfile) -> str:
    """Extract net quantity in tons from OCR text.

    Uses profile.labels["net_row"] as the row identifier and
    profile.weight_table["tons_column_index"] (0-based) to select the
    correct number from the weight row (default: index 1 = second number).
    """
    net_label  = profile.labels.get("net_row", "Net")
    tons_col   = profile.weight_table.get("tons_column_index", 1)  # 0-based
    net_escape = re.escape(net_label)

    # Primary: Net row with two numbers — pick column by tons_column_index
    net_re = re.compile(
        rf'\b{net_escape}\b\s+(\d[\d,]*\.?\d*)\s+(\d[\d,]*\.?\d*)',
        re.IGNORECASE,
    )
    for line in text.splitlines():
        m = net_re.match(line.strip())
        if m:
            # group(1) = index 0 (pounds), group(2) = index 1 (tons)
            group = tons_col + 1
            try:
                return m.group(group).replace(",", "").strip()
            except IndexError:
                return m.group(2).replace(",", "").strip()

    # Fallback: explicit tons label
    fallback_patterns = [
        rf'{net_escape}\s*tons?[:\s]+(\d[\d,]*\.?\d*)',
        r'\btons?[:\s]+(\d[\d,]*\.?\d*)',
        r'(\d[\d,]*\.\d{1,4})\s*tons?\b',
    ]
    lbs_label = re.compile(r'\blbs?\b|\bpounds?\b', re.IGNORECASE)
    for pattern in fallback_patterns:
        for line in text.splitlines():
            if lbs_label.search(line):
                continue
            m = re.search(pattern, line, re.IGNORECASE)
            if m:
                return m.group(1).replace(",", "").strip()

    return ""


def _ocr_customer(text: str, profile: TicketProfile) -> str:
    """Extract the customer / company name from OCR text.

    Label candidates come from profile.labels["customer"].
    The value ends before "Pounds" (weight column header) or end of line.
    """
    customer_labels = profile.labels.get("customer", ["Customer"])
    if isinstance(customer_labels, str):
        customer_labels = [customer_labels]

    for lbl in customer_labels:
        clean_lbl = re.escape(lbl.rstrip(":. "))
        m = re.search(
            rf'{clean_lbl}[.:\s]+(.+?)(?=\s*pounds|\n|$)',
            text,
            re.IGNORECASE,
        )
        if m:
            return m.group(1).strip()
    return ""


# ============================================================
# EXCEL OPERATIONS
# ============================================================
def _load_toc_materials() -> dict:
    """Read the TOC tab and return a QR-code → anticipated-material lookup dict.

    Scans every cell in the TOC tab for a value matching the QR code pattern
    (XXXX-YYYY-CCCCCC-CC).  The anticipated material for each row is read from
    column K (11th column, 0-based index 10).  Blank values and "??" are skipped.

    Called once at startup; the returned dict is passed through the call chain
    so it is available for every ticket processed in the run.

    Returns an empty dict if the workbook is missing, the TOC tab does not
    exist, or no QR codes can be found in the sheet.
    """
    excel_path = Path(EXCEL_FILE)
    if not excel_path.exists():
        logging.warning("_load_toc_materials: workbook not found — skipping.")
        return {}

    try:
        workbook = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    except Exception as exc:
        logging.warning(f"_load_toc_materials: could not open workbook: {exc}")
        return {}

    if _TOC_TAB_NAME not in workbook.sheetnames:
        logging.warning(f"_load_toc_materials: tab '{_TOC_TAB_NAME}' not found.")
        workbook.close()
        return {}

    toc    = workbook[_TOC_TAB_NAME]
    lookup: dict[str, str] = {}

    for row in toc.iter_rows(values_only=True):
        # Scan every cell in this row for a QR code match
        qr_key: Optional[str] = None
        for cell_val in row:
            if cell_val is None:
                continue
            m = QR_PATTERN.search(str(cell_val).strip())
            if m:
                qr_key = m.group(0).upper()
                break

        if qr_key is None:
            continue

        # Column K is the 11th column — index 10 when 0-based
        mat_raw = row[10] if len(row) > 10 else None
        if mat_raw is None:
            continue
        anticipated = str(mat_raw).strip()
        if anticipated and anticipated.lower() not in ("??", "n/a", ""):
            lookup[qr_key] = anticipated

    workbook.close()
    logging.info(f"TOC materials loaded: {len(lookup)} entry/entries.")
    return lookup


def _append_note(existing: str, new: str) -> str:
    """Append *new* to *existing* using ' | ' as the separator.

    Returns *new* alone when *existing* is empty, *existing* alone when
    *new* is empty, and the joined string when both are non-empty.
    """
    if existing and new:
        return f"{existing} | {new}"
    return existing or new


def _find_duplicate_ticket(
    ticket_number: str,
    excel_path: str = None,
    col_ticket_num: int = None,
    data_start_row: int = None,
) -> Optional[str]:
    """Search all tabs in an Excel workbook for an existing ticket number.

    *excel_path* defaults to the rock tracker (EXCEL_FILE).
    *col_ticket_num* defaults to rock column E (_COL_TICKET_NUM).
    *data_start_row* defaults to rock data start row (_DATA_START_ROW).

    Scans every non-TOC sheet from *data_start_row* and checks the specified
    column.  Returns the tab name where the ticket is found, or None.
    """
    path = Path(excel_path) if excel_path else Path(EXCEL_FILE)
    if col_ticket_num is None:
        col_ticket_num = _COL_TICKET_NUM
    if data_start_row is None:
        data_start_row = _DATA_START_ROW

    if not path.exists() or not ticket_number:
        return None

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in workbook.sheetnames:
            if sheet_name == _TOC_TAB_NAME:
                continue
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
                if len(row) >= col_ticket_num:
                    cell_val = row[col_ticket_num - 1]   # iter_rows is 0-based via values_only
                    if cell_val and str(cell_val).strip() == str(ticket_number).strip():
                        return sheet_name
    finally:
        workbook.close()

    return None


def _read_existing_ticket_numbers(
    excel_path: str = None,
    col_ticket_num: int = None,
    data_start_row: int = None,
) -> frozenset:
    """Collect all ticket numbers already written in an Excel workbook.

    Opens the file read-only and scans every non-TOC sheet from
    *data_start_row*, returning a frozenset of stripped ticket-number strings.

    Called once per email before processing begins.  The resulting snapshot
    lets the retry logic skip tickets that were successfully written in an
    earlier attempt so they are not written a second time.

    Returns an empty frozenset when the file does not exist, is unreadable,
    or contains no ticket numbers.
    """
    path = Path(excel_path) if excel_path else Path(EXCEL_FILE)
    if col_ticket_num is None:
        col_ticket_num = _COL_TICKET_NUM
    if data_start_row is None:
        data_start_row = _DATA_START_ROW

    if not path.exists():
        return frozenset()

    result: set = set()
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet_name in workbook.sheetnames:
                if sheet_name == _TOC_TAB_NAME:
                    continue
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
                    if len(row) >= col_ticket_num:
                        cell_val = row[col_ticket_num - 1]
                        if cell_val:
                            result.add(str(cell_val).strip())
        finally:
            workbook.close()
    except Exception as exc:
        logging.warning(
            f"  Could not read existing ticket numbers from {path.name}: {exc}"
        )

    return frozenset(result)


def _pick_tab_with_ai(
    matching_tabs: list,
    qr_raw: str,
    sticker_text: str,
    material: str,
    toc_materials: dict,
) -> Optional[object]:
    """Ask Claude to choose among multiple tabs that all share the same QR suffix.

    Reads B4 (QR Description) and B5 (QR Short Desc.) from each candidate tab
    and presents them alongside the ticket's sticker text, extracted material,
    and TOC anticipated material.

    Returns the selected worksheet, or None when the AI answer cannot be parsed
    as a valid tab number.
    """
    if not ANTHROPIC_API_KEY:
        return None

    ai_client  = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
    tab_lines: list[str] = []

    for i, ws in enumerate(matching_tabs, start=1):
        desc       = str(ws["B4"].value or "").strip()
        short_desc = str(ws["B5"].value or "").strip()
        anticipated = toc_materials.get(qr_raw.strip().upper(), "")
        tab_lines.append(
            f"Tab {i} (name: {ws.title!r}): "
            f"QR Description={desc!r}, "
            f"Short Desc={short_desc!r}, "
            f"Anticipated Material={anticipated!r}"
        )

    prompt = (
        f"A ticket has QR code sticker text: {sticker_text!r}\n"
        f"and extracted material: {material!r}\n\n"
        f"There are {len(matching_tabs)} possible tabs this ticket could belong to:\n"
        + "\n".join(tab_lines)
        + "\n\nWhich tab number does this ticket belong to?\n"
        "Reply with only the tab number (1, 2, etc.) and nothing else."
    )

    try:
        response = ai_client.messages.create(
            model=_AI_MODEL,
            max_tokens=10,
            messages=[{"role": "user", "content": prompt}],
        )
        answer = response.content[0].text.strip()
        logging.info(
            f"  Tab disambiguation: Claude answered {answer!r} "
            f"for QR {qr_raw!r} "
            f"(candidates: {[ws.title for ws in matching_tabs]})"
        )
        idx = int(answer) - 1   # 1-based → 0-based
        if 0 <= idx < len(matching_tabs):
            return matching_tabs[idx]
    except Exception as exc:
        logging.warning(f"  Tab disambiguation AI call failed: {exc}")

    return None


def _find_tab_for_qr(
    workbook,
    qr_raw: str,
    ticket_data: Optional[dict] = None,
    toc_materials: Optional[dict] = None,
) -> Optional[object]:
    """Find the pre-built tab whose B3 cost code matches the QR code suffix.

    Each data tab has a cost code in cell B3 (e.g. "0010-015436-10").
    The QR code string looks like "2601-0010-015436-10".  A tab matches
    when qr_raw ends with the value in its B3 cell (case-insensitive).

    Single match:
        Returns the matching worksheet immediately.

    Multiple matches:
        Calls Claude with the ticket's qr_sticker_text, extracted material,
        and each tab's B4/B5 descriptions to pick the right one.
        Logs the AI selection.
        Raises AmbiguousTabError when AI cannot determine — the caller should
        flag the ticket for manual review.

    No match:
        Returns None.

    All tabs are pre-built by humans — this function never creates new tabs.
    Skips the TOC tab.
    """
    qr_upper      = qr_raw.strip().upper()
    matching_tabs = []

    for ws in workbook.worksheets:
        if ws.title == _TOC_TAB_NAME:
            continue
        b3 = ws["B3"].value
        if b3 is None:
            continue
        b3_str = str(b3).strip().upper()
        if b3_str and qr_upper.endswith(b3_str):
            matching_tabs.append(ws)

    if not matching_tabs:
        return None

    if len(matching_tabs) == 1:
        return matching_tabs[0]

    # Multiple tabs share the same cost-code suffix — need AI disambiguation
    logging.info(
        f"  Multiple tabs matched QR {qr_raw!r}: "
        + ", ".join(repr(ws.title) for ws in matching_tabs)
        + " — asking Claude to disambiguate"
    )

    td   = ticket_data   or {}
    toc  = toc_materials or {}
    sticker_text = td.get("qr_sticker_text", "") or ""
    material     = td.get("material",        "") or ""

    selected = _pick_tab_with_ai(matching_tabs, qr_raw, sticker_text, material, toc)

    if selected is not None:
        logging.info(
            f"  Multiple tabs matched QR {qr_raw!r} — AI selected tab "
            f"{selected.title!r} based on material {material!r}"
        )
        return selected

    raise AmbiguousTabError(
        f"REVIEW REQUIRED: Multiple tabs match QR code {qr_raw!r} and material "
        "could not be matched. Please assign manually."
    )


def _find_concrete_tab_for_qr(workbook, qr_raw: str):
    """Find the concrete tracker tab for *qr_raw* via the TOC.

    Scans column B of the TOC tab for an exact match with *qr_raw*.
    The tab whose title equals the matched row number (as a string) is returned.

    Example: QR code "2601-0180-313713-99" in TOC row 2 → returns sheet "2".

    Returns the matching worksheet, or None when:
        • the TOC tab is missing
        • the QR code is not in column B of the TOC
        • the expected tab (row-number name) does not exist yet
    In all failure cases a descriptive warning is logged.
    """
    if _TOC_TAB_NAME not in workbook.sheetnames:
        logging.warning(
            f"Concrete TOC tab '{_TOC_TAB_NAME}' not found in workbook."
        )
        return None

    toc      = workbook[_TOC_TAB_NAME]
    qr_upper = qr_raw.strip().upper()

    for row_cells in toc.iter_rows(min_col=2, max_col=2):
        cell = row_cells[0]
        if cell.value and str(cell.value).strip().upper() == qr_upper:
            tab_name = str(cell.row)
            if tab_name in workbook.sheetnames:
                return workbook[tab_name]
            logging.warning(
                f"WARNING: QR code {qr_raw!r} found in concrete TOC row {cell.row}, "
                f"but tab '{tab_name}' does not exist. "
                "Please create the corresponding tab."
            )
            return None

    logging.warning(
        f"WARNING: QR code {qr_raw!r} not found in concrete TOC. "
        "Please add it to the TOC and create the corresponding tab."
    )
    return None


def _next_data_row(
    sheet,
    col_ticket_num: int = None,
    data_start_row: int = None,
) -> int:
    """Return the index of the first empty row in the data area.

    Uses *col_ticket_num* (default: rock tracker column E) as the sentinel —
    a row is considered occupied when that cell has a value.
    """
    if col_ticket_num is None:
        col_ticket_num = _COL_TICKET_NUM
    if data_start_row is None:
        data_start_row = _DATA_START_ROW
    for row_idx in range(data_start_row, sheet.max_row + 2):
        if sheet.cell(row=row_idx, column=col_ticket_num).value is None:
            return row_idx
    return data_start_row


def write_to_excel(qr_data: dict, ticket_data: dict, notes: str = "", toc_materials: Optional[dict] = None, excel_path: str = None) -> None:
    """Write one ticket row into the rock Excel tracker for a job.

    - Opens the existing workbook (raises FileNotFoundError if absent).
    - Finds the pre-built tab whose B3 cost code matches the QR code suffix.
    - If no tab matches, logs a warning and returns without writing.
    - Finds the first empty data row (row 9+) and writes to the mapped columns.
    - Does NOT touch rows 1-8 (pre-built header), the TOC tab, or columns
      G, H, I, K (those are human-filled).

    *excel_path* — explicit local path to the workbook.  Defaults to EXCEL_FILE
    (the primary-job tracker) when not provided.

    Column mapping:
        A (_COL_TICKET_DATE) — ticket date
        B (_COL_LOGGED_DATE) — today's date (script run date)
        C (_COL_FACILITY)    — facility / quarry
        D (_COL_MATERIAL)    — material description
        E (_COL_TICKET_NUM)  — 5-digit ticket number
        F (_COL_QTY_TN)      — net tons
        J (_COL_NOTES)       — helper notes / flags

    Highlighting:
        notes="DUPLICATE TICKET" → all written cells orange
        empty required field     → that cell yellow + warning logged
    """
    excel_path = Path(excel_path) if excel_path else Path(EXCEL_FILE)
    if not excel_path.exists():
        raise FileNotFoundError(
            f"Excel workbook not found: {excel_path}\n"
            "Ensure the tracker has been downloaded from SharePoint."
        )

    workbook = openpyxl.load_workbook(excel_path)
    qr_raw   = qr_data.get("raw", "")
    sheet    = _find_tab_for_qr(workbook, qr_raw, ticket_data, toc_materials)

    if sheet is None:
        logging.warning(
            f"WARNING: No matching tab found for QR code {qr_raw!r}. "
            "Please create a tab for this cost code manually."
        )
        workbook.close()
        return

    tab_name = sheet.title
    row_idx  = _next_data_row(sheet)
    today    = datetime.now().strftime("%m/%d/%Y")

    # Write to the mapped columns only
    sheet.cell(row=row_idx, column=_COL_TICKET_DATE).value = ticket_data.get("date", "")
    sheet.cell(row=row_idx, column=_COL_LOGGED_DATE).value = today
    sheet.cell(row=row_idx, column=_COL_FACILITY).value    = ticket_data.get("facility", "")
    sheet.cell(row=row_idx, column=_COL_MATERIAL).value    = ticket_data.get("material", "")
    sheet.cell(row=row_idx, column=_COL_TICKET_NUM).value  = ticket_data.get("ticket_number", "")
    sheet.cell(row=row_idx, column=_COL_QTY_TN).value      = ticket_data.get("net_tons", "")
    _td_notes      = ticket_data.get("helper_notes", "")
    combined_notes = _append_note(_td_notes, notes) if notes else _td_notes
    if combined_notes:
        sheet.cell(row=row_idx, column=_COL_NOTES).value = combined_notes

    # Highlighting
    _orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    _yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    if "DUPLICATE" in notes:
        for col in _WRITTEN_COLS:
            sheet.cell(row=row_idx, column=col).fill = _orange
    else:
        # Yellow highlight for any required field that came back empty
        ticket_number = ticket_data.get("ticket_number", "N/A")
        required = {
            _COL_TICKET_DATE: "Ticket Date",
            _COL_FACILITY:    "Facility",
            _COL_MATERIAL:    "Material",
            _COL_TICKET_NUM:  "Ticket #",
            _COL_QTY_TN:      "Quantity TN",
        }
        for col, field_name in required.items():
            if not sheet.cell(row=row_idx, column=col).value:
                sheet.cell(row=row_idx, column=col).fill = _yellow
                logging.warning(f"  WARNING: Empty field '{field_name}' for ticket {ticket_number}")
                log_error(f"ticket {ticket_number}", f"Empty field: {field_name}")

        # Yellow-highlight col J when a material mismatch was flagged
        if "Correct material?" in notes:
            sheet.cell(row=row_idx, column=_COL_NOTES).fill = _yellow

    workbook.save(excel_path)
    logging.info(
        f"Excel: wrote ticket {ticket_data.get('ticket_number', 'N/A')} "
        f"to tab {tab_name!r} row {row_idx}."
    )


def write_to_excel_concrete(
    qr_data: dict,
    ticket_data: dict,
    notes: str = "",
    toc_materials: Optional[dict] = None,
    excel_path: str = None,
) -> None:
    """Write one concrete ticket row into the concrete Excel tracker for a job.

    *excel_path* — explicit local path to the workbook.  Defaults to
    EXCEL_FILE_CONCRETE (the primary-job tracker) when not provided.

    Column mapping (data starts at row 5):
        A (_CONCRETE_COL_TICKET_DATE) — ticket date
        B (_CONCRETE_COL_LOGGED_DATE) — today's date
        C (_CONCRETE_COL_SUPPLIER)    — supplier name
        D (_CONCRETE_COL_SLUMP)       — slump on arrival
        E (_CONCRETE_COL_QTY_CY)      — qty delivered (CY)
        F (_CONCRETE_COL_TICKET_NUM)  — 7-digit ticket number
        G (_CONCRETE_COL_NOTES)       — helper notes (material description)
        H                             — left blank for human notes

    Tab matching: looks up the QR code in column B of the concrete TOC tab
    and writes to the tab whose title equals that TOC row number (e.g. "2").
    """
    excel_path = Path(excel_path) if excel_path else Path(EXCEL_FILE_CONCRETE)
    if not excel_path.exists():
        raise FileNotFoundError(
            f"Concrete Excel workbook not found: {excel_path}\n"
            "Ensure the tracker has been downloaded from SharePoint."
        )

    workbook = openpyxl.load_workbook(excel_path)
    qr_raw   = qr_data.get("raw", "")
    sheet    = _find_concrete_tab_for_qr(workbook, qr_raw)

    if sheet is None:
        # Warning already logged inside _find_concrete_tab_for_qr
        workbook.close()
        return

    tab_name = sheet.title
    row_idx  = _next_data_row(
        sheet,
        col_ticket_num=_CONCRETE_COL_TICKET_NUM,
        data_start_row=_CONCRETE_DATA_START_ROW,
    )
    today = datetime.now().strftime("%m/%d/%Y")

    # Helper notes: product description from ticket; merge with any flag note
    helper_from_ticket = (
        ticket_data.get("helper_notes", "") or ticket_data.get("material", "")
    )
    combined_notes = _append_note(helper_from_ticket, notes) if notes else helper_from_ticket

    sheet.cell(row=row_idx, column=_CONCRETE_COL_TICKET_DATE).value = ticket_data.get("date", "")
    sheet.cell(row=row_idx, column=_CONCRETE_COL_LOGGED_DATE).value = today
    sheet.cell(row=row_idx, column=_CONCRETE_COL_SUPPLIER).value    = (
        ticket_data.get("supplier", "") or "CalPortland"
    )
    sheet.cell(row=row_idx, column=_CONCRETE_COL_SLUMP).value       = ticket_data.get("slump", "")
    sheet.cell(row=row_idx, column=_CONCRETE_COL_QTY_CY).value      = ticket_data.get("qty_delivered", "")
    sheet.cell(row=row_idx, column=_CONCRETE_COL_TICKET_NUM).value   = ticket_data.get("ticket_number", "")
    if combined_notes:
        sheet.cell(row=row_idx, column=_CONCRETE_COL_NOTES).value = combined_notes

    # Highlighting
    _orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    _yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    if "DUPLICATE" in notes:
        for col in _CONCRETE_WRITTEN_COLS:
            sheet.cell(row=row_idx, column=col).fill = _orange
    else:
        ticket_number = ticket_data.get("ticket_number", "N/A")
        # All AI-filled columns A–G except B (Logged Date, always set by us)
        required = {
            _CONCRETE_COL_TICKET_DATE: "Ticket Date",
            _CONCRETE_COL_SUPPLIER:    "Supplier",
            _CONCRETE_COL_SLUMP:       "Slump",
            _CONCRETE_COL_QTY_CY:      "Qty Delivered (CY)",
            _CONCRETE_COL_TICKET_NUM:  "Ticket #",
            _CONCRETE_COL_NOTES:       "Helper Notes",
        }
        for col, field_name in required.items():
            val = sheet.cell(row=row_idx, column=col).value
            if not val or str(val).strip().upper() == "NOT FOUND":
                sheet.cell(row=row_idx, column=col).fill = _yellow
                logging.warning(
                    f"  WARNING: Empty field '{field_name}' for ticket {ticket_number}"
                )
                log_error(f"ticket {ticket_number}", f"Empty field: {field_name}")

    workbook.save(excel_path)
    logging.info(
        f"Excel (concrete): wrote ticket {ticket_data.get('ticket_number', 'N/A')} "
        f"to tab {tab_name!r} row {row_idx}."
    )


# ============================================================
# SHAREPOINT UPLOAD
# ============================================================
def _ensure_sharepoint_folder(
    client: GraphClient,
    site_id: str,
    drive_id: str,
    folder_path: str,
) -> bool:
    """Ensure a folder exists in SharePoint at the given drive-relative path.

    Makes no change if the folder already exists.  Creates the immediate
    folder (parent must already exist) when a 404 is returned.  Logs on
    creation.  Treats a 409 on creation as "already exists" (race-safe).

    Returns:
        True  — folder was just created.
        False — folder already existed (or a 409 race resolved to exists).

    Raises on any other HTTP error.
    """
    check_url = (
        f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}"
        f"/root:/{folder_path}"
    )
    try:
        client.get(check_url)
        return False   # folder already exists
    except requests.exceptions.HTTPError as exc:
        if exc.response is not None and exc.response.status_code == 404:
            pass  # does not exist — fall through to create
        else:
            raise

    # Split into parent path and the new folder name
    if "/" in folder_path:
        parent_path, folder_name = folder_path.rsplit("/", 1)
        create_url = (
            f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}"
            f"/root:/{parent_path}:/children"
        )
    else:
        folder_name = folder_path
        create_url  = (
            f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}/root/children"
        )

    try:
        client.post(create_url, json={
            "name": folder_name,
            "folder": {},
            "@microsoft.graph.conflictBehavior": "fail",
        })
        logging.info(f"Created folder: {folder_path}")
        return True
    except requests.exceptions.HTTPError as exc:
        # 409 = folder was created between our check and our create (race) — fine
        if exc.response is not None and exc.response.status_code == 409:
            logging.debug(f"Folder already exists (concurrent creation): {folder_path}")
            return False
        else:
            raise


def _ensure_job_folder_structure(
    client: GraphClient,
    site_id: str,
    drive_id: str,
    job_number: str,
) -> None:
    """Ensure the standard two-level folder layout exists for a job.

    Creates these paths if they do not already exist:
        /MJHughes OPEN JOBS/{job_number}/
        /MJHughes OPEN JOBS/{job_number}/Ticket Scans/

    Logs a single "Created folder structure for new job: {job_number}" message
    when either folder had to be created.  Silent when both already exist.
    """
    job_path   = f"{SHAREPOINT_FOLDER}/{job_number}"
    scans_path = f"{job_path}/Ticket Scans"

    job_created   = _ensure_sharepoint_folder(client, site_id, drive_id, job_path)
    scans_created = _ensure_sharepoint_folder(client, site_id, drive_id, scans_path)

    if job_created or scans_created:
        logging.info(f"Created folder structure for new job: {job_number}")


def _sanitize_sharepoint_folder_name(name: str) -> str:
    """Return *name* safe for use as a SharePoint folder name.

    Replaces forward-slashes with hyphens and removes characters that
    SharePoint prohibits in folder names.
    """
    name = name.replace("/", "-")
    for ch in '\\*?"<>|:#':
        name = name.replace(ch, "")
    return name.strip()


def upload_to_sharepoint(
    client: GraphClient,
    pdf_bytes: bytes,
    job_number: str,
    ticket_numbers: list[str],
    subfolder: str = "",
    material_subfolder: str = "",
) -> str:
    """Upload a ticket PDF to the job's Ticket Scans folder on SharePoint.

    Destination path (no subfolder):
        /Shared Documents/MJHughes OPEN JOBS/{job_number}/Ticket Scans/{filename}

    Destination path (type subfolder only, e.g. "Rock"):
        /Shared Documents/MJHughes OPEN JOBS/{job_number}/Ticket Scans/Rock/{filename}

    Destination path (type + material subfolder, e.g. "Rock" / "Class 2 RipRap"):
        /Shared Documents/MJHughes OPEN JOBS/{job_number}/Ticket Scans/Rock/Class 2 RipRap/{filename}

    All missing folders are created automatically.

    File naming convention: "Tickets, [TicketNumber1], [TicketNumber2].pdf"
    Single ticket example : Tickets, 16800.pdf
    Multi-ticket example  : Tickets, 16800, 16801, 16802.pdf

    Returns:
        The SharePoint web URL of the uploaded file.
    """
    filename = "Tickets, " + ", ".join(ticket_numbers) + ".pdf"

    site_id  = _sharepoint_site_id(client)
    drive_id = _sharepoint_drive_id(client, site_id)

    # Ensure /MJHughes OPEN JOBS/{job_number}/Ticket Scans/ exists
    _ensure_job_folder_structure(client, site_id, drive_id, job_number)

    scans_path = f"{SHAREPOINT_FOLDER}/{job_number}/Ticket Scans"

    if subfolder:
        type_path = f"{scans_path}/{subfolder}"
        created   = _ensure_sharepoint_folder(client, site_id, drive_id, type_path)
        if created:
            logging.info(f"Created subfolder: Ticket Scans/{subfolder}")
        scans_path = type_path

    if material_subfolder:
        mat_path   = f"{scans_path}/{material_subfolder}"
        rel_path   = f"Ticket Scans/{subfolder}/{material_subfolder}" if subfolder else f"Ticket Scans/{material_subfolder}"
        created    = _ensure_sharepoint_folder(client, site_id, drive_id, mat_path)
        if created:
            logging.info(f"Created subfolder: {rel_path}")
        else:
            logging.info(f"Folder already exists: {rel_path}")
        scans_path = mat_path

    upload_url = (
        f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}"
        f"/root:/{scans_path}/{filename}:/content"
    )
    response = client.put_binary(upload_url, pdf_bytes, content_type="application/pdf")
    web_url  = response.json().get("webUrl", "(unknown)")

    logging.info(f"Uploaded PDF to SharePoint: {scans_path}/{filename}")
    return web_url


def _sharepoint_site_id(client: GraphClient) -> str:
    """Retrieve the Graph site ID for the configured SharePoint hostname.

    Tries four different discovery strategies in order and stops at the first
    that returns an HTTP 200.  Every attempt logs the exact URL, status code,
    and full raw response body so connection/permission errors can be diagnosed.
    """
    token   = client.access_token
    headers = {
        "Authorization":    f"Bearer {token}",
        "ConsistencyLevel": "eventual",
    }

    # ── Attempt 1: standard path-based site lookup ────────────────────────────
    url1 = f"{GRAPH_BASE}/sites/{SHAREPOINT_HOST}:/"
    logging.info(f"SharePoint site discovery attempt 1: GET {url1}")
    resp1 = requests.get(url1, headers=headers)
    logging.info(f"  → HTTP {resp1.status_code}  body: {resp1.text}")
    if resp1.status_code == 200:
        site_id = resp1.json()["id"]
        logging.info(f"SharePoint site found via: path lookup ({url1})")
        return site_id

    # ── Attempt 2: hostname-only lookup ───────────────────────────────────────
    url2 = f"{GRAPH_BASE}/sites/{SHAREPOINT_HOST}"
    logging.info(f"SharePoint site discovery attempt 2: GET {url2}")
    resp2 = requests.get(url2, headers=headers)
    logging.info(f"  → HTTP {resp2.status_code}  body: {resp2.text}")
    if resp2.status_code == 200:
        site_id = resp2.json()["id"]
        logging.info(f"SharePoint site found via: hostname lookup ({url2})")
        return site_id

    # ── Attempt 3: root site ──────────────────────────────────────────────────
    url3 = f"{GRAPH_BASE}/sites/root"
    logging.info(f"SharePoint site discovery attempt 3: GET {url3}")
    resp3 = requests.get(url3, headers=headers)
    logging.info(f"  → HTTP {resp3.status_code}  body: {resp3.text}")
    if resp3.status_code == 200:
        site_id = resp3.json()["id"]
        logging.info(f"SharePoint site found via: root site ({url3})")
        return site_id

    # ── Attempt 4: site search ────────────────────────────────────────────────
    url4 = f"{GRAPH_BASE}/sites?search=*"
    logging.info(f"SharePoint site discovery attempt 4: GET {url4}")
    resp4 = requests.get(url4, headers=headers)
    logging.info(f"  → HTTP {resp4.status_code}  body: {resp4.text}")
    if resp4.status_code == 200:
        sites = resp4.json().get("value", [])
        host_lower = SHAREPOINT_HOST.lower()
        for site in sites:
            web_url = (site.get("webUrl") or "").lower()
            name    = (site.get("name")   or "").lower()
            if host_lower.split(".")[0] in web_url or host_lower.split(".")[0] in name:
                site_id = site["id"]
                logging.info(
                    f"SharePoint site found via: search (matched {site.get('webUrl')!r})"
                )
                return site_id
        logging.warning(
            f"Site search returned {len(sites)} result(s) but none matched "
            f"{SHAREPOINT_HOST!r}. Sites found: "
            + ", ".join(s.get("webUrl", "?") for s in sites)
        )

    # ── Attempt 5: SharePoint REST API (bypasses Graph entirely) ─────────────
    rest_url = f"https://{SHAREPOINT_HOST}/_api/site"
    rest_headers = {
        "Authorization": f"Bearer {token}",
        "Accept":        "application/json;odata=verbose",
    }
    logging.info(f"SharePoint site discovery attempt 5 (REST API): GET {rest_url}")
    resp5 = requests.get(rest_url, headers=rest_headers)
    logging.info(f"  → HTTP {resp5.status_code}  body: {resp5.text}")
    if resp5.status_code == 200:
        # REST API returns the site URL; we still need the Graph site ID.
        # Re-try attempt 1 in case the REST success indicates auth is fine
        # but the Graph path format was the only issue.
        logging.info(
            "SharePoint REST API succeeded — Graph site ID still required. "
            "Re-trying Graph path lookup after REST confirmation."
        )
        resp1b = requests.get(url1, headers=headers)
        logging.info(f"  → HTTP {resp1b.status_code}  body: {resp1b.text}")
        if resp1b.status_code == 200:
            site_id = resp1b.json()["id"]
            logging.info("SharePoint site found via: REST confirmation + Graph path lookup")
            return site_id

    raise RuntimeError(
        f"All SharePoint site discovery attempts failed for {SHAREPOINT_HOST!r}. "
        "Check the logs above for full HTTP responses from each attempt."
    )


def _sharepoint_drive_id(client: GraphClient, site_id: str) -> str:
    """
    Find the 'Shared Documents' document library drive on the site.
    Falls back to the first available drive if the expected name is absent.
    """
    url    = f"{GRAPH_BASE}/sites/{site_id}/drives"
    drives = client.get(url).json().get("value", [])

    for drive in drives:
        if drive.get("name", "").lower() in ("shared documents", "documents"):
            return drive["id"]

    if not drives:
        raise RuntimeError("No document libraries found on SharePoint site.")

    logging.warning(
        f"'Shared Documents' drive not found; using '{drives[0]['name']}' instead."
    )
    return drives[0]["id"]


def _download_excel_file(
    client: GraphClient,
    sp_filename: str,
    local_path: str,
    job_folder: str = None,
) -> bool:
    """Download one Excel tracker from SharePoint to a local temp path.

    SharePoint path:
        /Shared Documents/MJHughes OPEN JOBS/{job_folder}/{sp_filename}

    *job_folder* defaults to _SP_JOB_FOLDER when not provided.
    Returns True on success, False on any failure (logs a warning).
    """
    job_folder = job_folder or _SP_JOB_FOLDER
    try:
        site_id  = _sharepoint_site_id(client)
        drive_id = _sharepoint_drive_id(client, site_id)
        url = (
            f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}"
            f"/root:/{SHAREPOINT_FOLDER}/{job_folder}/{sp_filename}:/content"
        )
        response  = client.get(url)
        temp_path = Path(local_path)
        temp_path.parent.mkdir(parents=True, exist_ok=True)
        temp_path.write_bytes(response.content)
        logging.info(
            f"Downloaded {sp_filename} from SharePoint "
            f"({len(response.content):,} bytes → {temp_path})"
        )
        return True
    except Exception as exc:
        logging.warning(f"Could not download {sp_filename} from SharePoint: {exc}")
        return False


def _upload_excel_file(
    client: GraphClient,
    sp_filename: str,
    local_path: str,
    job_folder: str = None,
) -> None:
    """Upload one Excel tracker to SharePoint, overwriting the existing file.

    SharePoint path:
        /Shared Documents/MJHughes OPEN JOBS/{job_folder}/{sp_filename}

    *job_folder* defaults to _SP_JOB_FOLDER when not provided.
    Raises on any HTTP or IO error — the caller logs a CRITICAL message and
    keeps the local temp copy so the data is not lost.
    """
    job_folder = job_folder or _SP_JOB_FOLDER
    temp_path = Path(local_path)
    if not temp_path.exists():
        raise FileNotFoundError(
            f"Temp Excel file not found for upload: {temp_path}"
        )

    site_id  = _sharepoint_site_id(client)
    drive_id = _sharepoint_drive_id(client, site_id)
    url = (
        f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}"
        f"/root:/{SHAREPOINT_FOLDER}/{job_folder}/{sp_filename}:/content"
    )
    data = temp_path.read_bytes()
    client.put_binary(
        url, data,
        content_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
    )
    logging.info(f"Uploaded {sp_filename} to SharePoint ({len(data):,} bytes)")


# ============================================================
# MULTI-JOB ROUTING HELPERS
# ============================================================

def _excel_local_path(job_number: str, is_concrete: bool = False) -> str:
    """Return the local temp path for a job's Excel tracker."""
    suffix   = "_concrete" if is_concrete else ""
    filename = f"ticket_tracker_{job_number}{suffix}.xlsx"
    return str(_LOCAL_TEMP_DIR / filename)


# Cache: job_number → True/False (whether the SP folder exists).
# Pre-populated with the configured fallback folder so we never make a
# round-trip to validate it — it must exist or the run would have failed earlier.
_sp_job_folder_cache: dict[str, bool] = {_SP_JOB_FOLDER: True}


def _sp_job_folder_exists(client: GraphClient, job_number: str) -> bool:
    """Return True if /MJHughes OPEN JOBS/{job_number} exists in SharePoint.

    Results are cached for the lifetime of the run.
    """
    if job_number in _sp_job_folder_cache:
        return _sp_job_folder_cache[job_number]
    try:
        site_id  = _sharepoint_site_id(client)
        drive_id = _sharepoint_drive_id(client, site_id)
        check_url = (
            f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}"
            f"/root:/{SHAREPOINT_FOLDER}/{job_number}"
        )
        client.get(check_url)
        _sp_job_folder_cache[job_number] = True
        return True
    except requests.exceptions.HTTPError as exc:
        if exc.response is not None and exc.response.status_code == 404:
            _sp_job_folder_cache[job_number] = False
            return False
        raise


def _resolve_sp_job_folder(client: GraphClient, job_number: str) -> str:
    """Return the SharePoint folder name to use for *job_number*.

    If /MJHughes OPEN JOBS/{job_number} exists, returns *job_number*.
    Otherwise falls back to *_SP_JOB_FOLDER* and logs a warning.
    """
    if _sp_job_folder_exists(client, job_number):
        return job_number
    logging.warning(
        f"Job folder {job_number!r} not found in SharePoint — "
        f"falling back to {_SP_JOB_FOLDER!r}"
    )
    return _SP_JOB_FOLDER


# Cache: (job_number, is_concrete) → True/False (download outcome).
_excel_downloaded: dict[tuple, bool] = {}


def _get_or_download_excel(
    client: GraphClient,
    job_number: str,
    is_concrete: bool = False,
) -> Optional[str]:
    """Return the local path of a job's Excel tracker, downloading if needed.

    On the first call for a given (job_number, is_concrete) pair, downloads the
    file from SharePoint using the exact job folder.  Returns None when the
    download fails or has already failed this run.
    """
    local_path = _excel_local_path(job_number, is_concrete)
    cache_key  = (job_number, is_concrete)

    if Path(local_path).exists():
        return local_path

    if _excel_downloaded.get(cache_key) is False:
        return None  # already tried and failed this run

    sp_filename = (
        f"ticket_tracker_{job_number}_concrete.xlsx"
        if is_concrete
        else f"ticket_tracker_{job_number}.xlsx"
    )
    ok = _download_excel_file(client, sp_filename, local_path, job_folder=job_number)
    _excel_downloaded[cache_key] = ok
    return local_path if ok else None


# ============================================================
# BATCH OUTLIER DETECTION
# ============================================================
def _check_batch_outliers(ticket_numbers: list[str]) -> None:
    """Flag ticket numbers that differ significantly from the batch median.

    Ticket numbers are sequential — within a single run they should cluster
    near each other.  A ticket that differs from the median by more than 2000
    is treated as a likely OCR misread and highlighted in Excel.

    Also detects the specific leading-digit-drop pattern (e.g. "07637" instead
    of "17637") where the leading "1" is silently dropped by OCR.

    Skips silently when fewer than 2 numeric ticket numbers are available.
    """
    if len(ticket_numbers) < 2:
        return

    # Convert to int — skip any non-numeric values
    numeric: list[tuple[int, str]] = []
    for tn in ticket_numbers:
        try:
            numeric.append((int(tn), tn))
        except (ValueError, TypeError):
            pass

    if len(numeric) < 2:
        return

    values     = [n for n, _ in numeric]
    median_val = statistics.median(values)
    outliers   = [(val, tn) for val, tn in numeric if abs(val - median_val) > 2000]

    if not outliers:
        logging.info(
            f"Batch outlier check: {len(numeric)} ticket(s), "
            f"median={int(median_val)}, no outliers detected."
        )
        return

    # Determine the most common leading digit to spot the drop pattern
    leading_digits    = [tn[0] for _, tn in numeric if tn]
    most_common_start = max(set(leading_digits), key=leading_digits.count) if leading_digits else ""

    # Update Excel: yellow ticket-number cell + note in Helper Notes column (J)
    excel_path = Path(EXCEL_FILE)
    _yellow    = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    workbook = openpyxl.load_workbook(excel_path) if excel_path.exists() else None

    for val, tn in outliers:
        diff = abs(val - int(median_val))
        logging.warning(
            f"WARNING: Ticket number {tn} may be a misread — differs from batch "
            f"median {int(median_val)} by {diff}. "
            f"Please verify against original PDF."
        )

        # Leading-digit-drop: starts with 0 while the rest of the batch starts with 1
        if tn.startswith("0") and most_common_start == "1":
            suggested = "1" + tn[1:]
            logging.warning(
                f"Possible misread: {tn} may be {suggested} (leading digit dropped)"
            )

        if workbook is None:
            continue

        for sheet in workbook.worksheets:
            if sheet.title == _TOC_TAB_NAME:
                continue
            for row_cells in sheet.iter_rows(min_row=_DATA_START_ROW):
                if len(row_cells) < _COL_TICKET_NUM:
                    continue
                cell = row_cells[_COL_TICKET_NUM - 1]
                if cell.value and str(cell.value).strip() == tn:
                    cell.fill = _yellow
                    notes_cell = row_cells[_COL_NOTES - 1]
                    existing   = str(notes_cell.value or "").strip()
                    notes_cell.value = (
                        f"{existing} | VERIFY TICKET NUMBER - outlier detected" if existing
                        else "VERIFY TICKET NUMBER - outlier detected"
                    )

    if workbook is not None:
        workbook.save(excel_path)
        logging.info(
            f"Batch outlier check: flagged {len(outliers)} ticket(s) "
            f"(median={int(median_val)})."
        )


# ============================================================
# TICKET BOUNDARY DETECTION
# ============================================================
def detect_ticket_boundaries(image: "Image.Image", page_num: int) -> list[dict]:
    """Use Claude vision to detect how many tickets appear on a page.

    Sends the page image (resized to ≤1500 px) to Claude and asks it to
    return each ticket's vertical extent as a percentage of image height.

    Returns a list of dicts with keys: ticket (int), top_pct (float),
    bottom_pct (float).

    Falls back to [{"ticket": 1, "top_pct": 0, "bottom_pct": 100}] when:
      - ANTHROPIC_API_KEY is not set
      - The API call raises
      - The response cannot be parsed as a valid list
    """
    _FALLBACK = [{"ticket": 1, "top_pct": 0, "bottom_pct": 100}]

    if not ANTHROPIC_API_KEY:
        return _FALLBACK

    img_copy = image.copy()
    max_side = 1500
    if max(img_copy.size) > max_side:
        ratio    = max_side / max(img_copy.size)
        new_size = (int(img_copy.size[0] * ratio), int(img_copy.size[1] * ratio))
        img_copy = img_copy.resize(new_size, Image.LANCZOS)

    buf = io.BytesIO()
    img_copy.convert("RGB").save(buf, format="JPEG", quality=85)
    b64_data = base64.b64encode(buf.getvalue()).decode("ascii")

    prompt = (
        "This image may contain one or more material delivery "
        "tickets on a single page. Each ticket has a header "
        "with the supplier company name at the top.\n\n"
        "Count how many tickets are on this page and provide "
        "the approximate pixel coordinates of each ticket as "
        "a percentage of the total image height.\n\n"
        "Return ONLY a JSON array like this:\n"
        "[\n"
        "  {\"ticket\": 1, \"top_pct\": 0, \"bottom_pct\": 48},\n"
        "  {\"ticket\": 2, \"top_pct\": 50, \"bottom_pct\": 100}\n"
        "]\n\n"
        "If there is only one ticket return:\n"
        "[{\"ticket\": 1, \"top_pct\": 0, \"bottom_pct\": 100}]\n\n"
        "No other text, no markdown, no explanation."
    )

    try:
        ai_client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        response  = ai_client.messages.create(
            model=_AI_MODEL,
            max_tokens=256,
            messages=[{
                "role": "user",
                "content": [
                    {
                        "type": "image",
                        "source": {
                            "type":       "base64",
                            "media_type": "image/jpeg",
                            "data":       b64_data,
                        },
                    },
                    {"type": "text", "text": prompt},
                ],
            }],
        )
        raw = response.content[0].text.strip()

        # Strip accidental markdown fences
        if raw.startswith("```"):
            raw = re.sub(r"^```[a-z]*\n?", "", raw)
            raw = re.sub(r"\n?```$",        "", raw)

        parsed = json.loads(raw)

        if not isinstance(parsed, list) or not parsed:
            logging.warning(
                f"  Page {page_num}: boundary detection returned "
                f"unexpected format: {raw!r} — using full-page fallback"
            )
            return _FALLBACK

        validated: list[dict] = []
        for item in parsed:
            if not isinstance(item, dict):
                continue
            validated.append({
                "ticket":     int(item.get("ticket",     len(validated) + 1)),
                "top_pct":    float(item.get("top_pct",    0)),
                "bottom_pct": float(item.get("bottom_pct", 100)),
            })

        if not validated:
            return _FALLBACK

        logging.info(f"  Page {page_num}: Claude detected {len(validated)} ticket(s)")
        for b in validated:
            logging.info(
                f"    Ticket {b['ticket']}: "
                f"top={b['top_pct']}%  bottom={b['bottom_pct']}%"
            )
        return validated

    except Exception as exc:
        logging.warning(
            f"  Page {page_num}: boundary detection failed ({exc}) "
            "— treating page as single ticket."
        )
        return _FALLBACK


def _crop_ticket_region(
    image: "Image.Image",
    top_pct: float,
    bottom_pct: float,
    overlap_pct: float = 2.0,
) -> "Image.Image":
    """Crop a ticket region from a page image using percentage coordinates.

    Adds *overlap_pct* percentage points above and below the specified bounds
    to avoid clipping edge content.  Result is clamped to image dimensions.
    """
    w, h   = image.size
    top    = max(0, int(h * (top_pct    - overlap_pct) / 100))
    bottom = min(h, int(h * (bottom_pct + overlap_pct) / 100))
    return image.crop((0, top, w, bottom))


def _pil_image_to_pdf_bytes(image: "Image.Image") -> bytes:
    """Encode a PIL Image as a single-page PDF and return the raw bytes.

    PIL's built-in PDF writer is used (no extra dependencies).  The image is
    converted to RGB first because PIL's PDF encoder requires that mode.
    """
    if image.mode != "RGB":
        image = image.convert("RGB")
    buf = io.BytesIO()
    image.save(buf, format="PDF")
    return buf.getvalue()


def _apply_duplicate_copy_filter(
    page_image: "Image.Image",
    boundaries: list[dict],
    profiles: list[TicketProfile],
    page_num: int,
    prefetched_qr: "tuple[Optional[dict], Optional[int]]" = (None, None),
) -> list[dict]:
    """Filter out duplicate-copy boundaries for suppliers that print two
    identical copies on one page (e.g. Knife River Customer/Office copy).

    When exactly two boundaries are detected and the page profile has
    ``layout.duplicate_copy_page = true``, only the boundary whose region
    contains a QR code is kept.

    Detection strategy:
      1. Use ``prefetched_qr`` (qr_data, y_center_px) if provided — the
         caller pre-scanned the full page before any cropping.
      2. If not prefetched, call ``_scan_full_page_for_qr`` directly.
      3. Fall back to scanning each crop individually via ``extract_qr_code``
         only when the full-page scan also fails.

    If both crops have a QR code the first is used; if neither has a QR code
    the boundaries are returned unchanged so normal processing can flag the
    page for review.

    For all other cases (≠ 2 boundaries, or profile does not flag duplicate
    copies) the original boundaries list is returned as-is.
    """
    if len(boundaries) != 2:
        return boundaries

    # Quick profile detection: OCR just the top 15 % of the page (fast).
    _, h      = page_image.size
    w, _h     = page_image.size
    top_strip = page_image.crop((0, 0, w, min(h, int(h * 0.15))))
    try:
        quick_text   = pytesseract.image_to_string(
            preprocess_for_ocr(top_strip), config="--oem 3 --psm 6"
        )
        page_profile = detect_profile(quick_text, profiles)
    except Exception:
        return boundaries

    if not page_profile or not page_profile.layout.get("duplicate_copy_page"):
        return boundaries

    # ── Strategy 1 & 2: full-page QR scan (prefetched or fresh) ────────────
    qr_data, qr_y = prefetched_qr
    if qr_data is None:
        qr_data, qr_y = _scan_full_page_for_qr(page_image, page_num)

    if qr_data is not None and qr_y is not None:
        # Assign the QR to whichever boundary's Y range contains qr_y.
        for boundary in boundaries:
            top_y = int(h * boundary["top_pct"] / 100)
            bot_y = int(h * boundary["bottom_pct"] / 100)
            if top_y <= qr_y < bot_y:
                logging.info(
                    f"  [p{page_num}] {page_profile.supplier_name} duplicate copy "
                    f"detected — processing only the copy with QR code "
                    f"(ticket {boundary['ticket']}, "
                    f"top={boundary['top_pct']}% bottom={boundary['bottom_pct']}%)"
                )
                return [boundary]
        # QR found but Y didn't fall in any boundary — fall through to crop scan.

    # ── Strategy 3: per-crop QR scan (fallback) ─────────────────────────────
    qr_boundary: Optional[dict] = None
    for boundary in boundaries:
        crop    = _crop_ticket_region(
            page_image, boundary["top_pct"], boundary["bottom_pct"]
        )
        crop_qr = extract_qr_code([crop])
        if crop_qr:
            qr_boundary = boundary
            break

    if qr_boundary is None:
        # Neither half has a QR — return both and let Phase 2 flag the page.
        return boundaries

    logging.info(
        f"  [p{page_num}] {page_profile.supplier_name} duplicate copy detected "
        f"— processing only the copy with QR code "
        f"(ticket {qr_boundary['ticket']}, "
        f"top={qr_boundary['top_pct']}% bottom={qr_boundary['bottom_pct']}%)"
    )
    return [qr_boundary]


# ============================================================
# PER-EMAIL PROCESSING PIPELINE
# ============================================================
def _build_success_subject(
    ticket_numbers: list,
    first_ticket_data: dict,
    types_seen: set,
    partial: bool = False,
) -> str:
    """Build the email subject line for a successfully processed email.

    Examples (single):   "Rock Ticket 16800 - 08/20/2025"
    Examples (multi):    "Rock Tickets 16800, 16801 - 08/20/2025"
    Examples (mixed):    "Mixed Tickets 16800, 2805126 - 08/20/2025"
    Examples (partial):  "Rock Tickets 16800, 16801 - 08/20/2025 (partial)"
    """
    # Type label — "Mixed" when tickets from different types were collected
    clean_types = {t for t in types_seen if t and t != "unknown"}
    if len(clean_types) > 1:
        type_label = "Mixed"
    elif len(clean_types) == 1:
        type_label = next(iter(clean_types)).title()
    else:
        type_label = "Unknown"

    plural   = "Tickets" if len(ticket_numbers) > 1 else "Ticket"
    tn_str   = ", ".join(ticket_numbers) if ticket_numbers else "UNKNOWN"
    raw_date = ((first_ticket_data or {}).get("date", "") or "").strip()
    date_str = raw_date if raw_date else datetime.now().strftime("%m/%d/%Y")

    subject = f"{type_label} {plural} {tn_str} - {date_str}"
    if partial:
        subject += " (partial)"
    return subject


def process_email(
    client: GraphClient,
    email: dict,
    profiles: list[TicketProfile],
    toc_materials: dict,
    ticket_types: dict,
) -> Optional[dict]:
    """Process one email end-to-end.

    For each PDF attachment:
      0. Download and render the first page; detect ticket type from full-page
         OCR.  Emails that don't match the configured TICKET_TYPE are sorted
         to the appropriate Exchange folder and skipped.
      1. Download and render to page images.
      2. Ask Claude to detect how many tickets are on each page.
      3. Crop each ticket region (with 2 % overlap) and process
         independently: QR extraction → OCR → confidence check →
         material verification → duplicate check → Excel write →
         SharePoint PDF upload.
      4. After all tickets on all pages are handled, send one email
         reply, rename the subject, and mark as read (or flag for
         review if any ticket needed human attention).

    Returns:
        dict with keys: tickets, review_needed, subject, duplicate_count,
             review_reasons, and optionally sorted_type when the email was
             moved to a sort folder rather than processed.
        None — unhandled exception; email stays unread for retry.
    """
    subject  = email.get("subject", "(no subject)")
    email_id = email["id"]

    logging.info(f"{'='*55}")
    logging.info(f"Processing: '{subject}'")

    # Iterate over every PDF attachment in the email
    for attachment in email["pdf_attachments"]:
        att_name = attachment.get("name", "attachment.pdf")

        _current_step = "initializing"
        try:
            # ---- 1. Download PDF ------------------------------------------------
            _current_step = f"downloading attachment '{att_name}'"
            logging.info(f"  Downloading '{att_name}'...")
            pdf_bytes = get_attachment_content(client, email_id, attachment["id"])

            # ---- 2. Convert PDF to images ---------------------------------------
            _current_step = "rendering PDF pages"
            logging.info("  Rendering PDF pages...")
            images = pdf_to_images(pdf_bytes, dpi=200)
            if not images:
                raise ValueError("PDF rendered 0 pages.")

            # ---- 2b. Ticket type detection (full-page OCR, first page only) -----
            # Run on the raw full-page image before any cropping so all text is
            # visible.  This determines whether this processor should handle the
            # email or sort it to another folder.
            # detected_type is initialized here so it is always defined when the
            # Phase 3 upload loop runs, even when ticket_types is empty.
            detected_type = ""
            _current_step = "detecting ticket type"
            if ticket_types:
                type_ocr = pytesseract.image_to_string(
                    images[0], config="--oem 3 --psm 6"
                )
                detected_type, type_score = detect_ticket_type(type_ocr, ticket_types)
                logging.info(
                    f"  Ticket type detected: {detected_type} "
                    f"(score: {type_score} keywords matched)"
                )

                if detected_type == "unknown":
                    # Cannot determine type — flag for manual review.
                    logging.info(
                        f"  Unknown ticket type — moving to '{_review_folder_name('unknown')}' folder."
                    )
                    new_id = move_email_to_review_folder(client, email_id, ticket_type="unknown")
                    flag_email_category(client, new_id)
                    mark_email_as_unread(client, new_id)
                    return {
                        "tickets":         [],
                        "review_needed":   True,
                        "subject":         subject,
                        "duplicate_count": 0,
                        "review_reasons":  ["Unknown ticket type — could not classify"],
                    }

                if _TICKET_TYPE and detected_type != _TICKET_TYPE:
                    # This processor handles a different type — sort and skip.
                    type_config   = ticket_types[detected_type]
                    folder_name   = type_config.get(
                        "target_folder", f"{detected_type.title()} Tickets"
                    )
                    sort_email_to_type_folder(client, email_id, subject, folder_name)
                    logging.info(
                        f"  Ticket type {detected_type!r} detected — sorted to "
                        f"{folder_name!r}. Processing for this type not yet implemented."
                    )
                    return {
                        "tickets":         [],
                        "review_needed":   False,
                        "subject":         subject,
                        "duplicate_count": 0,
                        "review_reasons":  [],
                        "sorted_type":     detected_type,
                    }
                # else: type matches _TICKET_TYPE (or no filter set) — continue.

            # ── Retry snapshot: record every ticket number already in Excel. ──
            # If this email is being reprocessed after a partial failure, any
            # ticket whose number appears here will be skipped (no duplicate
            # Excel row, no duplicate PDF upload) and counted as already-done.
            if detected_type == "concrete":
                already_in_excel: frozenset = _read_existing_ticket_numbers(
                    excel_path=EXCEL_FILE_CONCRETE,
                    col_ticket_num=_CONCRETE_COL_TICKET_NUM,
                    data_start_row=_CONCRETE_DATA_START_ROW,
                )
            else:
                already_in_excel: frozenset = _read_existing_ticket_numbers()
            if already_in_excel:
                logging.info(
                    f"  Retry snapshot: {len(already_in_excel)} ticket number(s) "
                    "already in Excel — will skip if re-encountered."
                )

            # ---- 3. Detect ticket boundaries per page; process each crop --------
            # all_ticket_numbers : ticket numbers from rows successfully written
            # any_review_needed  : True when at least one ticket needs human review
            # last_success       : (qr_data, ticket_data) from the last clean ticket
            all_ticket_numbers: list[str]          = []
            any_review_needed:  bool               = False
            duplicate_count:    int                = 0
            review_reasons:     list[str]          = []
            last_success:       Optional[tuple]    = None
            first_success:      Optional[tuple]    = None
            types_seen:         set                = set()

            for page_num, page_image in enumerate(images, start=1):
                _current_step = f"scanning QR code on page {page_num}"

                # Pre-scan the full uncropped page for a QR code.  AI vision
                # is tried first; pyzbar is the fallback.  Running before
                # any cropping ensures codes near a crop boundary are not lost.
                # The result is passed into both the duplicate-copy filter and
                # the per-ticket QR step below.
                page_qr_data, page_qr_y = _ai_scan_qr_from_image(
                    page_image, page_num
                )
                if page_qr_data:
                    logging.info(
                        f"  [p{page_num}] QR found via AI vision: "
                        f"'{page_qr_data['raw']}' "
                        f"(estimated Y={page_qr_y}px, "
                        f"page height {page_image.size[1]}px)"
                    )
                else:
                    page_qr_data, page_qr_y = _scan_full_page_for_qr(
                        page_image, page_num
                    )
                    if page_qr_data:
                        logging.info(
                            f"  [p{page_num}] QR found via pyzbar (AI missed it): "
                            f"'{page_qr_data['raw']}' at Y={page_qr_y}px "
                            f"(page height {page_image.size[1]}px)"
                        )
                    else:
                        # Step 3: OCR text search on the full page image.
                        # Runs after AI vision and pyzbar both fail.  The QR
                        # sticker always has the code printed in plain text
                        # beside the QR image, so even when the image can't be
                        # decoded the text can be OCR'd directly.
                        logging.info(
                            f"  [p{page_num}] QR: AI + pyzbar failed — "
                            "Step 3: trying full-page OCR text search..."
                        )
                        _page_ocr_text = pytesseract.image_to_string(
                            page_image, config="--psm 6 --oem 3"
                        )
                        _ocr_qr_match = QR_PATTERN.search(_page_ocr_text)
                        if _ocr_qr_match:
                            _ocr_qr_raw  = _ocr_qr_match.group(0)
                            page_qr_data = _parse_qr_string(_ocr_qr_raw)
                            page_qr_y    = page_image.size[1] // 2  # center; unknown
                            logging.info(
                                f"  [p{page_num}] QR found via OCR text: "
                                f"'{_ocr_qr_raw}' (Y estimated to page centre)"
                            )
                        else:
                            logging.info(
                                f"  [p{page_num}] QR not found by any method "
                                "(AI vision + pyzbar + OCR)"
                            )

                boundaries = detect_ticket_boundaries(page_image, page_num)
                boundaries = _apply_duplicate_copy_filter(
                    page_image, boundaries, profiles, page_num,
                    prefetched_qr=(page_qr_data, page_qr_y),
                )

                # ── Phase 1: Extract all ticket data for this page ──────────────
                # Process every ticket crop before making any write/upload
                # decisions.  Results are stored and evaluated together so that a
                # single missing ticket number causes the ENTIRE page to be
                # flagged rather than partial writes.
                page_results: list[dict] = []

                for boundary in boundaries:
                    t_idx   = boundary["ticket"]
                    top_pct = boundary["top_pct"]
                    bot_pct = boundary["bottom_pct"]
                    tag     = f"  [p{page_num}/t{t_idx}]"

                    res: dict = {
                        "tag":           tag,
                        "qr_data":       None,
                        "ticket_data":   None,
                        "material_note": "",
                        "is_duplicate":  False,
                        "duplicate_tab": None,
                        "failed":        False,
                        "fail_reason":   "",
                        "crop_img":      None,
                    }

                    # Single full-page ticket: skip the crop and use the full image
                    if len(boundaries) == 1 and top_pct <= 0 and bot_pct >= 100:
                        crop_imgs = [page_image]
                        res["crop_img"] = page_image
                        logging.info(f"{tag} Single full-page ticket — no crop needed")
                    else:
                        crop = _crop_ticket_region(page_image, top_pct, bot_pct)
                        crop_imgs = [crop]
                        res["crop_img"] = crop
                        logging.info(
                            f"{tag} Cropped: top={top_pct}%  bottom={bot_pct}%  "
                            f"({crop.size[0]}×{crop.size[1]} px)"
                        )

                    # QR code extraction — prefer the full-page pre-scan result
                    # when the QR's Y coordinate falls within this boundary's range.
                    # Only scan the crop when the full-page scan did not find a code.
                    logging.info(f"{tag} Scanning for QR code...")
                    qr_data = None
                    if page_qr_data is not None and page_qr_y is not None:
                        _ph   = page_image.size[1]
                        _ty   = int(_ph * top_pct / 100)
                        _by   = int(_ph * bot_pct / 100)
                        if _ty <= page_qr_y < _by:
                            qr_data = page_qr_data
                            logging.info(
                                f"{tag} QR assigned from full-page scan "
                                f"(Y={page_qr_y}px within boundary {_ty}–{_by}px)"
                            )
                            if OCR_DEBUG:
                                logging.info(
                                    f"QR found at Y={page_qr_y}px on full page "
                                    f"({_ph}px total) → assigned to ticket {t_idx}"
                                )
                    if qr_data is None:
                        qr_data = extract_qr_code(crop_imgs)
                    if not qr_data:
                        logging.warning(
                            f"{tag} QR DETECTION FAILED — no QR code found in region"
                        )
                        res["failed"]      = True
                        res["fail_reason"] = "QR code not found"
                        page_results.append(res)
                        continue

                    res["qr_data"] = qr_data
                    logging.info(
                        f"{tag} QR → Job: {qr_data['job_number']}  "
                        f"Location: {qr_data['location']}  "
                        f"Cost Code: {qr_data['cost_code']}"
                    )

                    # OCR + AI extraction
                    _current_step = f"running OCR/AI extraction on page {page_num} ticket {t_idx}"
                    logging.info(f"{tag} Running OCR...")
                    ticket_data = extract_ticket_data_ocr(
                        crop_imgs, profiles, subject=subject
                    )

                    # Unknown supplier
                    if ticket_data.get("_no_profile_match"):
                        company_name = ticket_data.get("_company_name", "unknown")
                        err = (
                            f"UNKNOWN TICKET FORMAT: {company_name}. "
                            "No profile found. "
                            "Please create a profile for this supplier."
                        )
                        logging.warning(f"{tag} {err}")
                        log_error(subject, err)
                        res["failed"]      = True
                        res["fail_reason"] = err
                        page_results.append(res)
                        continue

                    # OCR confidence failure
                    if not ticket_data.get("_ocr_confidence_passed", True):
                        failed_checks = ticket_data.get("_failed_checks", [])
                        logging.warning(f"{tag} TICKET FLAGGED FOR REVIEW:")
                        for chk in failed_checks:
                            logging.warning(f"{tag}   - {chk}")
                        log_error(
                            subject,
                            f"REVIEW REQUIRED: {subject} — "
                            + ("; ".join(failed_checks) or "unknown"),
                        )
                        res["failed"]      = True
                        res["fail_reason"] = (
                            "; ".join(failed_checks) or "OCR confidence failed"
                        )
                        res["ticket_data"] = ticket_data
                        page_results.append(res)
                        continue

                    res["ticket_data"] = ticket_data
                    ticket_number = ticket_data.get("ticket_number", "")
                    if detected_type == "concrete":
                        logging.info(
                            f"{tag} OCR → Ticket: {ticket_number or 'N/A'}  "
                            f"Date: {ticket_data.get('date', 'N/A')}  "
                            f"Supplier: {ticket_data.get('supplier', 'N/A')}  "
                            f"Qty CY: {ticket_data.get('qty_delivered', 'N/A')}"
                        )
                    else:
                        logging.info(
                            f"{tag} OCR → Ticket: {ticket_number or 'N/A'}  "
                            f"Date: {ticket_data.get('date', 'N/A')}  "
                            f"Facility: {ticket_data.get('facility', 'N/A')}  "
                            f"Net Tons: {ticket_data.get('net_tons', 'N/A')}"
                        )

                    # Material verification against TOC (rock only)
                    material_note = ""
                    if detected_type != "concrete" and toc_materials and qr_data:
                        qr_raw_upper = qr_data.get("raw", "").strip().upper()
                        anticipated  = toc_materials.get(qr_raw_upper, "")
                        actual_mat   = ticket_data.get("material", "")
                        if anticipated and actual_mat:
                            try:
                                if not _verify_material_match(anticipated, actual_mat):
                                    material_note = (
                                        f"Correct material? "
                                        f"Anticipated: {anticipated} "
                                        f"| Found: {actual_mat}"
                                    )
                                    logging.warning(
                                        f"{tag} WARNING: Material mismatch for "
                                        f"ticket {ticket_number} — "
                                        f"Anticipated: {anticipated} "
                                        f"| Found: {actual_mat}"
                                    )
                            except Exception as ve:
                                logging.warning(
                                    f"{tag} Material verification skipped: {ve}"
                                )
                    res["material_note"] = material_note

                    # Duplicate ticket check — route to the correct workbook
                    if detected_type == "concrete":
                        duplicate_tab = _find_duplicate_ticket(
                            ticket_number,
                            excel_path=EXCEL_FILE_CONCRETE,
                            col_ticket_num=_CONCRETE_COL_TICKET_NUM,
                            data_start_row=_CONCRETE_DATA_START_ROW,
                        )
                    else:
                        duplicate_tab = _find_duplicate_ticket(ticket_number)
                    if duplicate_tab:
                        logging.warning(
                            f"{tag} DUPLICATE TICKET: {ticket_number!r} already "
                            f"in tab '{duplicate_tab}'"
                        )
                        log_error(
                            subject,
                            f"DUPLICATE TICKET: {ticket_number} - {subject} - "
                            + datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        )
                        res["is_duplicate"]  = True
                        res["duplicate_tab"] = duplicate_tab

                    page_results.append(res)

                # ── Phase 2: Whole-page review gate ─────────────────────────────
                # Rule: if ANY ticket on the page is missing its ticket number
                # (extraction failed for any reason OR OCR returned an empty
                # string), do NOT write any rows for this page and flag the
                # email for review.  Other missing fields (facility, material,
                # etc.) still write with yellow highlighting per normal rules.
                failed_on_page    = [r for r in page_results if r["failed"]]
                missing_tn_on_page = [
                    r for r in page_results
                    if not r["failed"]
                    and not (r.get("ticket_data") or {}).get("ticket_number")
                ]
                bad_count   = len(failed_on_page) + len(missing_tn_on_page)
                total_count = len(page_results)

                if bad_count > 0:
                    logging.warning(
                        f"REVIEW REQUIRED: {bad_count} of {total_count} "
                        f"ticket(s) on page {page_num} missing ticket number "
                        f"— entire page flagged"
                    )
                    any_review_needed = True
                    review_reasons.append(
                        f"Page {page_num}: {bad_count} of {total_count} "
                        "ticket(s) missing ticket number"
                    )
                    continue   # skip Phase 3 — no Excel writes, no PDF upload

                # ── Phase 3: All tickets have ticket numbers — write and upload ──
                for res in page_results:
                    qr_d     = res["qr_data"]
                    td       = res["ticket_data"]
                    tn       = td.get("ticket_number", "")
                    notes    = res["material_note"]
                    job      = qr_d["job_number"]
                    crop_img = res["crop_img"]

                    # ── Retry skip: ticket already in Excel from a prior attempt ──
                    # Do not write a duplicate Excel row or re-upload the PDF.
                    # Count it as successfully done so the email can be marked read.
                    if tn and tn in already_in_excel:
                        logging.info(
                            f"{res['tag']} Skipping ticket {tn} — already "
                            "processed in a previous attempt"
                        )
                        all_ticket_numbers.append(tn)
                        last_success = (qr_d, td)
                        if first_success is None:
                            first_success = (qr_d, td)
                        if detected_type:
                            types_seen.add(detected_type)
                        continue   # skip Excel write AND PDF upload

                    # Resolve the effective SharePoint folder for this job.
                    # Falls back to _SP_JOB_FOLDER when the job folder doesn't
                    # exist in SharePoint yet (logged as a warning inside).
                    effective_job = _resolve_sp_job_folder(client, job)
                    _is_concrete  = (detected_type == "concrete")
                    _current_step = f"writing ticket {tn} to Excel"

                    # Lazy-download this job's Excel tracker if not yet on disk.
                    job_excel_path = _get_or_download_excel(
                        client, effective_job, is_concrete=_is_concrete
                    )
                    write_succeeded = True

                    if job_excel_path is None:
                        logging.warning(
                            f"{res['tag']} WARNING: Cannot fetch Excel for job "
                            f"{effective_job!r} — skipping Excel write."
                        )
                        write_succeeded = False
                    else:
                        try:
                            if res["is_duplicate"]:
                                logging.info(f"{res['tag']} Writing duplicate row to Excel...")
                                _dup_notes = _append_note("DUPLICATE TICKET", notes)
                                if _is_concrete:
                                    write_to_excel_concrete(
                                        qr_d, td, notes=_dup_notes,
                                        excel_path=job_excel_path,
                                    )
                                else:
                                    write_to_excel(
                                        qr_d, td,
                                        notes=_dup_notes,
                                        toc_materials=toc_materials,
                                        excel_path=job_excel_path,
                                    )
                                any_review_needed = True
                                duplicate_count  += 1
                                review_reasons.append(f"Duplicate ticket: {tn}")
                            else:
                                logging.info(f"{res['tag']} Writing to Excel...")
                                if _is_concrete:
                                    write_to_excel_concrete(
                                        qr_d, td, notes=notes,
                                        excel_path=job_excel_path,
                                    )
                                else:
                                    write_to_excel(
                                        qr_d, td,
                                        notes=notes,
                                        toc_materials=toc_materials,
                                        excel_path=job_excel_path,
                                    )
                        except AmbiguousTabError as amb_exc:
                            logging.warning(str(amb_exc))
                            any_review_needed = True
                            write_succeeded   = False
                            review_reasons.append(str(amb_exc))

                    if write_succeeded:
                        all_ticket_numbers.append(tn)
                        last_success = (qr_d, td)
                        if first_success is None:
                            first_success = (qr_d, td)
                        if detected_type:
                            types_seen.add(detected_type)

                    # Upload one crop PDF per ticket to the resolved job folder.
                    # Rock PDFs   → Ticket Scans/Rock/[material]/
                    # Concrete PDFs → Ticket Scans/Concrete/[supplier]/
                    # Falls back to the type root when no material/supplier name
                    # is available, and to the flat Ticket Scans root when type
                    # detection was not configured.
                    if crop_img is not None:
                        _current_step = f"uploading ticket {tn} PDF to SharePoint"
                        try:
                            logging.info(
                                f"{res['tag']} Uploading crop PDF to SharePoint "
                                f"(ticket: {tn}, job: {effective_job})..."
                            )
                            crop_pdf       = _pil_image_to_pdf_bytes(crop_img)
                            type_subfolder = detected_type.title() if detected_type else ""

                            # Build material-level subfolder name
                            if detected_type == "rock":
                                raw_mat = (td.get("qr_sticker_text") or "").strip()
                                if raw_mat:
                                    mat_subfolder = _sanitize_sharepoint_folder_name(raw_mat)
                                else:
                                    mat_subfolder = ""
                                    logging.info("No material name - uploading to root Rock folder")
                            elif detected_type == "concrete":
                                raw_mat = (td.get("supplier") or "").strip()
                                if raw_mat:
                                    mat_subfolder = _sanitize_sharepoint_folder_name(raw_mat)
                                else:
                                    mat_subfolder = ""
                                    logging.info("No supplier name - uploading to root Concrete folder")
                            else:
                                mat_subfolder = ""

                            upload_to_sharepoint(
                                client, crop_pdf, effective_job, [tn],
                                subfolder=type_subfolder,
                                material_subfolder=mat_subfolder,
                            )
                        except Exception as sp_exc:
                            logging.warning(
                                f"{res['tag']} PDF upload failed (will continue): "
                                f"{sp_exc}"
                            )

            # ---- Email-level operations (once per attachment) -------------------
            if any_review_needed and not all_ticket_numbers:
                # ALL tickets failed — no data written; use REVIEW REQUIRED prefix
                logging.info(
                    "  All tickets failed — flagging email for review."
                )
                try:
                    flag_email_category(client, email_id)
                except Exception as e:
                    logging.warning(f"  Category flag not applied: {e}")
                try:
                    rename_email_subject(
                        client, email_id,
                        f"REVIEW REQUIRED - {subject}",
                        current_subject=subject,
                    )
                except Exception as e:
                    logging.warning(f"  Subject rename not applied: {e}")
                moved_email_id = email_id
                try:
                    moved_email_id = move_email_to_review_folder(client, email_id, detected_type)
                except Exception as e:
                    logging.warning(f"  Email move not completed: {e}")
                mark_email_as_unread(client, moved_email_id)

            elif any_review_needed and all_ticket_numbers:
                # PARTIAL — some tickets written, some failed
                _partial_subject = _build_success_subject(
                    all_ticket_numbers,
                    (first_success or (None, {}))[1],
                    types_seen,
                    partial=True,
                )
                logging.info(
                    f"  Partial success ({len(all_ticket_numbers)} ticket(s) written, "
                    f"review flag set) — renaming to: {_partial_subject!r}"
                )
                try:
                    flag_email_category(client, email_id)
                except Exception as e:
                    logging.warning(f"  Category flag not applied: {e}")
                try:
                    rename_email_subject(
                        client, email_id, _partial_subject, current_subject=subject
                    )
                except Exception as e:
                    logging.warning(f"  Subject rename not applied: {e}")
                moved_email_id = email_id
                try:
                    moved_email_id = move_email_to_review_folder(client, email_id, detected_type)
                except Exception as e:
                    logging.warning(f"  Email move not completed: {e}")
                mark_email_as_unread(client, moved_email_id)

            elif last_success:
                # ALL tickets processed cleanly — rename and archive to processed folder
                _current_step  = "completing email operations (rename/mark-read/archive)"
                _clean_subject = _build_success_subject(
                    all_ticket_numbers,
                    first_success[1],
                    types_seen,
                    partial=False,
                )
                logging.info(f"  Renaming email to: {_clean_subject!r}")
                try:
                    rename_email_subject(client, email_id, _clean_subject)
                except Exception as e:
                    logging.warning(f"  Subject rename not applied: {e}")
                logging.info("  Marking email as read...")
                mark_email_as_read(client, email_id)
                try:
                    move_email_to_processed_folder(client, email_id, subject, detected_type)
                except Exception as e:
                    logging.warning(f"  Could not archive email to '{_processed_folder_name(detected_type)}' folder: {e}")

            logging.info(
                f"  Done: '{subject}' — "
                f"{len(all_ticket_numbers)} ticket(s) written, "
                f"review_needed={any_review_needed}"
            )
            return {
                "tickets":         all_ticket_numbers,
                "review_needed":   any_review_needed,
                "subject":         subject,
                "duplicate_count": duplicate_count,
                "review_reasons":  review_reasons,
                "detected_type":   detected_type,
            }

        except Exception as exc:
            tb     = traceback.format_exc()
            detail = f"{type(exc).__name__}: {exc}"
            logging.error(
                f"  FAILED on '{att_name}' at step '{_current_step}': {exc}"
            )
            logging.error(f"  Traceback:\n{tb}")
            log_error(
                subject,
                f"Attachment '{att_name}' [{_current_step}]: {detail}\n{tb}",
            )
            return {                        # Email stays unread → will retry next run
                "failed":       True,
                "subject":      subject,
                "att_name":     att_name,
                "error_step":   _current_step,
                "error_detail": detail,
                "traceback":    tb,
            }

    # No PDF attachments were processed (shouldn't happen given the filter above)
    return None


# ============================================================
# ENTRY POINT
# ============================================================
def main() -> None:
    """
    Main driver:
        1. Validates .env credentials
        2. Authenticates with Graph API
        3. Downloads ticket_tracker_2601.xlsx from SharePoint to a temp path
        4. Fetches unread emails with PDF attachments and processes each one
        5. Uploads the modified Excel back to SharePoint and deletes the temp copy
    """
    configure_logging()
    run_start = datetime.now()

    logging.info("=" * 55)
    logging.info("Ticket Processing Automation — starting")
    logging.info(f"Run time: {run_start.strftime('%Y-%m-%d %H:%M:%S')}")
    logging.info("=" * 55)
    logging.info(f"Platform: {platform.system()}")
    logging.info(f"Tesseract path: {pytesseract.pytesseract.tesseract_cmd}")

    # Fail fast if any required credential is missing
    missing_vars = [
        v for v in ("AZURE_CLIENT_ID", "AZURE_CLIENT_SECRET", "AZURE_TENANT_ID")
        if not os.getenv(v)
    ]
    if missing_vars:
        raise EnvironmentError(
            f"Missing required .env variables: {', '.join(missing_vars)}\n"
            f"  .env loaded from: {_env_path}\n"
            f"  Make sure the file exists at that path and contains the required keys."
        )

    # Confirm credentials are being read from the right place
    tenant_preview = os.getenv("AZURE_TENANT_ID", "")[:4]
    logging.info(f"Credentials loaded from: {_env_path}")
    logging.info(f"AZURE_TENANT_ID starts with: {tenant_preview!r}")

    # Load and validate supplier ticket profiles
    profiles = load_ticket_profiles()
    logging.info(f"Loaded {len(profiles)} ticket profile(s).")

    # Load ticket type definitions for sorting
    ticket_types = load_ticket_types()
    if _TICKET_TYPE:
        logging.info(f"Ticket type filter active: processing {_TICKET_TYPE!r} tickets only.")

    # Authenticate with Microsoft Graph API
    client = GraphClient()

    # ── Download primary job Excel trackers from SharePoint ──────────────────
    # Download the configured job's trackers upfront so _load_toc_materials()
    # has a file to read, and to fail fast if the primary job files are missing.
    # Other job trackers are downloaded lazily on first encounter during processing.
    _need_rock     = (not _TICKET_TYPE or _TICKET_TYPE == "rock")
    _need_concrete = (not _TICKET_TYPE or _TICKET_TYPE == "concrete")

    rock_ok     = _download_excel_file(client, _EXCEL_FILENAME,          EXCEL_FILE)
    concrete_ok = _download_excel_file(client, _EXCEL_FILENAME_CONCRETE, EXCEL_FILE_CONCRETE)

    # Mark primary-job files in the lazy-download cache so they are not
    # re-downloaded when first encountered during ticket processing.
    _excel_downloaded[(_SP_JOB_FOLDER, False)] = rock_ok
    _excel_downloaded[(_SP_JOB_FOLDER, True)]  = concrete_ok

    if _need_rock and not rock_ok:
        logging.error(
            f"Cannot download {_EXCEL_FILENAME} from SharePoint — "
            "aborting because rock tickets are in scope for this run."
        )
        return
    if _need_concrete and not concrete_ok:
        logging.error(
            f"Cannot download {_EXCEL_FILENAME_CONCRETE} from SharePoint — "
            "aborting because concrete tickets are in scope for this run."
        )
        return

    # Load anticipated materials from the rock TOC tab (reads the downloaded file)
    toc_materials = _load_toc_materials()
    logging.info(f"Loaded {len(toc_materials)} TOC anticipated material entry/entries.")

    # ── Process emails ────────────────────────────────────────────────────────
    emails = get_unread_emails_with_pdf(client)

    successes, failures  = 0, 0
    sorted_count         = 0
    review_count         = 0
    total_duplicates     = 0
    rock_tickets:        list[str]                     = []
    flagged_items:       list[tuple[str, list[str]]]   = []
    error_items:         list[dict]                    = []

    if not emails:
        logging.info("No unread emails with PDF attachments. Nothing to do.")
    else:
        for email in emails:
            result = process_email(client, email, profiles, toc_materials, ticket_types)
            if result is None or result.get("failed"):
                failures += 1
                error_items.append({
                    "subject":      result["subject"]           if result else email.get("subject", "(unknown)"),
                    "att_name":     result.get("att_name", "")  if result else "",
                    "error_step":   result.get("error_step",   "unknown") if result else "unknown",
                    "error_detail": result.get("error_detail", "")        if result else "",
                    "traceback":    result.get("traceback",    "")        if result else "",
                })
            elif result.get("sorted_type"):
                # Email was classified as a different ticket type and moved to
                # its sort folder — count separately, don't flag as failure.
                sorted_count += 1
            else:
                successes += 1
                if result["tickets"]:
                    # Only accumulate rock tickets for batch outlier detection;
                    # concrete ticket numbers are 7-digit (different range) and
                    # cannot be mixed with 5-digit rock numbers.
                    if result.get("detected_type") != "concrete":
                        rock_tickets.extend(result["tickets"])
                if result["review_needed"]:
                    review_count += 1
                    flagged_items.append(
                        (result["subject"], result["review_reasons"])
                    )
                total_duplicates += result["duplicate_count"]

        # Batch outlier detection — runs after all rock tickets are written
        if rock_tickets:
            logging.info(
                f"Running batch outlier detection on "
                f"{len(rock_tickets)} rock ticket(s): {rock_tickets}"
            )
            _check_batch_outliers(rock_tickets)

    # ── Upload all Excel trackers back to SharePoint ─────────────────────────
    # Discover every ticket_tracker_*.xlsx in the temp directory — this covers
    # the primary job and any additional jobs downloaded lazily during processing.
    for _local_path in sorted(_LOCAL_TEMP_DIR.glob("ticket_tracker_*.xlsx")):
        _sp_name = _local_path.name
        # Extract job number from filename:
        #   ticket_tracker_2601.xlsx          → job 2601, rock
        #   ticket_tracker_2601_concrete.xlsx → job 2601, concrete
        _m = re.match(r"ticket_tracker_(\w+?)(?:_concrete)?\.xlsx$", _sp_name)
        if not _m:
            continue
        _job_num   = _m.group(1)
        _sp_folder = _resolve_sp_job_folder(client, _job_num)
        try:
            _upload_excel_file(client, _sp_name, str(_local_path), job_folder=_sp_folder)
            _local_path.unlink(missing_ok=True)
            logging.info(f"Deleted local temp copy: {_local_path}")
        except Exception as exc:
            logging.critical(
                f"CRITICAL: Failed to upload {_sp_name} to SharePoint: {exc}. "
                f"Local copy saved at {_local_path}. Upload manually."
            )

    # ── Send run summary email ────────────────────────────────────────────────
    send_run_summary_email(
        client,
        processed_count  = len(rock_tickets),
        review_count     = review_count,
        duplicate_count  = total_duplicates,
        error_count      = failures,
        flagged_items    = flagged_items,
        error_items      = error_items,
    )

    logging.info("=" * 55)
    logging.info(
        f"Complete — success: {successes}  sorted: {sorted_count}  failed: {failures}  "
        f"(failed emails remain unread for retry)"
    )
    logging.info("=" * 55)


if __name__ == "__main__":
    main()
