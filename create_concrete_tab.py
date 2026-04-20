"""
create_concrete_tab.py
======================
One-time utility to create a new cost-code tab in the concrete ticket tracker.

The tab is named after the TOC row number where the QR code lives, so that
ticket_processor.py can find it via the TOC lookup (column B → row number →
tab name).  Run update_concrete_toc.py first to add the QR code row to the
TOC before running this script.

QR code     : 2601-0180-313713-99
Source file : ticket_tracker_2601_concrete.xlsx
SharePoint  : /Shared Documents/MJHughes OPEN JOBS/2601/

Azure App Registration permissions required (Application, not Delegated):
    - Files.ReadWrite.All
    - Sites.ReadWrite.All

Usage:
    python create_concrete_tab.py
"""

import os
import sys
import logging
from pathlib import Path

import msal
import requests
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# ── Configuration ──────────────────────────────────────────────────────────────
_env_path = Path(__file__).resolve().parent / ".env"
load_dotenv(dotenv_path=_env_path)

AZURE_CLIENT_ID     = os.getenv("AZURE_CLIENT_ID")
AZURE_CLIENT_SECRET = os.getenv("AZURE_CLIENT_SECRET")
AZURE_TENANT_ID     = os.getenv("AZURE_TENANT_ID")

SHAREPOINT_HOST = "vancouvermjhughes.sharepoint.com"
FILENAME        = "ticket_tracker_2601_concrete.xlsx"
SP_PATH         = f"MJHughes OPEN JOBS/2601/{FILENAME}"
LOCAL_TEMP      = Path(r"C:\Users\dawson.h\AppData\Local\Temp") / FILENAME

# The QR code that must already exist in the TOC (column B) before this script runs.
QR_CODE  = "2601-0180-313713-99"
TOC_TAB  = "TOC"

GRAPH_BASE   = "https://graph.microsoft.com/v1.0"
GRAPH_SCOPES = ["https://graph.microsoft.com/.default"]

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler()],
)


# ── Authentication ─────────────────────────────────────────────────────────────
def _get_token() -> str:
    authority = f"https://login.microsoftonline.com/{AZURE_TENANT_ID}"
    app = msal.ConfidentialClientApplication(
        AZURE_CLIENT_ID,
        authority=authority,
        client_credential=AZURE_CLIENT_SECRET,
    )
    result = app.acquire_token_for_client(scopes=GRAPH_SCOPES)
    if "access_token" not in result:
        raise RuntimeError(
            "Authentication failed: "
            + result.get("error_description", result.get("error", "unknown"))
        )
    return result["access_token"]


def _auth_header(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


# ── SharePoint helpers ─────────────────────────────────────────────────────────
def _get_site_id(token: str) -> str:
    resp = requests.get(
        f"{GRAPH_BASE}/sites/{SHAREPOINT_HOST}:/",
        headers=_auth_header(token),
    )
    resp.raise_for_status()
    return resp.json()["id"]


def _get_drive_id(token: str, site_id: str) -> str:
    resp = requests.get(
        f"{GRAPH_BASE}/sites/{site_id}/drives",
        headers=_auth_header(token),
    )
    resp.raise_for_status()
    drives = resp.json().get("value", [])
    for drive in drives:
        if drive.get("name", "").lower() in ("shared documents", "documents"):
            return drive["id"]
    if not drives:
        raise RuntimeError("No document libraries found on the SharePoint site.")
    logging.warning(
        f"'Shared Documents' drive not found; using '{drives[0]['name']}' as fallback."
    )
    return drives[0]["id"]


def _download(token: str, site_id: str, drive_id: str) -> None:
    url = (
        f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}"
        f"/root:/{SP_PATH}:/content"
    )
    resp = requests.get(url, headers=_auth_header(token))
    resp.raise_for_status()
    LOCAL_TEMP.parent.mkdir(parents=True, exist_ok=True)
    LOCAL_TEMP.write_bytes(resp.content)
    logging.info(
        f"Downloaded {FILENAME} from SharePoint "
        f"({len(resp.content):,} bytes → {LOCAL_TEMP})"
    )


def _upload(token: str, site_id: str, drive_id: str) -> None:
    data = LOCAL_TEMP.read_bytes()
    url = (
        f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}"
        f"/root:/{SP_PATH}:/content"
    )
    resp = requests.put(
        url,
        headers={
            **_auth_header(token),
            "Content-Type": (
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            ),
        },
        data=data,
    )
    resp.raise_for_status()
    logging.info(
        f"Uploaded {FILENAME} to SharePoint ({len(data):,} bytes)"
    )


# ── TOC lookup ─────────────────────────────────────────────────────────────────
def _toc_row_for_qr(wb) -> int:
    """Return the TOC row number where QR_CODE appears in column B.

    Raises RuntimeError when the TOC tab is missing or the QR code is not
    found — the tab cannot be named without this information.
    """
    if TOC_TAB not in wb.sheetnames:
        raise RuntimeError(
            f"TOC tab '{TOC_TAB}' not found in {FILENAME}. "
            f"Available tabs: {wb.sheetnames}"
        )

    toc      = wb[TOC_TAB]
    qr_upper = QR_CODE.strip().upper()

    for row_cells in toc.iter_rows(min_col=2, max_col=2):
        cell = row_cells[0]
        if cell.value and str(cell.value).strip().upper() == qr_upper:
            return cell.row

    raise RuntimeError(
        f"QR code '{QR_CODE}' not found in column B of the TOC tab. "
        "Run update_concrete_toc.py first to add it, then re-run this script."
    )


# ── Tab creation ───────────────────────────────────────────────────────────────
def _thin_border() -> Border:
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(
        start_color=hex_color,
        end_color=hex_color,
        fill_type="solid",
    )


def _font(bold: bool = False, size: int = 11) -> Font:
    return Font(name="Calibri", size=size, bold=bold)


def _align(horizontal: str = "left", vertical: str = "center") -> Alignment:
    return Alignment(
        horizontal=horizontal,
        vertical=vertical,
        wrap_text=False,
    )


def create_tab() -> None:
    """Open the local workbook, look up the TOC row, create the tab, and save."""
    wb = openpyxl.load_workbook(LOCAL_TEMP)

    toc_row      = _toc_row_for_qr(wb)
    tab_name     = str(toc_row)

    logging.info(
        f"QR code '{QR_CODE}' found in TOC row {toc_row} → tab will be named '{tab_name}'"
    )

    if tab_name in wb.sheetnames:
        logging.warning(
            f"Tab '{tab_name}' already exists — skipping creation."
        )
        return

    ws = wb.create_sheet(title=tab_name)

    GREY   = "D3D3D3"
    ORANGE = "FFC000"
    border = _thin_border()

    # ── Row 1: Title ────────────────────────────────────────────────────────
    ws.merge_cells("A1:H1")
    cell            = ws["A1"]
    cell.value      = "Concrete Ticket Tracking Form"
    cell.font       = _font(bold=True, size=14)
    cell.alignment  = _align(horizontal="center")
    cell.border     = border

    # ── Row 2: Job Name / Job Number ────────────────────────────────────────
    ws["A2"].value     = "Job Name:"
    ws["A2"].font      = _font(bold=True)
    ws["A2"].fill      = _fill(GREY)
    ws["A2"].alignment = _align()
    ws["A2"].border    = border

    ws.merge_cells("B2:F2")
    for col in range(2, 7):          # B–F: apply border to all merged cells
        ws.cell(row=2, column=col).border = border

    ws["G2"].value     = "Job Number:"
    ws["G2"].font      = _font(bold=True)
    ws["G2"].fill      = _fill(GREY)
    ws["G2"].alignment = _align()
    ws["G2"].border    = border

    ws["H2"].value     = 2601
    ws["H2"].font      = _font()
    ws["H2"].alignment = _align()
    ws["H2"].border    = border

    # ── Row 3: Cost Code / QR Description ───────────────────────────────────
    ws["A3"].value     = "Cost Code #:"
    ws["A3"].font      = _font(bold=True)
    ws["A3"].fill      = _fill(GREY)
    ws["A3"].alignment = _align()
    ws["A3"].border    = border

    ws.merge_cells("B3:D3")
    ws["B3"].value     = "0180-313713-99"
    ws["B3"].font      = _font()
    ws["B3"].alignment = _align()
    for col in range(2, 5):          # B–D
        ws.cell(row=3, column=col).border = border

    ws["E3"].value     = "QR Description:"
    ws["E3"].font      = _font(bold=True)
    ws["E3"].fill      = _fill(GREY)
    ws["E3"].alignment = _align()
    ws["E3"].border    = border

    ws.merge_cells("F3:H3")
    ws["F3"].value     = "BUY Class 2 Rip Rap by TON"
    ws["F3"].font      = _font()
    ws["F3"].alignment = _align()
    for col in range(6, 9):          # F–H
        ws.cell(row=3, column=col).border = border

    # ── Row 4: Column headers ────────────────────────────────────────────────
    headers = [
        ("A4", "Ticket Date",   ORANGE),
        ("B4", "Logged Date",   ORANGE),
        ("C4", "Supplier",      ORANGE),
        ("D4", "Slump",         ORANGE),
        ("E4", "Qty Delivered", ORANGE),
        ("F4", "Ticket #",      ORANGE),
        ("G4", "Helper Notes",  ORANGE),
        ("H4", "Human Notes",   GREY),
    ]
    for addr, label, color in headers:
        cell            = ws[addr]
        cell.value      = label
        cell.font       = _font(bold=True)
        cell.fill       = _fill(color)
        cell.alignment  = _align(horizontal="center")
        cell.border     = border

    # ── Row heights ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[4].height = 15

    # ── Column widths ────────────────────────────────────────────────────────
    col_widths = {1: 13, 2: 13, 3: 22, 4: 10, 5: 16, 6: 11, 7: 20, 8: 20}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(LOCAL_TEMP)
    logging.info(f"Created tab: '{tab_name}' (QR code: {QR_CODE})")


# ── Entry point ────────────────────────────────────────────────────────────────
def main() -> None:
    missing = [
        v for v in ("AZURE_CLIENT_ID", "AZURE_CLIENT_SECRET", "AZURE_TENANT_ID")
        if not os.getenv(v)
    ]
    if missing:
        logging.error(f"Missing required .env variables: {', '.join(missing)}")
        sys.exit(1)

    logging.info("Authenticating with Microsoft Graph API...")
    token    = _get_token()
    site_id  = _get_site_id(token)
    drive_id = _get_drive_id(token, site_id)
    logging.info("Authenticated.")

    _download(token, site_id, drive_id)
    create_tab()
    _upload(token, site_id, drive_id)
    logging.info(f"Uploaded {FILENAME} to SharePoint")

    LOCAL_TEMP.unlink()
    logging.info(f"Deleted local temp copy: {LOCAL_TEMP}")


if __name__ == "__main__":
    main()
