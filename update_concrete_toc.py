"""
update_concrete_toc.py
======================
One-time utility to add a new row to the TOC tab of the concrete ticket
tracker workbook on SharePoint.

Steps:
    1. Download ticket_tracker_2601_concrete.xlsx from SharePoint to temp.
    2. Open the TOC tab and append a new row:
           A  JOB              : 2601
           B  QR Code          : 2601-0180-313713-99
           C  Long Description : BUY Class 2 Rip Rap by TON
           D  QR Short Desc.   : Class 2 RipRap
           E  Owner Bid Item   : (blank)
    3. Apply Calibri 11 font and the same thin border as the header row.
    4. Upload the modified file back to SharePoint (overwrites existing).
    5. Delete the local temp copy.

SharePoint path:
    /Shared Documents/MJHughes OPEN JOBS/2601/ticket_tracker_2601_concrete.xlsx

Azure App Registration permissions required (Application, not Delegated):
    - Files.ReadWrite.All
    - Sites.ReadWrite.All

Usage:
    python update_concrete_toc.py
"""

import os
import sys
import logging
from copy import copy
from pathlib import Path

import msal
import requests
import openpyxl
from openpyxl.styles import Font, Border, Side
from dotenv import load_dotenv

# ── Configuration ──────────────────────────────────────────────────────────────
_env_path = Path(__file__).resolve().parent / ".env"
load_dotenv(dotenv_path=_env_path)

AZURE_CLIENT_ID     = os.getenv("AZURE_CLIENT_ID")
AZURE_CLIENT_SECRET = os.getenv("AZURE_CLIENT_SECRET")
AZURE_TENANT_ID     = os.getenv("AZURE_TENANT_ID")

SHAREPOINT_HOST = "vancouvermjhughes.sharepoint.com"
FILENAME        = "ticket_tracker_2601_concrete.xlsx"
SP_PATH         = f"MJHughes OPEN JOBS/2601/{FILENAME}"
LOCAL_TEMP      = Path(r"C:\Users\dawson.h\AppData\Local\Temp") / FILENAME

TOC_TAB         = "TOC"

# New row values: (col_index, value)  — col E intentionally left blank
NEW_ROW = {
    1: "2601",
    2: "2601-0180-313713-99",
    3: "BUY Class 2 Rip Rap by TON",
    4: "Class 2 RipRap",
    5: "",
}

GRAPH_BASE   = "https://graph.microsoft.com/v1.0"
GRAPH_SCOPES = ["https://graph.microsoft.com/.default"]

# ── Logging ────────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler()],
)


# ── Authentication ─────────────────────────────────────────────────────────────
def _get_token() -> str:
    authority = f"https://login.microsoftonline.com/{AZURE_TENANT_ID}"
    app = msal.ConfidentialClientApplication(
        AZURE_CLIENT_ID,
        authority=authority,
        client_credential=AZURE_CLIENT_SECRET,
    )
    result = app.acquire_token_for_client(scopes=GRAPH_SCOPES)
    if "access_token" not in result:
        raise RuntimeError(
            "Authentication failed: "
            + result.get("error_description", result.get("error", "unknown"))
        )
    return result["access_token"]


def _auth_header(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


# ── SharePoint helpers ─────────────────────────────────────────────────────────
def _get_site_id(token: str) -> str:
    resp = requests.get(
        f"{GRAPH_BASE}/sites/{SHAREPOINT_HOST}:/",
        headers=_auth_header(token),
    )
    resp.raise_for_status()
    return resp.json()["id"]


def _get_drive_id(token: str, site_id: str) -> str:
    resp = requests.get(
        f"{GRAPH_BASE}/sites/{site_id}/drives",
        headers=_auth_header(token),
    )
    resp.raise_for_status()
    drives = resp.json().get("value", [])
    for drive in drives:
        if drive.get("name", "").lower() in ("shared documents", "documents"):
            return drive["id"]
    if not drives:
        raise RuntimeError("No document libraries found on the SharePoint site.")
    logging.warning(
        f"'Shared Documents' drive not found; using '{drives[0]['name']}' as fallback."
    )
    return drives[0]["id"]


def _download(token: str, site_id: str, drive_id: str) -> None:
    url = (
        f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}"
        f"/root:/{SP_PATH}:/content"
    )
    resp = requests.get(url, headers=_auth_header(token))
    resp.raise_for_status()
    LOCAL_TEMP.parent.mkdir(parents=True, exist_ok=True)
    LOCAL_TEMP.write_bytes(resp.content)
    logging.info(
        f"Downloaded {FILENAME} from SharePoint "
        f"({len(resp.content):,} bytes → {LOCAL_TEMP})"
    )


def _upload(token: str, site_id: str, drive_id: str) -> None:
    url = (
        f"{GRAPH_BASE}/sites/{site_id}/drives/{drive_id}"
        f"/root:/{SP_PATH}:/content"
    )
    data = LOCAL_TEMP.read_bytes()
    resp = requests.put(
        url,
        headers={
            **_auth_header(token),
            "Content-Type": (
                "application/vnd.openxmlformats-officedocument"
                ".spreadsheetml.sheet"
            ),
        },
        data=data,
    )
    resp.raise_for_status()
    logging.info(f"Uploaded {FILENAME} to SharePoint ({len(data):,} bytes)")


# ── TOC edit ──────────────────────────────────────────────────────────────────
def _thin_border_from_row(sheet, row_idx: int, num_cols: int) -> Border:
    """Read the border from the first non-empty cell in *row_idx* and return
    a Border object with thin sides on all four edges.  Falls back to a
    plain thin border when the row has no border set.
    """
    for col in range(1, num_cols + 1):
        cell = sheet.cell(row=row_idx, column=col)
        b = cell.border
        if b and any([b.left, b.right, b.top, b.bottom]):
            # Re-create border from the live values so it is detached from
            # the source cell and safe to apply to new cells.
            def _side(s):
                return Side(style=s.border_style, color=s.color) if s and s.border_style else Side(style="thin")
            return Border(
                left=_side(b.left),
                right=_side(b.right),
                top=_side(b.top),
                bottom=_side(b.bottom),
            )
    # Fallback: plain thin border on all sides
    thin = Side(style="thin")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def add_toc_row() -> None:
    """Open the local workbook, append the new row to the TOC tab, and save."""
    wb = openpyxl.load_workbook(LOCAL_TEMP)

    if TOC_TAB not in wb.sheetnames:
        raise ValueError(
            f"Tab '{TOC_TAB}' not found in {FILENAME}. "
            f"Available tabs: {wb.sheetnames}"
        )

    ws = wb[TOC_TAB]

    # Find the first empty row (check column A).
    next_row = 1
    for row_idx in range(1, ws.max_row + 2):
        if ws.cell(row=row_idx, column=1).value is None:
            next_row = row_idx
            break

    num_cols = max(len(NEW_ROW), 5)

    # Copy border style from the header row (row 1).
    border = _thin_border_from_row(ws, 1, num_cols)
    font   = Font(name="Calibri", size=11)

    for col_idx, value in NEW_ROW.items():
        cell        = ws.cell(row=next_row, column=col_idx, value=value if value else None)
        cell.font   = font
        cell.border = border

    wb.save(LOCAL_TEMP)
    logging.info(
        f"Added TOC row for QR code {NEW_ROW[2]} at row {next_row} "
        f"(tab '{TOC_TAB}', file saved locally)"
    )


# ── Entry point ────────────────────────────────────────────────────────────────
def main() -> None:
    missing = [
        v for v in ("AZURE_CLIENT_ID", "AZURE_CLIENT_SECRET", "AZURE_TENANT_ID")
        if not os.getenv(v)
    ]
    if missing:
        logging.error(f"Missing required .env variables: {', '.join(missing)}")
        sys.exit(1)

    logging.info("Authenticating with Microsoft Graph API...")
    token    = _get_token()
    site_id  = _get_site_id(token)
    drive_id = _get_drive_id(token, site_id)
    logging.info("Authenticated.")

    # 1. Download
    _download(token, site_id, drive_id)

    # 2. Edit TOC tab
    add_toc_row()

    # 3. Upload
    _upload(token, site_id, drive_id)
    logging.info(f"Uploaded {FILENAME} to SharePoint")

    # 4. Delete temp copy
    LOCAL_TEMP.unlink()
    logging.info(f"Deleted local temp copy: {LOCAL_TEMP}")

    logging.info(f"Added TOC row for QR code {NEW_ROW[2]}")


if __name__ == "__main__":
    main()
